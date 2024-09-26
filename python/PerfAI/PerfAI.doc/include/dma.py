#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
# @Time    : 2023/7/18 10:46
# @Author  : chongqing.zeng@sophgo.com
# @Project : PerfAI
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from definition.bm1684x_defs import dma_func_type_dict, data_type_dict
from definition.style import *
from utils.utils import *


class DmaNode:
    def __init__(self, reg):
        self.datasize = int(reg['DMA data size(B)'])
        self.cycle = int(reg['Asic Cycle'])
        self.direction = reg['Direction']


class Dma(object):
    def __init__(self, core_id, writer):
        """
        Initial DMA object, equals to a DMA sheet in Excel.
        :param core_id: the id of current core
        :param writer: the writer of Excel to write
        """
        self.columns = ['Engine Id', 'Core Id', 'Cmd Id', 'Layer Id', 'Layer Name',
                        'Function Type', 'Function Name', 'DMA data size(B)', 'Start Cycle', 'End Cycle',
                        'Asic Cycle', 'Stall Cycle', 'DDR Bandwidth(GB/s)', 'L2M Bandwidth(GB/s)', 'Direction', 'AvgBurstLength', 'Data Type', 'Non32ByteRatio',
                        'MaskWriteRatio', 'cmd_id_dep', 'cmd_special_function', 'src_start_addr', 'dst_start_addr',
                        'src_nsize', 'src_csize', 'src_hsize', 'src_wsize',
                        'dst_nsize', 'dst_csize', 'dst_hsize', 'dst_wsize',
                        'src_nstride', 'src_cstride', 'src_hstride', 'src_wstride',
                        'dst_nstride', 'dst_cstride', 'dst_hstride', 'dst_wstride',
                        'nchw_copy', 'stride_enable', 'src_data_format', 'cmd_type',
                        'index_csize', 'index_hsize', 'index_cstride', 'index_hstride',
                        'mask_start_addr_h8', 'mask_start_addr_l32', 'mask_data_format', 'localmem_mask_h32',
                        'localmem_mask_l32',
                        'fill_constant_en', 'constant_value', 'index', 'cmd_short', 'intr_en', 'Msg Id', 'Sd\Wt Count']
        self.reg_list = []
        self.core_id = str(core_id)
        self.height = None
        self.width = len(self.columns)
        self.dma_cycle = 0
        self.working_cycle = 0
        self.stall_cycle = 0
        self.stall_cycle_ratio = 0
        self.ddr_total_datasize = 0
        self.ddr_total_cycle = 0
        self.l2_total_datasize = 0
        self.l2_total_cycle = 0
        self.l2_avg_bandwidth = 0
        self.ddr_avg_bandwidth = 0
        self.ddr_burst_length_sum = 0
        self.ddr_xact_cnt = 0
        self.ddr_avg_burst_length = 0
        self.l2_burst_length_sum = 0
        self.l2_xact_cnt = 0
        self.l2_avg_burst_length = 0
        self.wait_msg_total_time = 0
        self.perf_dict = {}
        self.chip_arch_dict = None
        self.sheet_name = None
        self.sheet_color = None
        self.writer = writer
        self.start_time = sys.maxsize
        self.end_time = 0
        self.dma_time = 0

    def load(self, reg_info_file, dma_layer_map):
        """
        Load data from external file.
        :param reg_info_file: file records register information, usually obtained by TPUPerf
        :return: None
        """
        if os.path.exists(reg_info_file) and os.path.getsize(reg_info_file) != 0:
            last_underscore_index = reg_info_file.rfind('_')
            core_id = int(reg_info_file[last_underscore_index + 1 : -4])
            with open(reg_info_file) as f:
                rows = f.readlines()
                field_set = set()
                reg_count = 0
                for row in rows:
                    if "__TDMA_REG_INFO__" in row:
                        reg_count += 1
                    if "\t" in row and reg_count > 0:
                        attr = row.split(': ')[0][1:]
                        field_set.add(attr)
                reg_count = 0
                field_list = list(field_set) if len(field_set) >= len(self.columns) else self.columns
                reg_dict = dict.fromkeys(field_list, '')
                idx = 0
                for row in rows:
                    if "__CHIP_ARCH_ARGS__" in row:
                        self.chip_arch_dict = dict()
                    elif "__TDMA_REG_INFO__" in row:
                        reg_count += 1
                        if idx != 0:
                            k = int(reg_dict['Cmd Id'])
                            if any(key[0] == k for key in dma_layer_map.keys()):
                                layer_id_name = dma_layer_map[(k, core_id)]
                            else:
                                layer_id_name = ['-', '-']
                            reg_dict['Layer Id'] = layer_id_name[0]
                            reg_dict['Layer Name'] = layer_id_name[1]
                            self.reg_list.append(reg_dict)
                        reg_dict = dict.fromkeys(field_list, '')
                    elif reg_count == 0:
                        fields = row.split(': ')
                        attr = fields[0][1:]
                        val = fields[1][:-1]
                        self.chip_arch_dict[attr] = val
                        idx = 0
                    else:
                        fields = row.split(': ')
                        attr = fields[0][1:]
                        val = fields[1][:-1]
                        if val.isnumeric() and 'burst' not in attr.lower() and 'width' not in attr.lower():
                            val = int(val)
                        elif 'bandwidth' in attr.lower():
                            val = float(val)
                        reg_dict[attr] = val
                        idx += 1
                if idx != 0:
                    k = int(reg_dict['Cmd Id'])
                    if any(key[0] == k for key in dma_layer_map.keys()):
                        layer_id_name = dma_layer_map[(k, core_id)]
                    else:
                        layer_id_name = ['-', '-']
                    reg_dict['Layer Id'] = layer_id_name[0]
                    reg_dict['Layer Name'] = layer_id_name[1]
                self.reg_list.append(reg_dict)
        self.height = len(self.reg_list)

    def add_kpi_field(self):
        """
        Add some indicators which are convenient for performance analysis artificially.
        :return: None
        """
        for i in range(len(self.reg_list)):
            reg_dict = self.reg_list[i]
            name_key = (int(reg_dict['cmd_type']))
            if reg_dict['cmd_type'] == 6:
                reg_dict['Data Type'] = 'None'
                # dma_sys do not transfer data
                reg_dict['Direction'] = '-'
            self.dma_cycle += int(reg_dict['Asic Cycle'])
            self.stall_cycle += int(reg_dict['Stall Cycle'])
            if 'DDR' in reg_dict['Direction'] and isinstance(reg_dict['DMA data size(B)'], int):
                self.ddr_total_datasize += reg_dict['DMA data size(B)']
                self.ddr_total_cycle += reg_dict['Asic Cycle']
                self.ddr_burst_length_sum += reg_dict['gmem_bl_sum']
                self.ddr_xact_cnt += reg_dict['gmem_xact_cnt']
            elif 'L2' in reg_dict['Direction'] and isinstance(reg_dict['DMA data size(B)'], int):
                self.l2_total_datasize += reg_dict['DMA data size(B)']
                self.l2_total_cycle += reg_dict['Asic Cycle']
            if reg_dict['cmd_type'] == 6 and reg_dict['cmd_special_function'] == 4:
                self.wait_msg_total_time += reg_dict['Asic Cycle']
            if reg_dict['gmem_xact_cnt'] > 0:
                reg_dict['AvgBurstLength'] = get_ratio_float_2f(reg_dict['gmem_bl_sum'], reg_dict['gmem_xact_cnt'])
                reg_dict['Non32ByteRatio'] = get_ratio_float_2f(reg_dict['gmem_n32Ba_sa_cnt'], reg_dict['gmem_xact_cnt'])
                reg_dict['MaskWriteRatio'] = get_ratio_float_2f(reg_dict['gmem_msk_wr_cnt'], reg_dict['gmem_xact_cnt'])
            else:
                reg_dict['AvgBurstLength'] = 0
                reg_dict['Non32ByteRatio'] = 0
                reg_dict['MaskWriteRatio'] = 0
            self.start_time = min(self.start_time, get_time_by_cycle(reg_dict['Start Cycle'], self.chip_arch_dict['DMA Frequency(MHz)']))
            self.end_time = max(self.start_time, get_time_by_cycle(reg_dict['End Cycle'], self.chip_arch_dict['DMA Frequency(MHz)']))
        self.dma_time = get_time_by_cycle(self.dma_cycle, self.chip_arch_dict['DMA Frequency(MHz)']) if self.chip_arch_dict else 0
        self.working_cycle = self.dma_cycle - self.wait_msg_total_time
        self.ddr_avg_bandwidth = get_ratio_float_2f(self.ddr_total_datasize,
                                                    get_time_by_cycle(self.ddr_total_cycle, self.chip_arch_dict['DMA Frequency(MHz)'])) \
                                                    if self.chip_arch_dict else 0
        self.l2_avg_bandwidth = get_ratio_float_2f(self.l2_total_datasize,
                                                    get_time_by_cycle(self.l2_total_cycle, self.chip_arch_dict['DMA Frequency(MHz)'])) \
                                                    if self.chip_arch_dict else 0
        self.ddr_avg_burst_length = get_ratio_float_2f(self.ddr_burst_length_sum, self.ddr_xact_cnt)
        self.perf_dict = {
            'totalDmaCycle': [self.working_cycle],
            'workingCycle': [self.working_cycle],
            'totalStallCycle': [self.stall_cycle],
            'stallCycleRatio': [get_ratio_str_2f_zero(self.stall_cycle, self.dma_cycle)],
            'totalDdrDataSize(B)': [self.ddr_total_datasize],
            'totalL2DataSize(B)': [self.l2_total_datasize],
            'ddrAvgBandwidth(GB/s)': [self.ddr_avg_bandwidth],
            'l2AvgBandwidth(GB/s)': [self.l2_avg_bandwidth],
            'avgDdrBurstLength': [self.ddr_avg_burst_length],
            'waitMsgTotalTime': [self.wait_msg_total_time]
        }

    def pop_data(self):
        gdma_instance_map = dict()
        for reg in self.reg_list:
            gdma_instance_map[int(reg['Cmd Id'])] = DmaNode(reg)
        return gdma_instance_map

    def write(self):
        """
        Write register information and kpi field to Excel.
        :return: None
        """
        df = pd.DataFrame(self.reg_list)
        new_df = pd.DataFrame()
        if len(df) > 0:
            for column in self.columns:
                if column in df.columns:
                    new_df[column] = df[column]
                else:
                    new_df[column] = None
            pre_clos = df.columns
            df = new_df
            for col in self.columns:
                if ('addr' in col or 'mask' in col) and col in pre_clos:
                    df[col] = int2Hex(df[col].values)
            if self.chip_arch_dict['Platform'].lower() == 'simulator':
                df.rename(columns={'Asic Cycle': 'Simulator Cycle'}, inplace=True)
            pd.DataFrame(self.perf_dict).to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=0,
                                                  startcol=2,
                                                  engine='xlsxwriter', float_format='%g')
            df.to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=5, engine='xlsxwriter', float_format='%g')

    @classmethod
    def set_style(cls, file_path, core_id, engine_type, sheet_color, chip_arch, frozen=True):
        """
        Set style for Excel, and highlight some fields we care about.
        :param frozen: freeze some statistics above the sheet
        :param sheet_color: the color for sheet name in Excel bottem
        :param engine_type: the type of engine, like tiu, gdma, sdma etc.
        :param file_path: Excel output path
        :param core_id: the id of current core
        :return: None
        """
        wb = load_workbook(file_path)
        sheet_name = engine_type + '_' + str(core_id)
        if sheet_name not in wb.sheetnames:
            return
        df = pd.read_excel(file_path, sheet_name)
        ws = wb[sheet_name]
        ws.cell(1, 1).value = 'Performance'
        ws.cell(2, 1).value = 'Summary'
        for h, w in zip([1, 2], [1, 1]):
            ws.cell(h, w).fill = DetailsStyle.title_pattern
            ws.cell(h, w).font = DetailsStyle.title_font
        summary_start_cols = 3
        summary_end_cols = summary_start_cols + 10
        for w in range(summary_start_cols, summary_end_cols):
            # summary title style
            ws.cell(1, w).fill = DetailsStyle.title_header_pattern
            ws.cell(1, w).font = DetailsStyle.title_header_font
            # summary content style
            ws.cell(2, w).fill = DetailsStyle.title_content_pattern
            ws.cell(2, w).font = DetailsStyle.title_font
        for cell in ws[6:6]:
            # content title style
            cell.fill = DetailsStyle.content_pattern
            cell.font = DetailsStyle.title_header_font
        # set content style
        content_end_cols = 19
        for w in range(1, content_end_cols + 1):
            for h in range(7, len(df) + 2):
                # set key content border style
                ws.cell(h, w).fill = DetailsStyle.key_content_pattern
                ws.cell(h, w).border = DetailsStyle.border
                ws.cell(h, w).alignment = DetailsStyle.right_align
                ws.cell(h, w).font = DetailsStyle.title_font
        stall_cycle_pos = 12
        ddr_bandwidth_pos = 13
        l2m_bandwidth_pos = 14
        tsk_typ_pos = 42
        dma_cycle_pos = 11
        avg_burst_length_pos = 16
        direction_pos = 15
        data_type_pos = 17
        for h in range(5, len(df)):
            ddr_max_bd, l2_max_bd, ddr_max_bl, l2_max_bl = float(chip_arch['DDR Max BW(GB/s/Core)']),\
            float(chip_arch['L2 Max BW(GB/s)']), float(chip_arch['Bus Max Burst']), int(chip_arch['L2 Max Burst'])
            # bandwidth
            if int(ws.cell(h + 2, tsk_typ_pos).value) == 6 or ws.cell(h + 2, direction_pos).value in [None, "-"]:
                # sys do not have bandwidth
                continue
            else:
                if float(ws.cell(h + 2, ddr_bandwidth_pos).value) < (ddr_max_bd * 0.5):
                    ws.cell(h + 2, ddr_bandwidth_pos).fill = DetailsStyle.red_light
                    if int(ws.cell(h + 2, dma_cycle_pos).value) < 1000:
                        ws.cell(h + 2, dma_cycle_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h + 2, ddr_bandwidth_pos).value) < (ddr_max_bd * 0.75):
                    ws.cell(h + 2, ddr_bandwidth_pos).fill = DetailsStyle.yellow_light
                    if int(ws.cell(h + 2, dma_cycle_pos).value) < 1000:
                        ws.cell(h + 2, dma_cycle_pos).fill = DetailsStyle.red_light
                if float(ws.cell(h + 2, l2m_bandwidth_pos).value) < (l2_max_bd * 0.5):
                    ws.cell(h + 2, l2m_bandwidth_pos).fill = DetailsStyle.red_light
                    if int(ws.cell(h + 2, dma_cycle_pos).value) < 1000:
                        ws.cell(h + 2, dma_cycle_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h + 2, l2m_bandwidth_pos).value) < (l2_max_bd * 0.75):
                    ws.cell(h + 2, l2m_bandwidth_pos).fill = DetailsStyle.yellow_light
                    if int(ws.cell(h + 2, dma_cycle_pos).value) < 1000:
                        ws.cell(h + 2, dma_cycle_pos).fill = DetailsStyle.red_light
            # avg burst length
            if 'DDR' in ws.cell(h + 2, direction_pos).value:
                if float(ws.cell(h + 2, avg_burst_length_pos).value) < ddr_max_bl:
                    ws.cell(h + 2, avg_burst_length_pos).fill = DetailsStyle.yellow_light
            elif 'L2' in ws.cell(h + 2, direction_pos).value:
                if float(ws.cell(h + 2, avg_burst_length_pos).value) < l2_max_bl:
                    ws.cell(h + 2, avg_burst_length_pos).fill = DetailsStyle.yellow_light
            # data type
            if ws.cell(h + 2, data_type_pos).value == 'FP32':
                ws.cell(h + 2, data_type_pos).fill = DetailsStyle.red_light
            # stall cycle
            if int(ws.cell(h + 2, stall_cycle_pos).value) > 0:
                ws.cell(h + 2, stall_cycle_pos).fill = DetailsStyle.red_light
        h, w = 4, ddr_bandwidth_pos
        ws.cell(h, w).value = 'DDR_BW<75%'
        ws.cell(h, w).fill = DetailsStyle.yellow_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align
        h, w = 5, ddr_bandwidth_pos
        ws.cell(h, w).value = 'DDR_BW<50%'
        ws.cell(h, w).fill = DetailsStyle.red_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align
        h, w = 4, l2m_bandwidth_pos
        ws.cell(h, w).value = 'L2M_BW<75%'
        ws.cell(h, w).fill = DetailsStyle.yellow_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align
        h, w = 5, l2m_bandwidth_pos
        ws.cell(h, w).value = 'L2M_BW<50%'
        ws.cell(h, w).fill = DetailsStyle.red_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align
        h, w = 5, dma_cycle_pos
        ws.cell(h, w).value = '搬运量太少'
        ws.cell(h, w).fill = DetailsStyle.red_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align
        h, w = 5, avg_burst_length_pos
        ws.cell(h, w).value = '<' + str(ddr_max_bl) + '(DDR)|<' + str(l2_max_bl) + '(L2)'
        ws.cell(h, w).fill = DetailsStyle.yellow_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align
        h, w = 5, data_type_pos
        ws.cell(h, w).value = 'isFP32'
        ws.cell(h, w).fill = DetailsStyle.red_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align
        h, w = 5, stall_cycle_pos
        ws.cell(h, w).value = 'Engine Conflict'
        ws.cell(h, w).fill = DetailsStyle.red_light
        ws.cell(h, w).font = DetailsStyle.title_font
        ws.cell(h, w).alignment = DetailsStyle.center_align

        for w in range(summary_start_cols, summary_end_cols):
            # set header border style
            ws.cell(2, w).border = DetailsStyle.border
            ws.cell(2, w).alignment = DetailsStyle.center_align
        for col in df.columns:
            # auto set columns width
            index = list(df.columns).index(col)
            letter = get_column_letter(index + 1)
            collen = 0
            if index > content_end_cols + 2:
                collen = len(col)
            else:
                for h in range(1, len(df) + 6):
                    if h > 20:
                        break
                    if index == direction_pos - 1:
                        collen = max(collen, len(str(ws.cell(h, index).value)))
                    elif ws.cell(h, index + 1).value is not None:
                        collen = max(collen, len(str(ws.cell(h, index + 1).value)))
            ws.column_dimensions[letter].width = collen * 1.05
        ws.cell(5, 1).value = '*Stall Cycle indicates the waiting time when TIU and DMA attempting to access a bank simultaneously.'
        if frozen:
            _cell = ws.cell(7, 1)
            ws.freeze_panes = _cell
        ws.sheet_properties.tabColor = sheet_color
        wb.save(file_path)


class Gdma(Dma):
    def __init__(self, core_id, writer, sheet_name):
        """
        Inherited from the dma class, initialize a gdma object.
        :param core_id: the id of current core
        :param writer: the writer of Excel to write
        :param sheet_name: the name of gdma sheet
        """
        super().__init__(core_id, writer)
        self.sheet_name = sheet_name + '_' + str(core_id)

    def load(self, reg_info_file, gdma_layer_map):
        """
        Load gdma data from external file.
        :param gdma_layer_map:
        :param reg_info_file: file records DMA register information, usually obtained by TPUPerf
        :return: None
        """
        super().load(reg_info_file, gdma_layer_map)
        new_reg_list = []
        for reg_dict in self.reg_list:
            if reg_dict['Engine Id'] == 1:
                new_reg_list.append(reg_dict)
        self.reg_list = new_reg_list
        return self.chip_arch_dict

    @classmethod
    def set_style(cls, file_path, core_id, engine_type='GDMA', sheet_color='FFA500', chip_arch = None, frozen=True):
        """
        Set gdma sheet style for output Excel
        :param frozen: freeze some statistics above the sheet
        :param file_path: the Excel output file path
        :param core_id: the id of current core
        :param engine_type: the type of engine, default is 'GDMA'
        :param sheet_color: the type of sheet in Excel bottem, default is 'FFA500'
        :return:
        """
        super().set_style(file_path, core_id, engine_type, sheet_color, chip_arch, frozen=frozen)


class Sdma(Dma):
    def __init__(self, core_id, writer, sheet_name):
        """
        Inherited from the dma class, initialize a sdma object.
        :param core_id: the id of current core
        :param writer: the writer of Excel to write
        :param sheet_name: the name of sdma sheet
        """
        super().__init__(core_id, writer)
        self.sheet_name = sheet_name + '_' + str(core_id)

    def load(self, reg_info_file, sdma_layer_map):
        """
        Load data from external file.
        :param sdma_layer_map:
        :param reg_info_file: file records register information, usually obtained by TPUPerf
        :return: None
        """
        super().load(reg_info_file, sdma_layer_map)
        new_reg_list = []
        for reg_dict in self.reg_list:
            if reg_dict['Engine Id'] == 3:
                new_reg_list.append(reg_dict)
        self.reg_list = new_reg_list
        return self.chip_arch_dict

    @classmethod
    def set_style(cls, file_path, core_id, engine_type='SDMA', sheet_color='D0CECE', chip_arch=None, frozen=True):
        """
        Set gdma sheet style for output Excel
        :param frozen: freeze some statistics above the sheet
        :param file_path: the Excel output file path
        :param core_id: the id of current core
        :param engine_type: the type of engine, default is 'SDMA'
        :param sheet_color: the type of sheet in Excel bottem, default is 'D0CECE'
        :return:
        """
        super().set_style(file_path, core_id, engine_type, sheet_color, chip_arch, frozen=frozen)


class Cdma(Dma):
    def __init__(self, core_id, writer, sheet_name):
        """
        Inherited from the dma class, initialize a cdma object.
        :param core_id: the id of current core
        :param writer: the writer of Excel to write
        :param sheet_name: the name of cdma sheet
        """
        super().__init__(core_id, writer)
        self.sheet_name = sheet_name + '_' + str(core_id)

    def load(self, reg_info_file):
        """
        Load cdma data from external file.
        :param cdma_layer_map:
        :param reg_info_file: file records DMA register information, usually obtained by TPUPerf
        :return: None
        """
        if os.path.exists(reg_info_file) and os.path.getsize(reg_info_file) != 0:
            with open(reg_info_file) as f:
                rows = f.readlines()
                field_set = set()
                reg_count = 0
                for row in rows:
                    if "__CDMA_REG_INFO__" in row:
                        reg_count += 1
                    if "\t" in row and reg_count > 0:
                        attr = row.split(': ')[0][1:]
                        field_set.add(attr)
                reg_count = 0
                field_list = list(field_set) if len(field_set) >= len(self.columns) else self.columns
                reg_dict = dict.fromkeys(field_list, '')
                idx = 0
                for row in rows:
                    if "__CHIP_ARCH_ARGS__" in row:
                        self.chip_arch_dict = dict()
                    elif "__CDMA_REG_INFO__" in row:
                        reg_count += 1
                        if idx != 0:
                            layer_id_name = ['-', '-']
                            reg_dict['Layer Id'] = layer_id_name[0]
                            reg_dict['Layer Name'] = layer_id_name[1]
                            self.reg_list.append(reg_dict)
                        reg_dict = dict.fromkeys(field_list, '')
                    elif reg_count == 0:
                        fields = row.split(': ')
                        attr = fields[0][1:]
                        val = fields[1][:-1]
                        self.chip_arch_dict[attr] = val
                        idx = 0
                    else:
                        fields = row.split(': ')
                        attr = fields[0][1:]
                        val = fields[1][:-1]
                        if val.isnumeric() and 'burst' not in attr.lower() and 'width' not in attr.lower():
                            val = int(val)
                        reg_dict[attr] = val
                        idx += 1
                self.reg_list.append(reg_dict)
        self.height = len(self.reg_list)
        return self.chip_arch_dict

    @classmethod
    def set_style(cls, file_path, core_id, engine_type='CDMA', sheet_color='C0504D', chip_arch=None, frozen=True):
        """
        Set cdma sheet style for output Excel
        :param frozen: freeze some statistics above the sheet
        :param file_path: the Excel output file path
        :param core_id: the id of current core
        :param engine_type: the type of engine, default is 'CDMA'
        :param sheet_color: the type of sheet in Excel bottem, default is 'C0504D'
        :return:
        """
        super().set_style(file_path, core_id, engine_type, sheet_color, chip_arch, frozen=frozen)
