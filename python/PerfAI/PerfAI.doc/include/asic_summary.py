#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
# @Time    : 2023/7/18 10:50
# @Author  : chongqing.zeng@sophgo.com
# @Project : PerfAI
import pandas as pd
from numpy import transpose
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utils.utils import *
from definition.style import DetailsStyle


class AsicSummary(object):
    sheet_name = 'Engine Summary'

    def __init__(self, writer, tius, gdmas, sdmas, cdmas, act_core_num):
        """
        Initial a Summary object contains summary information of all engines such as tiu, gdma, sdma, etc.
        Equals to summary sheet in output Excel.
        :param writer: the writer of Excel to writer
        :param tius: tiu instance list
        :param gdmas: gdma instance list
        :param sdmas: sdma instance list, it may be empty
        :param act_core_num: represents the actual number of cores to run since the input file may be empty
        """
        self.columns = ['CoreId', 'TiuWorkingRatio', 'Parallelism', 'Concurrency',
                        'totalTime(us)', 'totalTiuCycle', 'totalAlgCycle', 'totalAlgOps', 'totalUArchOps', 'uArchURate',
                        'totalGdmaCycle', 'totalDdrDataSize', 'totalL2DataSize', 'ddrAvgBandwidth', 'l2AvgBandwidth', 'avgDdrBurstLength',
                        'totalSdmaCycle', 'totalDdrDataSize', 'ddrAvgBandwidth', 'avgDdrBurstLength',
                        'totalcdmaCycle', 'totalDdrDataSize', 'totalL2DataSize', 'ddrAvgBandwidth', 'l2AvgBandwidth', 'avgDdrBurstLength']
        self.writer = writer
        self.tius = tius
        self.gdmas = gdmas
        self.sdmas = sdmas
        self.cdmas = cdmas
        self.act_core_num = act_core_num
        self.sheet_color = ''
        self.data = []

    def load(self, chip_arch):
        """
        Compute all the values we care about
        :return: None
        """
        core_ids, tiu_work_ratios, prallelisms, concurrencys, total_times, tiu_cycles, alg_cycles, alg_opss, uArch_opss, uArch_rates = \
            [], [], [], [], [], [], [], [], [], []
        gdma_cycles, gdma_ddr_datasizes, gdma_l2_datasizes, gdma_ddr_avg_bds, \
            gdma_l2_avg_bds, gdma_ddr_avg_bls = [], [], [], [], [], []
        sdma_cycles, sdma_ddr_datasizes, sdma_ddr_avg_bds, sdma_ddr_avg_bls = [], [], [], []
        cdma_cycles, cdma_ddr_datasizes, cdma_l2_datasizes, cdma_ddr_avg_bds, \
            cdma_l2_avg_bds, cdma_ddr_avg_bls = [], [], [], [], [], []
        gdma_ddr_cycles, gdma_l2_cycles, sdma_ddr_cycles, sdma_l2_cycles, cdma_ddr_cycles, cdma_l2_cycles = 0, 0, 0, 0, 0, 0
        gdma_bl_sum, gdma_xact_sum, sdma_bl_sum, sdma_xact_sum, cdma_bl_sum, cdma_xact_sum = 0, 0, 0, 0, 0, 0
        tiu_frequency = int(chip_arch['TIU Frequency(MHz)'])
        dma_frequency = int(chip_arch['DMA Frequency(MHz)'])
        total_time = get_total_time(self.tius, self.gdmas, self.sdmas, self.cdmas)

        for core_id in range(0, self.act_core_num):
            core_ids.append(core_id)
            tiu_work_ratios.append(get_ratio_str_2f_zero(self.tius[core_id].tiu_time, total_time))
            if self.act_core_num > 1:
                prallelisms.append(get_ratio_str_2f_zero(self.tius[core_id].tiu_time + self.gdmas[core_id].working_cycle\
                                                    + self.sdmas[core_id].working_cycle, total_time))
            else:
                prallelisms.append(get_ratio_str_2f_zero(self.tius[core_id].tiu_time + self.gdmas[core_id].working_cycle, total_time))
            if self.tius[core_id].tiu_time > 0 and self.gdmas[core_id].working_cycle > 0:
                concurrencys.append(get_ratio_str_2f_zero(self.tius[core_id].tiu_time + self.gdmas[core_id].working_cycle - total_time,
                                                          min(self.tius[core_id].tiu_time, self.gdmas[core_id].working_cycle)))
            else:
                concurrencys.append('0.00%')
            total_times.append(total_time)
            tiu_cycles.append(self.tius[core_id].tiu_cycle)
            alg_cycles.append(self.tius[core_id].alg_total_cycle)
            alg_opss.append(self.tius[core_id].alg_total_ops)
            uArch_opss.append(self.tius[core_id].uArch_total_ops)
            uArch_rates.append(get_ratio_str_2f_zero(self.tius[core_id].alg_total_ops, self.tius[core_id].uArch_total_ops))

            gdma_cycles.append(self.gdmas[core_id].working_cycle)
            gdma_ddr_datasizes.append(self.gdmas[core_id].ddr_total_datasize)
            gdma_l2_datasizes.append(self.gdmas[core_id].l2_total_datasize)
            gdma_ddr_avg_bds.append(get_ratio_float_2f(self.gdmas[core_id].ddr_total_datasize,
                                                     get_time_by_cycle(self.gdmas[core_id].ddr_total_cycle, dma_frequency)))
            gdma_l2_avg_bds.append(get_ratio_float_2f(self.gdmas[core_id].l2_total_datasize,
                                                     get_time_by_cycle(self.gdmas[core_id].l2_total_cycle, dma_frequency)))
            gdma_ddr_avg_bls.append(get_ratio_float_2f(self.gdmas[core_id].ddr_burst_length_sum, self.gdmas[core_id].ddr_xact_cnt))
            gdma_ddr_cycles += self.gdmas[core_id].ddr_total_cycle
            gdma_l2_cycles += self.gdmas[core_id].l2_total_cycle
            gdma_bl_sum += self.gdmas[core_id].ddr_burst_length_sum
            gdma_xact_sum += self.gdmas[core_id].ddr_xact_cnt

            sdma_cycles.append(self.sdmas[core_id].working_cycle)
            sdma_ddr_datasizes.append(self.sdmas[core_id].ddr_total_datasize)
            sdma_ddr_avg_bds.append(get_ratio_float_2f(self.sdmas[core_id].ddr_total_datasize,
                                                     get_time_by_cycle(self.sdmas[core_id].ddr_total_cycle, dma_frequency)))
            sdma_ddr_avg_bls.append(get_ratio_float_2f(self.sdmas[core_id].ddr_burst_length_sum, self.sdmas[core_id].ddr_xact_cnt))
            sdma_ddr_cycles += self.sdmas[core_id].ddr_total_cycle
            sdma_l2_cycles += self.sdmas[core_id].l2_total_cycle
            sdma_bl_sum += self.sdmas[core_id].ddr_burst_length_sum
            sdma_xact_sum += self.sdmas[core_id].ddr_xact_cnt

            cdma_cycles.append(self.cdmas[core_id].working_cycle)
            cdma_ddr_datasizes.append(self.cdmas[core_id].ddr_total_datasize)
            cdma_l2_datasizes.append(self.cdmas[core_id].l2_total_datasize)
            cdma_ddr_avg_bds.append(get_ratio_float_2f(self.cdmas[core_id].ddr_total_datasize,
                                                     get_time_by_cycle(self.cdmas[core_id].ddr_total_cycle, dma_frequency)))
            cdma_l2_avg_bds.append(get_ratio_float_2f(self.cdmas[core_id].l2_total_datasize,
                                                     get_time_by_cycle(self.cdmas[core_id].l2_total_cycle, dma_frequency)))
            cdma_ddr_avg_bls.append(get_ratio_float_2f(self.cdmas[core_id].ddr_burst_length_sum, self.cdmas[core_id].ddr_xact_cnt))
            cdma_ddr_cycles += self.cdmas[core_id].ddr_total_cycle
            cdma_l2_cycles += self.cdmas[core_id].l2_total_cycle
            cdma_bl_sum += self.cdmas[core_id].ddr_burst_length_sum
            cdma_xact_sum += self.cdmas[core_id].ddr_xact_cnt
        core_ids.append('Overall')
        tiu_work_ratios.append(get_ratio_str_2f_zero(get_time_by_cycle(max(tiu_cycles), tiu_frequency),  total_time))
        if self.act_core_num > 1:
            prallelisms.append(get_ratio_str_2f_zero(get_time_by_cycle(max(tiu_cycles), tiu_frequency) +
                                                     get_time_by_cycle(max(gdma_cycles) + max(sdma_cycles), dma_frequency), total_time))
        else:
            prallelisms.append(get_ratio_str_2f_zero(get_time_by_cycle(max(tiu_cycles), tiu_frequency) +
                                                     get_time_by_cycle(max(gdma_cycles), dma_frequency), total_time))
        tiu_max_time = get_time_by_cycle(max(tiu_cycles), tiu_frequency)
        gdma_max_time = get_time_by_cycle(max(gdma_cycles), dma_frequency)
        concurrencys.append(get_ratio_str_2f_zero(tiu_max_time + gdma_max_time - total_time, min(tiu_max_time, gdma_max_time)))
        uArch_rates.append(get_ratio_str_2f_zero(sum(alg_opss), sum(uArch_opss)))
        total_times.append(max(total_times))
        tiu_cycles.append(max(tiu_cycles))
        alg_cycles.append(max(alg_cycles))
        alg_opss.append(sum(alg_opss))
        uArch_opss.append(sum(uArch_opss))
        gdma_cycles.append(max(gdma_cycles))
        gdma_ddr_datasizes.append(sum(gdma_ddr_datasizes))
        gdma_l2_datasizes.append(sum(gdma_l2_datasizes))
        gdma_ddr_avg_bds.append(get_ratio_float_2f(gdma_ddr_datasizes[-1], get_time_by_cycle(gdma_ddr_cycles, dma_frequency)))
        gdma_l2_avg_bds.append(get_ratio_float_2f(gdma_l2_datasizes[-1], get_time_by_cycle(gdma_l2_cycles, dma_frequency)))
        gdma_ddr_avg_bls.append(get_ratio_float_2f(gdma_bl_sum, gdma_xact_sum))

        sdma_cycles.append(max(sdma_cycles))
        sdma_ddr_datasizes.append(sum(sdma_ddr_datasizes))
        sdma_ddr_avg_bds.append(get_ratio_float_2f(sdma_ddr_datasizes[-1], get_time_by_cycle(sdma_ddr_cycles, dma_frequency)))
        sdma_ddr_avg_bls.append(get_ratio_float_2f(sdma_bl_sum, sdma_xact_sum))

        cdma_cycles.append(max(cdma_cycles))
        cdma_ddr_datasizes.append(sum(cdma_ddr_datasizes))
        cdma_l2_datasizes.append(sum(cdma_l2_datasizes))
        cdma_ddr_avg_bds.append(get_ratio_float_2f(cdma_ddr_datasizes[-1], get_time_by_cycle(cdma_ddr_cycles, dma_frequency)))
        cdma_l2_avg_bds.append(get_ratio_float_2f(cdma_l2_datasizes[-1], get_time_by_cycle(cdma_l2_cycles, dma_frequency)))
        cdma_ddr_avg_bls.append(get_ratio_float_2f(cdma_bl_sum, cdma_xact_sum))
        for idx in range(len(total_times)):
            total_times[idx] = cycle_to_us(total_times[idx], 1000)
        for idx in [len(total_times) - 1]:
            tiu_cycles[idx]  = cycle_to_us(tiu_cycles[idx], tiu_frequency, with_unit=True)
            alg_cycles[idx]  = cycle_to_us(alg_cycles[idx], tiu_frequency, with_unit=True)
            gdma_cycles[idx] = cycle_to_us(gdma_cycles[idx], dma_frequency, with_unit=True)
            sdma_cycles[idx] = cycle_to_us(sdma_cycles[idx], dma_frequency, with_unit=True)
            cdma_cycles[idx] = cycle_to_us(cdma_cycles[idx], dma_frequency, with_unit=True)
            gdma_ddr_datasizes[idx] = datasize_to_MB(gdma_ddr_datasizes[idx])
            gdma_l2_datasizes[idx] = datasize_to_MB(gdma_l2_datasizes[idx])
            sdma_ddr_datasizes[idx] = datasize_to_MB(sdma_ddr_datasizes[idx])
            cdma_ddr_datasizes[idx] = datasize_to_MB(cdma_ddr_datasizes[idx])
            cdma_l2_datasizes[idx] = datasize_to_MB(cdma_l2_datasizes[idx])
        # if chip_arch['Chip Arch'].lower() == 'a2' and self.act_core_num == 1:
        #     chip_arch['DDR Max BW(GB/s)'] = int(chip_arch['DDR Max BW(GB/s)']) * 2
        self.data = transpose([core_ids, tiu_work_ratios, prallelisms, concurrencys, total_times, tiu_cycles, alg_cycles,
                            alg_opss, uArch_opss, uArch_rates, gdma_cycles,
                            gdma_ddr_datasizes, gdma_l2_datasizes, gdma_ddr_avg_bds,
                            gdma_l2_avg_bds, gdma_ddr_avg_bls, sdma_cycles, sdma_ddr_datasizes,
                            sdma_ddr_avg_bds, sdma_ddr_avg_bls, cdma_cycles,
                            cdma_ddr_datasizes, cdma_l2_datasizes, cdma_ddr_avg_bds,
                            cdma_l2_avg_bds, cdma_ddr_avg_bls]).tolist()

    def write(self):
        """
        Write summary information to Excel.
        :return: None
        """
        df = pd.DataFrame(self.data, columns=self.columns, index=None)
        df.to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=4, engine='xlsxwriter', float_format='%g')
        para_desc = '(totalTiuCycle + totalGdmaCycle + totalSdmaCycle) / totalTime, presents the parallelism among all engines.'\
        if self.act_core_num > 1 else '(totalTiuCycle + totalGdmaCycle) / totalTime, presents the parallelism among all engines.'
        kpi_desc = pd.DataFrame(
                {
                    'Field': [
                        'Tiu Working Ratio',
                        'Parallelism(%)',
                        'Concurrency(%)',
                        'Total Alg Cycle',
                        'Total Alg Ops',
                        'Total uArch Ops',
                        'uArch Urate'],
                    'Description': [
                        'totalTiuCycle / totalTime, indicates the percentage of time tiu execution takes.',
                        para_desc,
                        '(totalTiuCycle + totalGdmaCycle - totalTime) / min(totalTiuCycle, totalGdmaCycle), indicates the concurrency between tiu and gdma.',  # todo
                        'The time required to execute tiu instructions theoretically.',
                        'The theoretical OPs required to execute tiu instructions.',
                        'The actual number of OPs accounting for microarchitecture.',
                        'totalAlgOps / totalUArchOps, since the shape of tensor needs to be aligned in '
                        'micro-architecture, the actual ops will be greater than the algorithm value. '
                        "It's better to closer to 100%."
                        ]
                },
                index=None
            )
        kpi_desc.to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=9+self.act_core_num, startcol=0,
                                        engine='xlsxwriter', float_format='%g')

    @classmethod
    def set_style(cls, file_path, chip_arch):
        """
        Set style for output Excel.
        :param file_path: the output Excel path.
        :return: None
        """
        wb = load_workbook(file_path)
        if cls.sheet_name not in wb.sheetnames:
            return
        df = pd.read_excel(file_path, cls.sheet_name)
        ws = wb[cls.sheet_name]
        ws.cell(1, 1).value = 'Platform: ' + chip_arch['Chip Arch']
        ws.cell(1, 1).fill = DetailsStyle.title_pattern
        ws.cell(1, 1).font = DetailsStyle.title_font
        ws.merge_cells('A1:B1')
        # set header style
        ws.cell(4, 8).value = 'TIU'
        ws.cell(4, 14).value = 'GDMA'
        ws.cell(4, 19).value = 'SDMA'
        ws.cell(4, 24).value = 'CDMA'
        for h, w in zip([4, 4, 4, 4], [8, 14, 19, 24]):
            ws.cell(h, w).fill = DetailsStyle.title_pattern
            ws.cell(h, w).font = DetailsStyle.title_font
        for w in range(6, 12):
            ws.cell(4, w).fill = DetailsStyle.tiu_pattern
        for w in range(11, 18):
            ws.cell(4, w).fill = DetailsStyle.gdma_pattern
        for w in range(17, 22):
            ws.cell(4, w).fill = DetailsStyle.sdma_pattern
        for w in range(21, 27):
            ws.cell(4, w).fill = DetailsStyle.cdma_pattern
        # set title style
        for cell in ws['5:5']:
            cell.fill = DetailsStyle.content_pattern
            cell.font = DetailsStyle.title_header_font
        # set content style
        content_start_rows = 6
        content_end_rows = len(df) + content_start_rows - 4
        for w in range(len(df.columns)):
            for h in range(content_start_rows, len(df) + content_start_rows):
                ws.cell(h, w + 1).font = DetailsStyle.title_font
        # set highlight style
        ratio_pos = 2
        parallelismPos = 3
        concurrencyPos = 4
        uRate_pos = 10
        total_uArch_ops = 9
        gdma_ddr_bd_pos = 14
        gdma_l2_bd_pos = 15
        gdma_ddr_avg_bl_pos = 16
        sdma_ddr_bd_pos = 19
        sdma_ddr_avg_bl_pos = 20
        cdma_ddr_bd_pos = 24
        cdma_l2_bd_pos = 25
        cdma_ddr_avg_bl_pos = 26
        chart_len = len(df) - 9
        for h in range(content_start_rows, chart_len):
            if int(ws.cell(h, total_uArch_ops).value) > 0:
                # ratio_pos
                if float(ws.cell(h, ratio_pos).value[:-1]) < 15:
                    ws.cell(h, ratio_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, ratio_pos).value[:-1]) < 30:
                    ws.cell(h, ratio_pos).fill = DetailsStyle.yellow_light
                # uRate_pos
                if float(ws.cell(h, uRate_pos).value[:-1]) < 50:
                    ws.cell(h, uRate_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, uRate_pos).value[:-1]) < 75:
                    ws.cell(h, uRate_pos).fill = DetailsStyle.yellow_light
            # parallelismPos
            if chip_arch['Chip Arch'].lower() in ["sg2260", 'a2', 'A2']:
                if float(ws.cell(h, parallelismPos).value[:-1]) < 205:
                    ws.cell(h, parallelismPos).fill = DetailsStyle.yellow_light
            elif chip_arch['Chip Arch'].lower() in ['cv186x', 'bm1684x', 'mars3', 'bm1688']:
                if float(ws.cell(h, parallelismPos).value[:-1]) < 105:
                    ws.cell(h, parallelismPos).fill = DetailsStyle.red_light
            else:
                print('Not support chip arch')
                assert(0)
            # gdma_ddr_bd_pos
            ddr_max_bd, l2_max_bd, ddr_max_bl = float(chip_arch['DDR Max BW(GB/s/Core)']), float(chip_arch['L2 Max BW(GB/s)']), float(chip_arch['Bus Max Burst'])
            if float(ws.cell(h, gdma_ddr_bd_pos).value) > 0:
                if float(ws.cell(h, gdma_ddr_bd_pos).value) < (ddr_max_bd * 0.5):
                    ws.cell(h, gdma_ddr_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, gdma_ddr_bd_pos).value) < (ddr_max_bd * 0.75):
                    ws.cell(h, gdma_ddr_bd_pos).fill = DetailsStyle.yellow_light
            # gdma_l2_bd_pos
            if float(ws.cell(h, gdma_l2_bd_pos).value) > 0:
                if float(ws.cell(h, gdma_l2_bd_pos).value) < (l2_max_bd * 0.5):
                    ws.cell(h, gdma_l2_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, gdma_l2_bd_pos).value) < (l2_max_bd * 0.75):
                    ws.cell(h, gdma_l2_bd_pos).fill = DetailsStyle.yellow_light
                # gdma_ddr_avg_bl_pos
                if float(ws.cell(h, gdma_ddr_avg_bl_pos).value) < ddr_max_bl:
                    ws.cell(h, gdma_ddr_avg_bl_pos).fill = DetailsStyle.red_light
            # sdma_ddr_bd_pos
            if float(ws.cell(h, sdma_ddr_bd_pos).value) > 0:
                if float(ws.cell(h, sdma_ddr_bd_pos).value) < (ddr_max_bd * 0.5):
                    ws.cell(h, sdma_ddr_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, sdma_ddr_bd_pos).value) < (ddr_max_bd * 0.75):
                    ws.cell(h, sdma_ddr_bd_pos).fill = DetailsStyle.yellow_light
                # sdma_ddr_avg_bl_pos
                if float(ws.cell(h, sdma_ddr_avg_bl_pos).value) < ddr_max_bl:
                    ws.cell(h, sdma_ddr_avg_bl_pos).fill = DetailsStyle.red_light
            # cdma_l2_bd_pos
            if ws.cell(h, cdma_l2_bd_pos).value and float(ws.cell(h, cdma_l2_bd_pos).value) > 0:
                if float(ws.cell(h, cdma_l2_bd_pos).value) < (l2_max_bd * 0.5):
                    ws.cell(h, cdma_l2_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, cdma_l2_bd_pos).value) < (l2_max_bd * 0.75):
                    ws.cell(h, cdma_l2_bd_pos).fill = DetailsStyle.yellow_light
                # cdma_ddr_avg_bl_pos
                if float(ws.cell(h, cdma_ddr_avg_bl_pos).value) < ddr_max_bl:
                    ws.cell(h, cdma_ddr_avg_bl_pos).fill = DetailsStyle.red_light
        content_end_rows = chart_len
        if chip_arch['Chip Arch'].lower() in ["sg2260", 'a2', 'A2']:
            ws.cell(content_end_rows, parallelismPos).value = 'Parallelism<205%'
        elif chip_arch['Chip Arch'].lower() in ['cv186x', 'bm1684x', 'mars3', 'bm1688']:
            ws.cell(content_end_rows, parallelismPos).value = 'Parallelism<105%'
        else:
            print('Not support chip arch')
            assert(0)
        ws.cell(content_end_rows, ratio_pos).value = 'Ratio<30%'
        ws.cell(content_end_rows + 1, ratio_pos).value = 'Ratio<15%'
        ws.cell(content_end_rows, uRate_pos).value = 'Urate<75%'
        ws.cell(content_end_rows + 1, uRate_pos).value = 'Urate<50%'
        ws.cell(content_end_rows, gdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, gdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*50%'
        ws.cell(content_end_rows, gdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, gdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*50%'
        ws.cell(content_end_rows, gdma_ddr_avg_bl_pos).value = 'Length<' + str(ddr_max_bl)
        ws.cell(content_end_rows, sdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, sdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*50%'
        ws.cell(content_end_rows, sdma_ddr_avg_bl_pos).value = 'Length<' + str(ddr_max_bl)
        ws.cell(content_end_rows, cdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, cdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*50%'
        ws.cell(content_end_rows, cdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, cdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*50%'
        ws.cell(content_end_rows, cdma_ddr_avg_bl_pos).value = 'Length<' + str(ddr_max_bl)
        for h, w in zip([content_end_rows, content_end_rows, content_end_rows, content_end_rows,
                         content_end_rows, content_end_rows, content_end_rows, content_end_rows],
                        [ratio_pos, parallelismPos, uRate_pos, gdma_ddr_bd_pos, gdma_l2_bd_pos, sdma_ddr_bd_pos, cdma_ddr_bd_pos, cdma_l2_bd_pos]):
            ws.cell(h, w).fill = DetailsStyle.yellow_light
            ws.cell(h, w).font = DetailsStyle.title_font
            ws.cell(h, w).alignment = DetailsStyle.center_align
        for h, w in zip([content_end_rows + 1, content_end_rows + 1, content_end_rows + 1, content_end_rows + 1, content_end_rows,
                        content_end_rows + 1, content_end_rows, content_end_rows + 1, content_end_rows + 1, content_end_rows],
                        [ratio_pos, uRate_pos, gdma_ddr_bd_pos, gdma_l2_bd_pos, gdma_ddr_avg_bl_pos, sdma_ddr_bd_pos, sdma_ddr_avg_bl_pos,
                         cdma_ddr_bd_pos, cdma_l2_bd_pos, cdma_ddr_avg_bl_pos]):
            ws.cell(h, w).fill = DetailsStyle.red_light
            ws.cell(h, w).font = DetailsStyle.title_font
            ws.cell(h, w).alignment = DetailsStyle.center_align
        # set border style
        header_start_row = 5
        for h in range(header_start_row, chart_len + header_start_row - 5):
            for w in range(1, len(df.columns) + 1):
                ws.cell(h, w).border = DetailsStyle.border
                ws.cell(h, w).alignment = DetailsStyle.center_align
        # set auto columns width
        for col in df.columns:
            index = list(df.columns).index(col)
            letter = get_column_letter(index + 1)
            collen = len(col)
            for h in range(1, chart_len + 4):
                if ws.cell(h, index + 1).value != None:
                    collen = max(collen, len(str(ws.cell(h, index + 1).value)))
            ws.column_dimensions[letter].width = collen * 1.05
        ws.sheet_properties.tabColor = 'FFC7CE'
        wb.save(file_path)
