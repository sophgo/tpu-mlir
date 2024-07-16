#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
# @Time    : 2023/7/25 17:55
# @Author  : chongqing.zeng@sophgo.com
# @Project: PerfAI
import pandas as pd
import sys
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from decimal import Decimal
from PerfAI.utils.utils import remove_duplicate_path, get_simulator_total_cycle, get_profile_cycle, lcs_dp
from PerfAI.definition.style import DetailsStyle


class ABCSetSummary(object):

    def __init__(self, out_path, p_paths, f_paths, s_paths):
        self.out_path = out_path
        self.model_names = []
        self.p_paths = p_paths
        self.f_paths = f_paths
        self.s_paths = s_paths
        self.df = pd.DataFrame()
        self.model_nums = -1
        self.evb_cycles = []
        self.overalls = []
        self.diff1s = []
        self.tiu_evbs = []
        self.tiu_cycles = []
        self.diff2s = []
        self.dma_evbs = []
        self.dma_cycles = []
        self.diff3s = []
        self.notes = []

    def load(self):
        self.model_nums = len(self.p_paths)
        if self.model_nums > 1:
            self.model_names = remove_duplicate_path(self.f_paths)
        else:
            lcs_res = lcs_dp(self.p_paths[0], self.f_paths[0])
            model_name = self.f_paths[0]
            model_name = ''.join([model_name[i] for i in lcs_res if i != -1])
            self.model_names.append(model_name.split('/')[-1])
        tmp_names = []
        for name in self.model_names:
            tmp_names.append(name + '(Compiler)')
            tmp_names.append(name + '(Model)')
        self.model_names = tmp_names
        for i in range(0, self.model_nums):
            sim_total_cycle = get_simulator_total_cycle(self.s_paths[i])
            compiler_total_cycle = get_profile_cycle(self.p_paths[i])
            self.df = pd.read_csv(self.f_paths[i])
            real_tiu_cycle, real_dma_cycle, simulator_tiu_cycle, simulator_dma_cycle = 0, 0, 0, 0
            compiler_tiu_cycle, compiler_dma_cycle, start_cycle, end_cycle = 0, 0, sys.maxsize, -1
            for index in range(0, len(self.df)):
                if self.df['EngineId'][index] == 0:
                    real_tiu_cycle += int(self.df['1684x Cycle'][index])
                    simulator_tiu_cycle += int(self.df['CycleCount'][index])
                    compiler_tiu_cycle += int(self.df['OriginCycleCount'][index])
                elif self.df['EngineId'][index] == 2:
                    start_cycle = min(self.df['StartCycle'][index], start_cycle)
                    end_cycle = max(self.df['EndCycle'][index], end_cycle)
                    real_dma_cycle += int(self.df['1684x Cycle'][index])
                    simulator_dma_cycle += int(self.df['CycleCount'][index])
                    compiler_dma_cycle += int(self.df['OriginCycleCount'][index])
            real_total_cycle = end_cycle - start_cycle
            self.evb_cycles.extend([real_total_cycle, real_total_cycle])
            self.overalls.extend([compiler_total_cycle, sim_total_cycle])
            self.diff1s.extend([str((Decimal(((compiler_total_cycle - real_total_cycle) / real_total_cycle) * 100)).
                                    quantize(Decimal("0.00"))) + '%',
                                str((Decimal(((sim_total_cycle - real_total_cycle) / real_total_cycle) * 100)).
                                    quantize(Decimal("0.00"))) + '%'])
            self.tiu_cycles.extend([compiler_tiu_cycle, simulator_tiu_cycle])
            self.tiu_evbs.extend([real_tiu_cycle, real_tiu_cycle])
            self.diff2s.extend([str((Decimal(((compiler_tiu_cycle - real_tiu_cycle) / real_tiu_cycle) * 100)).
                                    quantize(Decimal("0.00"))) + '%',
                                str((Decimal(((simulator_tiu_cycle - real_tiu_cycle) / real_tiu_cycle) * 100)).
                                    quantize(Decimal("0.00"))) + '%'])
            self.dma_evbs.extend([real_dma_cycle, real_dma_cycle])
            self.dma_cycles.extend([compiler_dma_cycle, simulator_dma_cycle])
            self.diff3s.extend([str((Decimal(((compiler_dma_cycle - real_dma_cycle) / real_dma_cycle) * 100)).
                                    quantize(Decimal("0.00"))) + '%',
                                str((Decimal(((simulator_dma_cycle - real_dma_cycle) / real_dma_cycle) * 100)).
                                    quantize(Decimal("0.00"))) + '%'])
            self.notes.extend(['', ''])

    def write(self):
        total_table = dict()
        total_table['Model Name'] = self.model_names
        total_table['1684x EVB(cycle)'] = self.evb_cycles
        total_table['Overall(cycle)'] = self.overalls
        total_table['Diff1(%)'] = self.diff1s
        total_table['1684x Tiu EVB'] = self.tiu_evbs
        total_table['Tiu Cycle'] = self.tiu_cycles
        total_table['Diff2(%)'] = self.diff2s
        total_table['1684x Tdma EVB'] = self.dma_evbs
        total_table['Tdma Cycle'] = self.dma_cycles
        total_table['Diff3(%)'] = self.diff3s
        total_table['Note'] = self.notes
        df = pd.DataFrame(total_table)
        df.to_excel(self.out_path, index=None)


    def set_style(self):
        tiuDf = pd.read_excel(self.out_path)
        wb = load_workbook(self.out_path)
        wsTiu = wb['Sheet1']
        for cell in wsTiu['1:1']:
            # title style
            cell.fill = DetailsStyle.content_pattern
            cell.font = DetailsStyle.title_header_font
        endIndex = 10
        for w in range(endIndex):
            for h in range(len(tiuDf)):
                # key content style
                wsTiu.cell(h + 2, w + 1).fill = DetailsStyle.key_content_pattern
                wsTiu.cell(h + 2, w + 1).font = DetailsStyle.title_font
        for w in range(endIndex, len(tiuDf.columns)):
            for h in range(len(tiuDf)):
                # other content style
                wsTiu.cell(h + 2, w + 1).font = DetailsStyle.title_font
        for col in tiuDf.columns:
            # auto set columns width
            index = list(tiuDf.columns).index(col)
            letter = get_column_letter(index + 1)
            collen = max(tiuDf[col].apply(lambda x: len(str(x).encode())).max(), len(col))
            wsTiu.column_dimensions[letter].width = collen * 1.1
        for h in range(1, len(tiuDf) + 2):
            for w in range(1, len(tiuDf.columns) + 1):
                # set border style
                wsTiu.cell(h, w).border = DetailsStyle.border
                if h != 1:
                    wsTiu.cell(h, w).alignment = DetailsStyle.right_align
        wsTiu.sheet_properties.tabColor = '008000'
        wb.save(self.out_path)
