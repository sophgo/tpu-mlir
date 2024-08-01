#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
# @Time    : 2023/8/7 11:24
# @Author  : chongqing.zeng@sophgo.com
# @Project: PerfAI
import pandas as pd
import os

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from definition.style import LayerStyle
from utils.utils import get_dtype_size


def get_trans_cost(ddr_bytes, lmem_bytes, ddr_bw, lmem_bw):
    if lmem_bw:
        cost = get_ratio(ddr_bytes / ddr_bw + lmem_bytes / lmem_bw, 1)
    else:
        cost = get_ratio(ddr_bytes / ddr_bw, 1)
    return cost


def get_layer_nchw(layer, is_in):
    shape = [1, 1, 1, 1]
    tensors = layer.in_tensors if is_in else layer.out_tensors
    if tensors:
        for i, s in enumerate(tensors[0].shape):
            if i < 4:
                shape[i] = s
            else:
                shape[3] * s
    return shape


def get_layer_fm(layer, is_in:bool):
    total_bytes = 0
    tensors = layer.in_tensors if is_in else layer.out_tensors
    for tensor in tensors:
        tensor_bytes = get_dtype_size(tensor.dtype)
        for s in tensor.shape:
            tensor_bytes *= s
        total_bytes += tensor_bytes
    return total_bytes


def get_layer_dtype(layer):
    if layer.in_tensors:
        return layer.in_tensors[0].dtype
    if layer.out_tensors:
        return layer.out_tensors[0].dtype
    return None


def get_layer_bytes(layer, is_in):
    total_bytes = 0
    tensors = layer.in_tensors if is_in else layer.out_tensors
    for tensor in tensors:
        tensor_bytes = get_layer_dtype(tensor.dtype)
        for s in tensor.shape:
            tensor_bytes *= s
        total_bytes += tensor_bytes
    return total_bytes


def get_layer_weight_size(layer):
    total_bytes = 0
    tensors = layer.in_tensors
    for tensor in tensors:
        if not tensor.is_const:
            continue
        tensor_bytes = get_dtype_size(tensor.dtype)
        for s in tensor.shape:
            tensor_bytes *= s
        total_bytes += tensor_bytes
    return total_bytes

def get_layer_feature_size(layer):
    total_bytes = 0
    tensors = layer.in_tensors
    for tensor in tensors:
        if tensor.is_const:
            continue
        tensor_bytes = get_dtype_size(tensor.dtype)
        for s in tensor.shape:
            tensor_bytes *= s
        total_bytes += tensor_bytes
    return total_bytes

def get_ratio_str(x, y):
    return '%.2f%%' % (x / y * 100) if y != 0 else "--"


def get_ratio(x, y):
    return "%.2f" % (x / y) if y != 0 else 0


def get_golden_slice(m, k, n, dtype):
    matmul_tool_path = 'bin/matmulTiling'
    output = matmul_tool_path + '.csv'
    assert os.path.exists(matmul_tool_path)
    A = 1
    B = 1
    C = 1
    prec = 0
    if dtype == "FP32":
        prec = 1
    elif dtype == "INT8" or dtype == "UINT8":
        prec = 2
        # not support now
        return '-', '-', '-', '-', '-', '-', '-'
    cmd = './' + matmul_tool_path + ' --M ' + str(m) + ' --K ' + str(k) + ' --N ' + str(n) + \
          ' --A ' + str(A) + ' --B ' + str(B) + ' --C ' + str(C) + \
          ' --prec ' + str(prec) + ' --out ' + output
    os.system(cmd)
    assert os.path.exists(output)
    df = pd.read_csv(output, skiprows=1)
    if df.empty:
        return '-', '-', '-', '-', '-', '-', '-'
    else:
        df.sort_values(by="Cost", inplace=True, ascending=True, ignore_index=True)
        return df['Msecs_o'][0], df['Ksecs_o'][0], df['Nsecs_o'][0], df['Mslice'][0], df['Kslice'][0], df['Nslice'][0], df['Cost'][0]


class Layer:
    store_bytes = 0
    load_bytes = 0
    load_cycles = 0
    store_cycles = 0
    input_shape = []
    output_shape = []

    def __init__(self):
        self.layer_id = -1
        self.layer_name = ''
        self.data_type = ''
        self.layer_type = None
        self.engine_type = 'TPU'
        self.load_bandwidth = 0
        self.store_bandwidth = 0
        self.kh = 0
        self.kw = 0
        self.k_stride_h = 0
        self.k_stride_w = 0
        self.iN = 0
        self.iC = 0
        self.iH = 0
        self.iW = 0
        self.oN = 0
        self.oC = 0
        self.oH = 0
        self.oW = 0
        self.ifm = 0
        self.ofm = 0
        self.m = 0
        self.k = 0
        self.n = 0
        self.weight_size = 0
        self.feature_size = 0

        self.alg_ops = 0
        self.other_info = ''
        self.uarch_rate = 0
        self.uarch_ops = 0
        self.alg_cycle = 0
        self.alg_cycle_ratio = 0

        self.g_m_secs_o = 0
        self.g_n_secs_o = 0
        self.g_k_secs_o = 0
        self.g_m_slice = 0
        self.g_n_slice = 0
        self.g_k_slice = 0
        self.g_cost = 0

        self.ddr_bytes = 0
        self.lmem_bytes = 0

        self.a_m_secs_o = 0
        self.a_n_secs_o = 0
        self.a_k_secs_o = 0
        self.a_m_slice = 0
        self.a_n_slice = 0
        self.a_k_slice = 0
        self.a_cost = 0
        self.sim_cycle = 0
        self.sim_cycle_ratio = 0


class LayerInfo:
    def __init__(self):
        self.layer_id = -1
        self.core_id = 0
        self.layer_type = None
        self.layer_name = ""
        self.is_local = False
        self.in_tensors = []
        self.out_tensors = []
        self.group_id = -1
        self.total_size = 0
        self.feature_size = 0
        self.weight_size = 0
        self.gdma_op = None
        self.gdma_tensor = None
        self.gdma_nodes = []
        self.bd_nodes = []
        self.engine_type = None

    def add_input(self, tensor):
        if tensor in self.in_tensors:
            return
        self.in_tensors.append(tensor)
        tensor.out_layers.append(self)

    def add_output(self, tensor):
        if tensor in self.out_tensors:
            return
        self.out_tensors.append(tensor)
        tensor.in_layer = self

    def set_gdma_tensor(self, tensor):
        self.gdma_tensor = tensor


class TotalLayerInfo:
    def __init__(self, writer, layer_info_list):
        self.layer_infos = layer_info_list
        self.writer = writer
        self.custom_layer_info = []
        self.custom_layers = []
        self.sheet_name = 'Layer by Layer Info'

    def add_kpi_field(self, tiu_instances, gdma_instances, chip_arch):
        layer_id_map = dict()
        total_tiu_cycles = 0
        total_gdma_cycles = 0
        total_alg_cycles = 0
        ddr_bw, l2_bw = chip_arch['DDR Max BW(GB/s/Core)'], chip_arch['L2 Max BW(GB/s)']
        for layer_info in self.layer_infos:
            if len(layer_info.bd_nodes) == 0 or layer_info.layer_name in ("Load", "Store"):
                continue
            if layer_info.layer_id not in layer_id_map.keys():
                layer = Layer()
                layer.layer_id = layer_info.layer_id
                layer.data_type = get_layer_dtype(layer_info)
                layer.layer_name = layer_info.layer_name
                layer.layer_type = layer_info.layer_type
                layer.engine_type = layer_info.engine_type
                layer.input_shape = get_layer_nchw(layer_info, True)
                layer.iN, layer.iC, layer.iH, layer.iW = layer.input_shape[0], layer.input_shape[1], \
                    layer.input_shape[2], layer.input_shape[3]
                layer.output_shape = get_layer_nchw(layer_info, False)
                layer.oN, layer.oC, layer.oH, layer.oW = layer.output_shape[0], layer.output_shape[1], \
                    layer.output_shape[2], layer.output_shape[3]
                layer.ifm = get_layer_fm(layer_info, True)
                layer.ofm = get_layer_fm(layer_info, False)
                if layer.layer_name.lower() == 'matmul' and len(layer_info.in_tensors) == 3:
                    in_tensor_0 = layer_info.in_tensors[0]
                    in_tensor_1 = layer_info.in_tensors[1]
                    # assert in_tensor_0.shape[-1] == in_tensor_1.shape[-2]
                    # layer.m, layer.k, layer.n = in_tensor_0.shape[-2], in_tensor_0.shape[-1], in_tensor_1.shape[-1]
                elif layer.layer_name.lower() == 'batch_matmul' and len(layer_info.in_tensors) == 3:
                    in_tensor_0 = layer_info.in_tensors[0]
                    in_tensor_1 = layer_info.in_tensors[1]
                    # assert in_tensor_0.shape[2] == in_tensor_1.shape[1]
                    # layer.m, layer.k, layer.n = in_tensor_0.shape[1], in_tensor_0.shape[2], in_tensor_1.shape[2]
                layer.weight_size = get_layer_weight_size(layer_info)
                layer.feature_size = get_layer_feature_size(layer_info)
                if len(layer_info.in_tensors) > 1:
                    for i in range(len(layer_info.in_tensors)):
                        layer.other_info += "tensor_" + str(i) + ': ' + str(layer_info.in_tensors[i].shape)
                for n in layer_info.bd_nodes:
                    if (n.bd_id, n.core_id) not in tiu_instances.keys():
                        continue
                    tiu_node = tiu_instances[n.bd_id, n.core_id]
                    layer.alg_ops += tiu_node.alg_ops
                    layer.uarch_ops += tiu_node.uarch_ops
                    layer.sim_cycle += tiu_node.cycle
                    layer.alg_cycle += tiu_node.alg_cycle
                    total_tiu_cycles += tiu_node.cycle
                    total_alg_cycles += tiu_node.alg_cycle
                    if tiu_node.des_tsk_typ in [0, 1]:  # conv、pord
                        layer.kh = max(layer.kh, int(tiu_node.des_opd1_h))
                        layer.kw = max(layer.kw, int(tiu_node.des_opd1_w))
                        tiu_node.des_opd1_h_str = 0 if not str(tiu_node.des_opd1_h_str).isnumeric() \
                            else tiu_node.des_opd1_h_str
                        tiu_node.des_opd1_w_str = 0 if not str(tiu_node.des_opd1_w_str).isnumeric() \
                            else tiu_node.des_opd1_w_str
                        layer.k_stride_h = max(layer.k_stride_h, int(tiu_node.des_opd1_h_str))
                        layer.k_stride_w = max(layer.k_stride_w, int(tiu_node.des_opd1_w_str))

                for n in layer_info.gdma_nodes:
                    if (n.gdma_id, n.core_id) not in gdma_instances.keys():
                        continue
                    gdma_node = gdma_instances[(n.gdma_id, n.core_id)]
                    if gdma_node.direction.lower().endswith('ddr') or \
                        gdma_node.direction.lower().endswith('l2'):
                        layer.store_bytes += gdma_node.datasize
                        layer.store_cycles += gdma_node.cycle
                        total_gdma_cycles += gdma_node.cycle
                    elif gdma_node.direction.lower().endswith('lmem'):
                        layer.load_bytes += gdma_node.datasize
                        layer.load_cycles += gdma_node.cycle
                        total_gdma_cycles += gdma_node.cycle
                    if 'ddr' in gdma_node.direction.lower():
                        layer.ddr_bytes += gdma_node.datasize
                    else:
                        layer.lmem_bytes += gdma_node.datasize
                # if 'matmul' in layer.layer_name.lower():
                #     layer.g_m_secs_o, layer.g_n_secs_o, layer.g_k_secs_o, layer.g_m_slice, \
                #         layer.g_n_slice, layer.g_k_slice, layer.g_cost = get_golden_slice(layer.m, layer.k, layer.n, layer.data_type)
                #     layer.a_m_secs_o, layer.a_n_secs_o, layer.a_k_secs_o, \
                #         layer.a_m_slice, layer.a_n_slice, layer.a_k_slice, layer.a_cost = 1, 1, 1, 1, 1, 1, \
                #         get_trans_cost(layer.ddr_bytes, layer.lmem_bytes, ddr_bw, l2_bw)
                # else:
                layer.a_m_slice, layer.a_n_slice, layer.a_k_slice, layer.a_cost = '-', '-', '-', '-'
                layer.g_m_slice, layer.g_n_slice, layer.g_k_slice, layer.g_cost = '-', '-', '-', '-'
                layer.a_m_secs_o, layer.a_n_secs_o, layer.a_k_secs_o = '-', '-', '-'
                layer.g_m_secs_o, layer.g_n_secs_o, layer.g_k_secs_o = '-', '-', '-'
                layer_id_map[layer_info.layer_id] = layer
            else:
                for n in layer_info.bd_nodes:
                    tiu_node = tiu_instances[n.bd_id, n.core_id]
                    layer_id_map[layer_info.layer_id].sim_cycle += tiu_node.cycle
                    layer_id_map[layer_info.layer_id].alg_cycle += tiu_node.alg_cycle
                    layer_id_map[layer_info.layer_id].uarch_ops += tiu_node.uarch_ops
                    layer_id_map[layer_info.layer_id].alg_ops += tiu_node.alg_ops
                    total_tiu_cycles += tiu_node.cycle
                    total_alg_cycles += tiu_node.alg_cycle
                for n in layer_info.gdma_nodes:
                    gdma_node = gdma_instances[n.gdma_id, n.core_id]
                    # print(layer_info.layer_id, gdma_node.__dict__)
                    if gdma_node.direction.lower().endswith('ddr') or \
                            gdma_node.direction.lower().endswith('l2'):
                        layer_id_map[layer_info.layer_id].store_bytes += gdma_node.datasize
                        layer_id_map[layer_info.layer_id].store_cycles += gdma_node.cycle
                        total_gdma_cycles += gdma_node.cycle
                    elif gdma_node.direction.lower().endswith('lmem'):
                        layer_id_map[layer_info.layer_id].load_bytes += gdma_node.datasize
                        layer_id_map[layer_info.layer_id].load_cycles += gdma_node.cycle
                        total_gdma_cycles += gdma_node.cycle
                    if 'ddr' in gdma_node.direction.lower():
                        layer_id_map[layer_info.layer_id].ddr_bytes += gdma_node.datasize
                    else:
                        layer_id_map[layer_info.layer_id].lmem_bytes += gdma_node.datasize
                if 'matmul' in layer_id_map[layer_info.layer_id].layer_name.lower():
                    layer_id_map[layer_info.layer_id] = get_trans_cost(layer_id_map[layer_info.layer_id].ddr_bytes,
                                                                       layer_id_map[layer_info.layer_id].lmem_bytes, ddr_bw, l2_bw)
        for k in layer_id_map.keys():
            layer = layer_id_map[k]
            layer.uarch_rate = get_ratio_str(layer.alg_ops, layer.uarch_ops)
            layer.sim_cycle += layer.load_cycles + layer.store_cycles
            layer.load_bandwidth = get_ratio(layer.load_bytes, layer.load_cycles)
            layer.store_bandwidth = get_ratio(layer.store_bytes, layer.store_cycles)
            layer.alg_cycle_ratio = get_ratio_str(layer.alg_cycle, total_alg_cycles) if total_alg_cycles > 0 else 0
        total_sim_cycle = sum(layer.sim_cycle for layer in layer_id_map.values())
        self.custom_layers = layer_id_map.values()
        for layer in layer_id_map.values():
            layer.sim_cycle_ratio = get_ratio_str(layer.sim_cycle, total_sim_cycle) if total_sim_cycle > 0 else 0
            tmp_dict = layer.__dict__
            pop_field = ['input_shape', 'output_shape', 'load_bytes', 'load_cycles', 'store_bytes', 'store_cycles',
                         'ddr_bytes', 'lmem_bytes', 'sim_cycle', 'sim_cycle_ratio']
            tmp_ddr_bytes = tmp_dict['ddr_bytes']
            tmp_lmem_bytes = tmp_dict['lmem_bytes']
            tmp_sim_cycle = tmp_dict['sim_cycle']
            tmp_sim_cycle_ratio = tmp_dict['sim_cycle_ratio']
            for field in pop_field:
                if field in tmp_dict.keys():
                    tmp_dict.pop(field)
            tmp_dict['ddr<->l1/l2 bytes(B)'] = tmp_ddr_bytes
            tmp_dict['l1<->l2 bytes(B)'] = tmp_lmem_bytes
            tmp_dict['sim_cycle'] = tmp_sim_cycle
            tmp_dict['sim_cycle_ratio'] = tmp_sim_cycle_ratio
            self.custom_layer_info.append(tmp_dict)
        # self.custom_layer_info.sort(key=lambda x: (int(x['sim_cycle']), int(x['layer_id'])), reverse=True)

    def pop_data(self):
        return self.custom_layers

    def write(self, chip_arch):
        network = chip_arch['network']
        platform = chip_arch['Chip Arch']
        if platform.lower() == 'sg2260':
            int8_ops = 256
            fp32_ops = 16
        elif platform.lower() == 'bm1684x':
            int8_ops = 32
            fp32_ops = 2.2
        elif platform.lower() == 'a2':
            int8_ops = 14.4
            fp32_ops = 0.45
        ddr_bw = round(float(chip_arch['DDR Max BW(GB/s/Core)']) * int(chip_arch['Core Num']), 2)
        tpu_freq = float(chip_arch['TIU Frequency(MHz)']) / 1000
        dma_freq = float(chip_arch['DMA Frequency(MHz)']) / 1000
        condition_dict = {
            'Network': [network],
            'platform': [platform],
            'TPU INT8 TOPS': [int8_ops],
            'TPU FP16 TOPS': [int8_ops / 2],
            'TPU FP32 TOPS': [fp32_ops],
            'DDR BW(GB/s/Core)': [ddr_bw],
            'TPU Frequency(GHz)': [tpu_freq],
            'DMA Frequency(GHz)': [dma_freq]
        }
        npu_num = int(chip_arch['NPU Num'])
        Cube_IC_Align = int(chip_arch['Cube IC Align(8bits)'])
        Conv_OHOW_Align = int(chip_arch['Cube OHOW Align(8bits)'])
        Vector_OHOW_Align = int(chip_arch['Vector OHOW Align(8bits)'])
        Conv_ops = int(npu_num * Cube_IC_Align * Conv_OHOW_Align * tpu_freq * 2 / 1000)
        Vector_ops = round(npu_num * Vector_OHOW_Align * tpu_freq / 1000, 2)
        pooling_ops = Conv_ops / 4
        detail_spec = {
            'NPU NUM (lane)': [npu_num],
            'IC Alignment (8bits)': [Cube_IC_Align],
            'Conv OHOW Align': [Conv_OHOW_Align],
            'Vector OHOW Align(8bits)': [Vector_OHOW_Align],
            'Conv Ops/s': [Conv_ops],
            'Vect Ops/s': [Vector_ops],
            'Pool OPs/s': [pooling_ops]
        }

        pd.DataFrame(condition_dict).to_excel(self.writer, index=False, sheet_name=self.sheet_name, engine='xlsxwriter',
                                              startrow=0, startcol=1, float_format='%g')
        pd.DataFrame(detail_spec).to_excel(self.writer, index=False, sheet_name=self.sheet_name, engine='xlsxwriter',
                                           startrow=3, startcol=1, float_format='%g')

        if len(self.custom_layer_info) > 0:
            pd.DataFrame(self.custom_layer_info).to_excel(self.writer, index=False, sheet_name=self.sheet_name,
                                                          startrow=8, startcol=0,
                                                          engine='xlsxwriter', float_format='%g')

    @classmethod
    def set_style(cls, file_path):
        """
        Set style for Excel, and highlight some fields we care about.
        :param file_path: Excel output path
        :return: None
        """
        wb = load_workbook(file_path)
        layer_style = LayerStyle()
        sheet_name = 'Layer by Layer Info'
        df = pd.read_excel(file_path, sheet_name)
        ws = wb[sheet_name]
        ws.cell(1, 1).value = 'Condition'
        ws.cell(4, 1).value = 'Detail Spec'
        ws.cell(8, 1).value = '*Yellow background indicates the topN time consuming layer'
        # ws.merge_cells('A8: C8')
        for h, w in zip([1, 4], [1, 1]):
            ws.cell(h, w).fill = layer_style.title_pattern
            ws.cell(h, w).font = layer_style.title_font
        summary_start_cols = 2
        summary_end_cols = summary_start_cols + 8
        for w in range(summary_start_cols, summary_end_cols):
            for h in [1, 4]:
                if h == 1:
                    ws.cell(h, w).fill = layer_style.title_header_pattern
                    ws.cell(h, w).font = layer_style.title_header_font
                    ws.cell(h, w).border = layer_style.border
                    ws.cell(h + 1, w).fill = layer_style.title_content_pattern
                    ws.cell(h + 1, w).font = layer_style.title_header_font
                    ws.cell(h + 1, w).alignment = layer_style.center_align
                    ws.cell(h + 1, w).border = layer_style.border
                elif h == 4 and w < summary_end_cols - 1:
                    # summary title style
                    ws.cell(h, w).fill = layer_style.title_header_pattern
                    ws.cell(h, w).font = layer_style.title_header_font
                    ws.cell(h, w).border = layer_style.border
                    # summary content style
                    ws.cell(h + 1, w).fill = layer_style.title_content_pattern
                    ws.cell(h + 1, w).font = layer_style.title_header_font
                    ws.cell(h + 1, w).alignment = layer_style.center_align
                    ws.cell(h + 1, w).border = layer_style.border

        content_start_cols = 1
        content_end_cols = content_start_cols + len(df.columns)
        content_header_row = 8
        sim_cycle_index = content_end_cols - 2
        sort_dict = dict()
        for w in range(content_start_cols, content_end_cols):
            ws.cell(content_header_row + 1, w).fill = layer_style.content_header2_pattern
            ws.cell(content_header_row + 1, w).font = layer_style.content_header_font
            for h in range(content_header_row + 2, len(df) + 2):
                ws.cell(h, w).font = layer_style.title_header_font
                ws.cell(h, w).alignment = layer_style.center_align
                ws.cell(h, w).border = layer_style.border
        for h in range(content_header_row + 2, len(df) + 2):
            sort_dict[h] = ws.cell(h, sim_cycle_index).value
        ws.cell(content_header_row, 16).value = 'Algorithm Parameter'
        ws.cell(content_header_row, 34).value = 'uArch Performance'
        ws.cell(content_header_row, 46).value = 'Simulator Performance'
        for h, w in zip([content_header_row, content_header_row, content_header_row], [16, 34, 46]):
            ws.cell(h, w).font = layer_style.title_font
            ws.cell(h, w).alignment = layer_style.center_align
        for w in range(8, 29):
            ws.cell(8, w).fill = layer_style.alg_pattern
        for w in range(29, 40):
            ws.cell(8, w).fill = layer_style.uarch_pattern
        for w in range(40, 51):
            ws.cell(8, w).fill = layer_style.sim_pattern
        items = sorted(sort_dict.items(), key=lambda x: x[1], reverse=True)
        top_k = 3
        index_list = items[:top_k]
        for h in index_list:
            for w in range(1, content_end_cols):
                ws.cell(h[0], w).fill = LayerStyle.yellow_light
        ws.cell(content_header_row, 42).alignment = layer_style.left_align
        for h in range(content_header_row + 2, len(df) + 2):
            g_cost_pos = content_end_cols - 9
            a_cost_pos = content_end_cols - 5
            ddr_pos = content_end_cols - 4
            lmem_pos = content_end_cols - 3
            if str(ws.cell(h, g_cost_pos).value).isnumeric():
                g_cost = float(ws.cell(h, g_cost_pos).value)
                a_cost = float(ws.cell(h, a_cost_pos).value)
                if a_cost - g_cost > g_cost * 0.1:
                    pass
        ddr_pos = content_end_cols - 4
        lmem_pos = content_end_cols - 3
        ws.cell(len(df) + 1, ddr_pos).fill = layer_style.red_light
        ws.cell(len(df) + 1, lmem_pos).fill = layer_style.red_light

        for col in df.columns:
            # auto set columns width
            index = list(df.columns).index(col)
            collen = 0
            letter = get_column_letter(index + 1)
            for h in range(1, len(df) + 8):
                if h > 30:
                    break
                if ws.cell(h, index + 1).value is not None:
                    if len(str(ws.cell(h, index + 1).value)) > 35:
                        continue
                    collen = max(collen, len(str(ws.cell(h, index + 1).value)))
            ws.column_dimensions[letter].width = collen * 1.05 if collen > 11 else 12
        ws.sheet_properties.tabColor = '0070C0'
        wb.save(file_path)
