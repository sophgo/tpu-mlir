#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
# @Time    : 2023/7/18 10:50
# @Author  : chongqing.zeng@sophgo.com
# @Project : PerfAI
import pandas as pd
from openpyxl import load_workbook


class InstrWorld(object):
    sheet_name = 'Instr World'

    def __init__(self, reg_list, columns, writer, split=False):
        """
        Initial an instr world object, equals to the instr world sheet in Excel.
        :param reg_list: the list containing all engine registers information
        :param columns: the union of all engine columns
        :param writer: the writer of output Excel to write
        :param split: whether generate a csv separately for instr world instead of putting it in Excel as a sheet
        """
        self.reg_list = reg_list
        self.columns = columns
        self.writer = writer
        self.split = split
        self.index = 1

    def write(self, out_file):
        """
        Write instr world information to Excel or CSV.
        :param out_file: the output Excel file path
        :return: None
        """
        df = pd.DataFrame(self.reg_list, columns=self.columns, index=None)
        if self.split:
            out_file = out_file.replace('xlsx', "csv")
            df.to_csv(out_file, index=False)
        else:
            df.to_excel(self.writer, sheet_name=self.sheet_name, index=False, startrow=4, engine='xlsxwriter', float_format='%g')

    @classmethod
    def set_style(cls, out_file, frozen=True):
        """
        Set style for instr world sheet.
        :param out_file: Excel output path
        :param frozen: freeze title above the sheet
        :return:
        """
        wb = load_workbook(out_file)
        ws = wb[cls.sheet_name]
        if frozen:
            _cell = ws.cell(6, 1)
            ws.freeze_panes = _cell
        wb.save(out_file)
