#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
# @Time    : 2023/7/18 10:39
# @Author  : chongqing.zeng@sophgo.com
# @Project : PerfAI
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from definition.bm1684x_defs import data_type_dict, tiu_func_name_dict
from utils.utils import *
from definition.style import *


class TiuNode:
    def __init__(self, reg):
        self.alg_ops = int(reg['Alg Ops'])
        self.uarch_ops = int(reg['uArch Ops'])
        self.alg_cycle = int(reg['Alg Cycle'])
        self.cycle = int(reg['Asic Cycle'])
        self.des_tsk_typ = int(reg['des_tsk_typ'])
        self.des_opd1_h = reg['des_opd1_h'] if 'des_opd1_h' in reg.keys() else None
        self.des_opd1_w = reg['des_opd1_w'] if 'des_opd1_w' in reg.keys() else None
        self.des_opd1_h_str = reg['des_opd1_h_str'] if 'des_opd1_h_str' in reg.keys() else None
        self.des_opd1_w_str = reg['des_opd1_w_str'] if 'des_opd1_w_str' in reg.keys() else None


class Tiu(object):

    def __init__(self, core_id, writer):
        """
        Initial Tiu object, equals to a tiu sheet in Excel.
        :param core_id: the id of current core
        :param writer: the writer of Excel to write
        """
        self.writer = writer
        self.columns = ['Engine Id', 'Core Id', 'Cmd Id', 'Layer Id', 'Layer Name', 'Function Type', 'Function Name',
                        'Alg Cycle', 'Asic Cycle', 'Start Cycle', 'End Cycle', 'Avg Cycle Last 200', 'Alg Ops',
                        'uArch Ops', 'uArch Rate', 'Bank Conflict Ratio',
                        'Initial Cycle Ratio', 'Data Type', 'Sim Power(W)', 'des_cmd_id_dep',
                        'des_res0_n', 'des_res0_c', 'des_res0_h', 'des_res0_w',
                        'des_res0_n_str', 'des_res0_c_str', 'des_res0_h_str', 'des_res0_w_str',
                        'des_opd0_n', 'des_opd0_c', 'des_opd0_h', 'des_opd0_w',
                        'des_opd0_n_str', 'des_opd0_c_str', 'des_opd0_h_str', 'des_opd0_w_str',
                        'des_opd1_n', 'des_opd1_c', 'des_opd1_h', 'des_opd1_w', 'des_opd1_n_str', 'des_opd1_c_str', 'des_opd1_h_str',
                        'des_opd1_w_str',
                        'des_opd2_n_str', 'des_res0_addr', 'des_res1_addr', 'des_opd0_addr', 'des_opd1_addr',
                        'des_opd2_addr',
                        'des_opd3_addr', 'des_tsk_typ', 'des_tsk_eu_typ', 'des_cmd_short',
                        'des_opt_res0_prec', 'des_opt_opd0_prec', 'des_opt_opd1_prec', 'des_opt_opd2_prec',
                        'des_short_opd0_str',
                        'des_opt_opd0_const', 'des_opt_opd1_const', 'des_opt_opd2_const', 'des_opt_opd3_const',
                        'des_opt_opd4_const', 'des_opt_opd5_const',
                        'des_opt_res_add', 'des_opt_res0_sign', 'des_opt_opd0_sign', 'des_opt_opd1_sign',
                        'des_opt_opd2_sign',
                        'des_opd0_rt_pad', 'des_opd1_x_ins0', 'des_opd0_up_pad', 'des_opd0_lf_pad', 'des_opt_left_tran',
                        'des_pad_mode', 'des_opd0_y_ins0', 'des_opd1_y_ins0',
                        'des_short_res0_str', 'des_short_opd1_str', 'des_sym_range', 'des_opt_rq', 'des_op_code',
                        'des_opt_kernel_rotate', 'des_res_op_x_str', 'des_res_op_y_str', 'des_opd0_x_ins0',
                        'des_tsk_opd_num', 'des_opd0_dn_pad', 'des_intr_en', 'des_opt_relu', 'des_pwr_step', 'Msg Id', 'Sd\Wt Count']
        self.reg_list = []
        self.perf_dict = dict()
        self.stati_list = []
        self.core_id = str(core_id)
        self.total_instr = 0
        self.sheet_name = 'TIU_' + str(core_id)
        self.height = None
        self.width = len(self.columns)
        self.sheet_color = '008000'
        # The architecture parameters of the SG2260 are set to the initial values
        # It will be changed as the chip architecture parameters change
        self.detail_spec = {
                            'Platform': ['simulator'],
                            'CHIP ARCH': ['sg2260'],
                            'Core Num': ['64'],
                            'NPU Num': ['64'],
                            'Cube IC Align(8bits)': ['32'],
                            'Cube OHOW Align': ['8'],
                            'Vector OHOW Align(8bits)': ['128'],
                            'Tiu Frequency(MHz)': ['1000'],
                            'DMA Frequency(MHz)': ['1000'],
                            'Dram Bandwidth': ['8533'],
                            'TPU Lmem Size(MiB)': ['16777216']}
        self.kpi_desc = pd.DataFrame({'Field': [
            'uArch Rate',
            'Bank Conflict Ratio',
            'Initial Cycle Ratio',
            'Data Type',
            'Avg Cycle Last 200'],
            'Description': [
                'Alg Ops / uArch Ops, since the shape of tensor needs to be aligned in '
                'micro-architecture, the actual ops will be greater than the algorithm value. '
                "It's better to closer to 100%.",
                'opd num with same address / opd num, probability of bank conflict '
                "occurred. It's better to closer to 0%.",
                'initial cycle / total cycle, proportion of initial cycle in the total '
                'cycles, the meaning of initial cycle is that some initialization and '
                "other operations will be performed before execution. It's better to closer to 0%.",
                'Data type transform, usually represents opd0->res0.',
                'The average cycle taken to execute the last 200 tiu commands. Consecutive short '
                'commands will increase the burden on PMU, leading to a decrease in performance']}, index=None)
        self.summary = dict()
        self.tiu_cycle = 0
        self.start_time = sys.maxsize
        self.end_time = 0
        self.tiu_time = 0
        self.alg_total_cycle = 0
        self.alg_total_ops = 0
        self.uArch_total_ops = 0
        self.total_power = 0.0
        self.wait_msg_time = 0
        self.chip_arch_dict = None

    def load(self, reg_info_file, tiu_layer_map):
        """
        Load data from external file.
        :param tiu_layer_map:
        :param reg_info_file: file records register information, usually obtained by TPUPerf
        :return: None
        """
        chip_arch_dict = None
        if os.path.exists(reg_info_file) and os.path.getsize(reg_info_file) != 0:
            last_underscore_index = reg_info_file.rfind('_')
            core_id = int(reg_info_file[last_underscore_index + 1 : -4])
            with open(reg_info_file) as f:
                rows = f.readlines()
                field_set = set()
                reg_count = 0
                for row in rows:
                    if "__TIU_REG_INFO__" in row:
                        reg_count += 1
                    if "\t" in row and reg_count > 0:
                        attr = row.split(': ')[0][1:]
                        field_set.add(attr)
                reg_count = 0
                field_list = list(field_set) if len(field_set) >= len(self.columns) else self.columns
                reg_dict = dict.fromkeys(field_list, '')
                idx = 0
                for row in rows:
                    if "__CHIP_ARCH_ARGS__" in row:
                        chip_arch_dict = dict()
                    if "__TIU_REG_INFO__" in row:
                        reg_count += 1
                        if idx != 0:
                            k = int(reg_dict['Cmd Id'])
                            if any(key[0] == k for key in tiu_layer_map.keys()):
                                layer_id_name = tiu_layer_map[(k, core_id)]
                            else:
                                layer_id_name = ['-', '-']
                            reg_dict['Layer Id'] = layer_id_name[0]
                            reg_dict['Layer Name'] = layer_id_name[1]
                            self.reg_list.append(reg_dict)
                        reg_dict = dict.fromkeys(field_list, '')
                    elif "\t" not in row:
                        reg_dict['Function Type'] = row[:-2]
                    elif reg_count == 0:
                        fields = row.split(': ')
                        attr = fields[0][1:]
                        val = fields[1][:-1]
                        chip_arch_dict[attr] = val
                        idx = 0
                    else:
                        fields = row.split(': ')
                        attr = fields[0][1:]
                        val = fields[1][:-1]
                        val = int(val) if val.isnumeric() else val
                        reg_dict[attr] = val
                        idx += 1
                if 'Platform' not in chip_arch_dict.keys():
                    chip_arch_dict['Platform'] = 'pmu'
                if idx != 0:
                    k = int(reg_dict['Cmd Id'])
                    if any(key[0] == k for key in tiu_layer_map.keys()):
                        layer_id_name = tiu_layer_map[(k, core_id)]
                    else:
                        layer_id_name = ['-', '-']
                    reg_dict['Layer Id'] = layer_id_name[0]
                    reg_dict['Layer Name'] = layer_id_name[1]
                self.reg_list.append(reg_dict)
                self.detail_spec = {
                    'Platform': [chip_arch_dict['Platform']],
                    'CHIP ARCH': [chip_arch_dict['Chip Arch']],
                    'Core Num': [chip_arch_dict['Core Num']],
                    'NPU Num': [chip_arch_dict['NPU Num']],
                    'Cube IC Align(8bits)': [chip_arch_dict['Cube IC Align(8bits)']],
                    'Cube OHOW Align': [chip_arch_dict['Cube OHOW Align(8bits)']],
                    'Vector OHOW Align(8bits)': [chip_arch_dict['Vector OHOW Align(8bits)']],
                    'TIU Frequency(MHz)': [chip_arch_dict['TIU Frequency(MHz)']],
                    'DMA Frequency(MHz)': [chip_arch_dict['DMA Frequency(MHz)']],
                    'DDR Frequency(GHz)': [chip_arch_dict['DDR Frequency(GHz)']],
                    'TPU Lmem Size(MiB)': [chip_arch_dict['TPU Lmem Size(MiB)']]}
        self.height = len(self.reg_list)
        self.chip_arch_dict = chip_arch_dict
        return chip_arch_dict

    def add_kpi_field(self):
        """
        Add some indicators which are convenient for performance analysis artificially.
        :return: None
        """
        for i in range(len(self.reg_list)):
            reg_dict = self.reg_list[i]
            continous_gap = 200
            if i < continous_gap - 1 :
                reg_dict['Avg Cycle Last 200'] = round(int(reg_dict['End Cycle']) / int(reg_dict['Cmd Id']))
            else:
                reg_dict['Avg Cycle Last 200'] = round((reg_dict['End Cycle'] - self.reg_list[i-199]['Start Cycle']) / continous_gap)
            if not (reg_dict['des_tsk_typ'] == 15 and reg_dict['des_tsk_eu_typ'] == 9):
                # wait msg time do not add to tiu cycles
                self.tiu_cycle += int(reg_dict['Asic Cycle'])
            self.alg_total_cycle += int(reg_dict['Alg Cycle'])
            self.alg_total_ops += int(reg_dict['Alg Ops'])
            self.uArch_total_ops += int(reg_dict['uArch Ops'])
            if isinstance(reg_dict['des_opt_opd0_prec'], int):
                reg_dict['Data Type'] = data_type_dict[reg_dict['des_opt_opd0_prec']] + \
                                        ' -> ' + data_type_dict[reg_dict['des_opt_res0_prec']]
            else:
                reg_dict['Data Type'] = data_type_dict[reg_dict['des_opt_res0_prec']] + \
                                        ' -> ' + data_type_dict[reg_dict['des_opt_res0_prec']]
            if int(reg_dict['des_tsk_typ']) == 15 and int(reg_dict['des_tsk_eu_typ']) == 9:
                self.wait_msg_time += int(reg_dict['Asic Cycle'])
            self.start_time = min(self.start_time, get_time_by_cycle(reg_dict['Start Cycle'], self.chip_arch_dict['TIU Frequency(MHz)']))
            self.end_time = max(self.start_time, get_time_by_cycle(reg_dict['End Cycle'], self.chip_arch_dict['TIU Frequency(MHz)']))
            self.total_power += float(reg_dict['Sim Power(W)'])
            if reg_dict['Function Type'] not in self.perf_dict.keys():
                func_dict = {
                    'Function Name': reg_dict['Function Type'],
                    'Instr Num': 1,
                    'Alg Ops': int(reg_dict['Alg Ops']),
                    'Alg Ops Ratio': 0,
                    'Alg Cycle': int(reg_dict['Alg Cycle']),
                    'Alg Cycle Ratio': 0,
                    'uArch Ops': int(reg_dict['uArch Ops']),
                    'uArch URate': 0,
                    'uArch Ops Ratio': 0,
                    'Asic Cycle': int(reg_dict['Asic Cycle']),
                    'Asic Cycle Ratio': 0,
                    'Sim Power': float(reg_dict['Sim Power(W)']),
                    'Sim Power Ratio': 0
                }
                self.perf_dict[reg_dict['Function Type']] = func_dict
            else:
                func_dict = self.perf_dict[reg_dict['Function Type']]
                func_dict['Instr Num'] += 1
                func_dict['Alg Ops'] += int(reg_dict['Alg Ops'])
                func_dict['uArch Ops'] += int(reg_dict['uArch Ops'])
                func_dict['Alg Cycle'] += int(reg_dict['Alg Cycle'])
                func_dict['Asic Cycle'] += int(reg_dict['Asic Cycle'])
                func_dict['Sim Power'] += float(reg_dict['Sim Power(W)'])
                self.perf_dict[reg_dict['Function Type']] = func_dict
            self.total_instr += 1
        self.tiu_time = get_time_by_cycle(self.tiu_cycle, self.chip_arch_dict['TIU Frequency(MHz)']) if self.chip_arch_dict else 0

    def pop_data(self, core_id):
        tiu_instance_map = dict()
        for reg in self.reg_list:
            tiu_instance_map[(int(reg['Cmd Id']), core_id)] = TiuNode(reg)
        return tiu_instance_map

    def write(self):
        """
        Write register information and kpi field to Excel.
        :return: None
        """
        if len(self.reg_list) <= 0:
            return
        df = pd.DataFrame(self.reg_list)
        new_cols = []
        for col in self.columns:
            if col in df.columns.values.tolist():
                new_cols.append(col)
        self.columns = new_cols
        df = df[self.columns]
        for col in self.columns:
            if 'addr' in col or 'mask' in col:
                df[col] = int2Hex(df[col].values)
        self.summary = {
            'totalTiuCycle': [self.tiu_cycle],
            'totalAlgCycle': [self.alg_total_cycle],
            'algTotalOps': [self.alg_total_ops],
            'totalUArchOps': [self.uArch_total_ops],
            'uArchURate': [get_ratio_str_2f_zero(self.alg_total_ops, self.uArch_total_ops)],
            'totalTiuPower(W)': [self.total_power],
            'waitMsgTotalTime': [self.wait_msg_time]
        }
        for func in self.perf_dict.keys():
            tmp_func_dict = self.perf_dict[func]
            tmp_func_dict['Function Name'] = func
            tmp_func_dict['Alg Ops Ratio'] = get_ratio_str_2f_zero(tmp_func_dict['Alg Ops'], self.alg_total_ops)
            tmp_func_dict['Alg Cycle Ratio'] = get_ratio_str_2f_zero(tmp_func_dict['Alg Cycle'], self.alg_total_cycle)
            tmp_func_dict['uArch URate'] = get_ratio_str_2f_zero(tmp_func_dict['Alg Ops'], tmp_func_dict['uArch Ops'])
            tmp_func_dict['uArch Ops Ratio'] = get_ratio_str_2f_zero(tmp_func_dict['uArch Ops'], self.uArch_total_ops)
            tmp_func_dict['Asic Cycle Ratio'] = get_ratio_str_2f_zero(tmp_func_dict['Asic Cycle'], self.tiu_cycle)
            tmp_func_dict['Sim Power Ratio'] = get_ratio_str_2f_zero_f(tmp_func_dict['Sim Power'], self.total_power)
            self.perf_dict[func] = tmp_func_dict
            self.stati_list.append(tmp_func_dict)
        self.perf_dict['Overall'] = {
            'Function Name': 'Overall',
            'Instr Num': self.total_instr,
            'Alg Ops': self.alg_total_ops,
            'Alg Ops Ratio': '100.00%',
            'Alg Cycle': self.alg_total_cycle,
            'Alg Cycle Ratio': '100.00%',
            'uArch Ops': self.uArch_total_ops,
            'uArch URate': get_ratio_str_2f_zero(self.alg_total_ops, self.uArch_total_ops),
            'uArch Ops Ratio': '100.00%',
            'Sim Power': self.total_power,
            'Sim Power Ratio': '100.00%',
            'Asic Cycle': self.tiu_cycle,
            'Asic Cycle Ratio': '100.00%'
        }
        self.stati_list.append(self.perf_dict['Overall'])
        if self.chip_arch_dict['Platform'].lower() == 'simulator':
            for d in self.stati_list:
                d['Simulator Cycle'] = d.pop('Asic Cycle')
                d['Simulator Cycle Ratio'] = d.pop('Asic Cycle Ratio')
            df.rename(columns={'Asic Cycle': 'Simulator Cycle'}, inplace=True)
        if len(df) > 0:
            content_start_rows = len(self.stati_list) + 8
            pd.DataFrame(self.summary).to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=0,
                                                startcol=1,
                                                engine='xlsxwriter', float_format='%g')
            pd.DataFrame(self.detail_spec).to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=0,
                                                    startcol=15,
                                                    engine='xlsxwriter', float_format='%g')
            pd.DataFrame(self.kpi_desc).to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=3,
                                                 startcol=15,
                                                 engine='xlsxwriter', float_format='%g')
            pd.DataFrame(self.stati_list).to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=5,
                                                   startcol=1,
                                                   engine='xlsxwriter', float_format='%g')
            pd.DataFrame(df).to_excel(self.writer, index=False, sheet_name=self.sheet_name,
                                      startrow=content_start_rows, startcol=0,
                                      engine='xlsxwriter', float_format='%g')

    @classmethod
    def set_style(cls, file_path, core_id, frozen=False):
        """
        Set style for Excel, and highlight some fields we care about.
        :param frozen: freeze some statistics above the sheet
        :param file_path: Excel output path
        :param core_id: the id of current core
        :return: None
        """
        wb = load_workbook(file_path)
        detail_style = DetailsStyle()
        sheet_name = 'TIU_' + str(core_id)
        if sheet_name not in wb.sheetnames:
            return
        df = pd.read_excel(file_path, sheet_name)
        ws = wb[sheet_name]
        ws.cell(1, 1).value = 'Performance'
        ws.cell(2, 1).value = 'Summary'
        ws.cell(1, 15).value = 'Detail Spec'
        ws.cell(6, 1).value = 'Statistics'
        ws.cell(4, 15).value = 'Description'
        for h, w in zip([1, 2, 1, 6, 4], [1, 1, 15, 1, 15]):
            ws.cell(h, w).fill = detail_style.title_pattern
            ws.cell(h, w).font = detail_style.title_font
        summary_start_cols = 2
        summary_end_cols = summary_start_cols + 7
        for w in range(summary_start_cols, summary_end_cols):
            # summary title style
            ws.cell(1, w).fill = detail_style.title_header_pattern
            ws.cell(1, w).font = detail_style.title_header_font
            # summary content style
            ws.cell(2, w).fill = detail_style.title_content_pattern
            ws.cell(2, w).font = detail_style.title_font
        detail_start_cols = summary_end_cols + 7
        detail_end_cols = detail_start_cols + 11
        for w in range(detail_start_cols, detail_end_cols):
            # detail title style
            ws.cell(1, w).fill = detail_style.title_header_pattern
            ws.cell(1, w).font = detail_style.title_header_font
            # detail content style
            ws.cell(2, w).fill = detail_style.title_content_pattern
            ws.cell(2, w).font = detail_style.title_font
        perf_start_cols = summary_start_cols
        perf_end_cols = perf_start_cols + 13
        perf_start_rows = 6
        perf_df_len = df.query('totalTiuCycle == "Core Id"').index[0] - perf_start_rows - 1
        for w in range(perf_start_cols, perf_end_cols):
            # perf title style
            ws.cell(perf_start_rows, w).fill = detail_style.title_header_pattern
            ws.cell(perf_start_rows, w).font = detail_style.title_header_font
            # perf content style
            for h in range(7, 7 + perf_df_len):
                ws.cell(h, w).fill = detail_style.title_content_pattern
                ws.cell(h, w).font = detail_style.title_font
        tiuDescStartIndex = detail_start_cols
        tiuDescEndIndex = tiuDescStartIndex + 2
        for w in range(detail_start_cols, tiuDescEndIndex):
            # desc title style
            ws.cell(4, w).fill = detail_style.title_header_pattern
            ws.cell(4, w).font = detail_style.title_header_font
            # desc content style
            for h in range(5, 10):
                ws.cell(h, w).fill = detail_style.title_content_pattern
                ws.cell(h, w).font = detail_style.title_font
        rangeStr = str(perf_df_len + 9) + ':' + str(perf_df_len + 9)
        for cell in ws[rangeStr]:
            # content title style
            cell.fill = detail_style.content_pattern
            cell.font = detail_style.title_header_font
        # set content style
        content_end_cols = 19
        content_start_rows = 8 + perf_df_len
        initial_cycle_pos = 17
        uArch_ratio_pos = 15
        bank_conflict_pos = 16
        fp32_pos = 18
        avg_cycle_last_200_pos = 12
        for h in range(content_start_rows + 2, len(df) + 2):
            for w in range(1, content_end_cols + 1):
                # set key content border style
                ws.cell(h, w).fill = detail_style.key_content_pattern
                ws.cell(h, w).border = detail_style.border
                ws.cell(h, w).alignment = detail_style.right_align
                ws.cell(h, w).font = detail_style.title_font
        for h in range(content_start_rows, len(df)):
            # ------------------ set highlight style --------------------
            # initial cycle ratio
            if float(ws.cell(h + 2, initial_cycle_pos).value[:-1]) > 1:
                ws.cell(h + 2, initial_cycle_pos).fill = detail_style.yellow_light
            # uArch rate
            if float(ws.cell(h + 2, uArch_ratio_pos).value[:-1]) < 50:
                ws.cell(h + 2, uArch_ratio_pos).fill = detail_style.red_light
            elif float(ws.cell(h + 2, uArch_ratio_pos).value[:-1]) < 75:
                ws.cell(h + 2, uArch_ratio_pos).fill = detail_style.yellow_light
            # bank conflicted
            if float(ws.cell(h + 2, bank_conflict_pos).value[:-1]) > 0:
                ws.cell(h + 2, bank_conflict_pos).fill = detail_style.yellow_light
            # FP32
            if 'fp32' in ws.cell(h + 2, fp32_pos).value.lower():
                ws.cell(h + 2, fp32_pos).fill = detail_style.red_light
            # Avg Cycle Last 200
            if float(ws.cell(h + 2, avg_cycle_last_200_pos).value) < 200:
                ws.cell(h + 2, avg_cycle_last_200_pos).fill = detail_style.red_light

        # set tiu description content and style
        ws.cell(content_start_rows, initial_cycle_pos).value = 'Ratio>1%'
        ws.cell(content_start_rows - 1, uArch_ratio_pos).value = 'Urate<75%'
        ws.cell(content_start_rows, uArch_ratio_pos).value = 'Urate<50%'
        ws.cell(content_start_rows, bank_conflict_pos).value = 'Ratio>0'
        ws.cell(content_start_rows, fp32_pos).value = 'IsFP32'
        ws.cell(content_start_rows, avg_cycle_last_200_pos).value = 'AvgCycle<200'

        for h, w in zip([content_start_rows, content_start_rows - 1, content_start_rows],
                        [bank_conflict_pos, uArch_ratio_pos, initial_cycle_pos]):
            ws.cell(h, w).fill = detail_style.yellow_light
            ws.cell(h, w).font = detail_style.title_font
            ws.cell(h, w).alignment = detail_style.center_align
        for h, w in zip([content_start_rows, content_start_rows, content_start_rows], [uArch_ratio_pos, fp32_pos, avg_cycle_last_200_pos]):
            ws.cell(h, w).fill = detail_style.red_light
            ws.cell(h, w).font = detail_style.title_font
            ws.cell(h, w).alignment = detail_style.center_align
        for start_rows, end_rows, start_cols, end_cols in zip([1, 1, 6, 4],
                                                              [3, 3, 7 + perf_df_len, 10],
                                                              [2, 16, 2, 16], [9, 27, 15, 18]):
            for h in range(start_rows, end_rows):
                for w in range(start_cols, end_cols):
                    # set all header border style
                    ws.cell(h, w).border = detail_style.border
                    if h > 4 and w > 14:
                        ws.cell(h, w).alignment = detail_style.left_align
                    else:
                        ws.cell(h, w).alignment = detail_style.center_align
        for col in df.columns:
            # auto set columns width
            index = list(df.columns).index(col)
            collen = 0
            letter = get_column_letter(index + 1)
            if index > content_end_cols + 6:
                collen = len(str(ws.cell(content_start_rows + 1, index + 1).value))
            else:
                for h in range(1, len(df) + 8):
                    if h > 30:
                        break
                    if ws.cell(h, index + 1).value is not None:
                        if len(str(ws.cell(h, index + 1).value)) > 35:
                            continue
                        collen = max(collen, len(str(ws.cell(h, index + 1).value)))
            ws.column_dimensions[letter].width = collen * 1.05
        desc_start_rows, desc_end_rows = 5, 9
        for row in range(desc_start_rows, desc_end_rows):
            ws.merge_cells(start_row=row, end_row=row, start_column=17, end_column=30)
        ws.sheet_properties.tabColor = '008000'
        if frozen:
            _cell = ws.cell(content_start_rows + 2, 1)
            ws.freeze_panes = _cell
        wb.save(file_path)
