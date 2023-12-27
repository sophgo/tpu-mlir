#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/7/18 10:50
# @Author  : chongqing.zeng@sophgo.com
# @Project : PerfAI
import pandas as pd
from decimal import Decimal
from numpy import transpose
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math

from definition.style import DetailsStyle


class SimulatorSummary(object):
    sheet_name = 'Simulator Summary'

    def __init__(self, writer, tius, gdmas, sdmas, cdmas, act_core_num):
        """
        Initial a Summary object contains summary information of all engines such as tiu, gdma, sdma, etc.
        Equals to summary sheet in output Excel.
        :param writer: the writer of Excel to writer
        :param tius: tiu instance list
        :param gdmas: gdma instance list
        :param sdmas: sdma instance list, it may be empty
        :param act_core_num: represents the actual number of cores to run since the input file may be empty
        """
        self.columns = ['CoreId', 'TiuWorkingRatio', 'Parallelism(%)', 'Concurrency(%)',
                        'simTotalCycle', 'simTiuCycle', 'totalAlgCycle', 'totalAlgOps', 'totalUArchOps', 'uArchURate',
                        'simGdmaCycle', 'totalDdrDataSize', 'totalL2DataSize', 'ddrAvgBandwidth', 'l2AvgBandwidth', 'avgDdrBurstLength',
                        'simSdmaCycle', 'totalDdrDataSize', 'ddrAvgBandwidth', 'avgDdrBurstLength',
                        'simCdmaCycle', 'totalDdrDataSize', 'totalL2DataSize', 'ddrAvgBandwidth', 'l2AvgBandwidth', 'avgDdrBurstLength']
        self.writer = writer
        self.tius = tius
        self.gdmas = gdmas
        self.sdmas = sdmas
        self.cdmas = cdmas
        self.act_core_num = act_core_num
        self.sheet_color = ''
        self.data = []

    def load(self, chip_arch):
        """
        Compute all the values we care about
        :return: None
        """
        core_ids, tiu_work_ratios, prallelisms, concurrencys, sim_cycles, tiu_cycles, alg_cycles, alg_opss, uArch_opss, uArch_rates = \
            [], [], [], [], [], [], [], [], [], []
        gdma_cycles, gdma_ddr_datasizes, gdma_l2_datasizes, gdma_ddr_avg_bds, \
            gdma_l2_avg_bds, gdma_ddr_avg_bls = [], [], [], [], [], []
        sdma_cycles, sdma_ddr_datasizes, sdma_ddr_avg_bds, sdma_ddr_avg_bls = [], [], [], []
        cdma_cycles, cdma_ddr_datasizes, cdma_l2_datasizes, cdma_ddr_avg_bds, \
            cdma_l2_avg_bds, cdma_ddr_avg_bls = [], [], [], [], [], []
        gdma_ddr_cycles, gdma_l2_cycles, sdma_ddr_cycles, sdma_l2_cycles, cdma_ddr_cycles, cdma_l2_cycles = 0, 0, 0, 0, 0, 0
        gdma_bl_sum, gdma_xact_sum, sdma_bl_sum, sdma_xact_sum, cdma_bl_sum, cdma_xact_sum = 0, 0, 0, 0, 0, 0
        frequency = int(chip_arch['Frequency(MHz)'])
        for core_id in range(0, self.act_core_num):
            core_ids.append(core_id)
            tiu_work_ratios.append('0.00%' if self.tius[core_id].sim_total_cycle == 0 else str(
                Decimal((self.tius[core_id].sim_cycle * 100 / self.tius[core_id].sim_total_cycle))
                .quantize(Decimal("0.00"))) + '%')
            if self.act_core_num > 1:
                prallelisms.append('0.00%' if self.tius[core_id].sim_total_cycle == 0 else
                                        str((Decimal(((self.tius[core_id].sim_cycle + self.gdmas[core_id].sim_cycle\
                                                    + self.sdmas[core_id].sim_cycle) / self.tius[core_id].sim_total_cycle)* 100))\
                                                    .quantize(Decimal("0.00"))) + '%')
            else:
                prallelisms.append('0.00%' if self.tius[core_id].sim_total_cycle == 0 else
                                        str((Decimal(((self.tius[core_id].sim_cycle + self.gdmas[core_id].sim_cycle)\
                                                      / self.tius[core_id].sim_total_cycle) * 100)).quantize(Decimal("0.00"))) + '%')
            if self.tius[core_id].sim_cycle > 0 and self.gdmas[core_id].sim_cycle > 0:
                concurrencys.append(str(Decimal((self.tius[core_id].sim_cycle + self.gdmas[core_id].sim_cycle - self.tius[core_id].sim_total_cycle) /
                                                min(self.tius[core_id].sim_cycle, self.gdmas[core_id].sim_cycle) * 100).quantize(Decimal("0.00"))) + '%')
            else:
                concurrencys.append('0.00%')
            sim_cycles.append(self.tius[core_id].sim_total_cycle)
            tiu_cycles.append(self.tius[core_id].sim_cycle)
            alg_cycles.append(self.tius[core_id].alg_total_cycle)
            alg_opss.append(self.tius[core_id].alg_total_ops)
            uArch_opss.append(self.tius[core_id].uArch_total_ops)
            uArch_rates.append('0.00%' if self.tius[core_id].uArch_total_ops == 0 else str(
                Decimal((self.tius[core_id].alg_total_ops * 100 / self.tius[core_id].uArch_total_ops)).
                quantize(Decimal("0.00"))) + '%')

            gdma_cycles.append(self.gdmas[core_id].sim_cycle)
            gdma_ddr_datasizes.append(self.gdmas[core_id].ddr_total_datasize)
            gdma_l2_datasizes.append(self.gdmas[core_id].l2_total_datasize)
            gdma_ddr_avg_bds.append(0 if self.gdmas[core_id].ddr_total_cycle == 0 else str(
                Decimal((self.gdmas[core_id].ddr_total_datasize / self.gdmas[core_id].ddr_total_cycle * frequency / 1000)).
                quantize(Decimal("0.00"))))
            gdma_l2_avg_bds.append(0 if self.gdmas[core_id].l2_total_cycle == 0 else str(
                Decimal((self.gdmas[core_id].l2_total_datasize / self.gdmas[core_id].l2_total_cycle * frequency / 1000)).
                quantize(Decimal("0.00"))))
            gdma_ddr_avg_bls.append(0 if self.gdmas[core_id].ddr_xact_cnt == 0 else str(
                Decimal((self.gdmas[core_id].ddr_burst_length_sum / self.gdmas[core_id].ddr_xact_cnt)).
                quantize(Decimal("0.00"))))
            gdma_ddr_cycles += self.gdmas[core_id].ddr_total_cycle
            gdma_l2_cycles += self.gdmas[core_id].l2_total_cycle
            gdma_bl_sum += self.gdmas[core_id].ddr_burst_length_sum
            gdma_xact_sum += self.gdmas[core_id].ddr_xact_cnt

            sdma_cycles.append(self.sdmas[core_id].sim_cycle)
            sdma_ddr_datasizes.append(self.sdmas[core_id].ddr_total_datasize)
            sdma_ddr_avg_bds.append(0 if self.sdmas[core_id].ddr_total_cycle == 0 else str(
                Decimal((self.sdmas[core_id].ddr_total_datasize / self.sdmas[core_id].ddr_total_cycle * frequency / 1000)).
                quantize(Decimal("0.00"))))
            sdma_ddr_avg_bls.append(0 if self.sdmas[core_id].ddr_xact_cnt == 0 else str(
                Decimal((self.sdmas[core_id].ddr_burst_length_sum / self.sdmas[core_id].ddr_xact_cnt)).
                quantize(Decimal("0.00"))))
            sdma_ddr_cycles += self.sdmas[core_id].ddr_total_cycle
            sdma_l2_cycles += self.sdmas[core_id].l2_total_cycle
            sdma_bl_sum += self.sdmas[core_id].ddr_burst_length_sum
            sdma_xact_sum += self.sdmas[core_id].ddr_xact_cnt

            cdma_cycles.append(self.cdmas[core_id].sim_cycle)
            cdma_ddr_datasizes.append(self.cdmas[core_id].ddr_total_datasize)
            cdma_l2_datasizes.append(self.cdmas[core_id].l2_total_datasize)
            cdma_ddr_avg_bds.append(0 if self.cdmas[core_id].ddr_total_cycle == 0 else str(
                Decimal((self.cdmas[core_id].ddr_total_datasize / self.cdmas[core_id].ddr_total_cycle * frequency / 1000)).
                quantize(Decimal("0.00"))))
            cdma_l2_avg_bds.append(0 if self.cdmas[core_id].l2_total_cycle == 0 else str(
                Decimal((self.cdmas[core_id].l2_total_datasize / self.cdmas[core_id].l2_total_cycle * frequency / 1000)).
                quantize(Decimal("0.00"))))
            cdma_ddr_avg_bls.append(0 if self.cdmas[core_id].ddr_xact_cnt == 0 else str(
                Decimal((self.cdmas[core_id].ddr_burst_length_sum / self.cdmas[core_id].ddr_xact_cnt)).
                quantize(Decimal("0.00"))))
            cdma_ddr_cycles += self.cdmas[core_id].ddr_total_cycle
            cdma_l2_cycles += self.cdmas[core_id].l2_total_cycle
            cdma_bl_sum += self.cdmas[core_id].ddr_burst_length_sum
            cdma_xact_sum += self.cdmas[core_id].ddr_xact_cnt
        core_ids.append('Overall')
        tiu_work_ratios.append('0.00%' if sum(sim_cycles) == 0 else
                               str((Decimal((sum(tiu_cycles) / sum(sim_cycles)) * 100)).quantize(
                                   Decimal("0.00"))) + '%')
        if self.act_core_num > 1:
            prallelisms.append('0.00%' if sum(sim_cycles) == 0 else
                                    str((Decimal((max(tiu_cycles) + max(gdma_cycles)\
                                                + max(sdma_cycles)) / max(sim_cycles) * 100))\
                                                .quantize(Decimal("0.00"))) + '%')
        else:
            prallelisms.append('0.00%' if sum(sim_cycles) == 0 else
                                    str((Decimal((max(tiu_cycles) + max(gdma_cycles)) / max(sim_cycles) * 100))\
                                                .quantize(Decimal("0.00"))) + '%')
        if max(tiu_cycles) > 0 and max(gdma_cycles) > 0:
            concurrencys.append(str(Decimal((max(tiu_cycles) + max(gdma_cycles) - max(sim_cycles)) /
                                            min(max(tiu_cycles), max(gdma_cycles)) * 100).quantize(Decimal("0.00"))) + '%')
        else:
            concurrencys.append('0.00%')
        uArch_rates.append('0.00%' if sum(uArch_opss) == 0 else
                    str((Decimal((sum(alg_opss) / sum(uArch_opss)) * 100)).quantize(
                        Decimal("0.00"))) + '%')
        sim_cycles.append(max(sim_cycles))
        tiu_cycles.append(max(tiu_cycles))
        alg_cycles.append(max(alg_cycles))
        alg_opss.append(sum(alg_opss))
        uArch_opss.append(sum(uArch_opss))
        gdma_cycles.append(max(gdma_cycles))
        gdma_ddr_datasizes.append(sum(gdma_ddr_datasizes))
        gdma_l2_datasizes.append(sum(gdma_l2_datasizes))
        gdma_ddr_avg_bds.append(0 if gdma_ddr_cycles == 0 else str(
            Decimal((gdma_ddr_datasizes[-1] / gdma_ddr_cycles) * frequency / 1000).
            quantize(Decimal("0.00"))))
        gdma_l2_avg_bds.append(0 if gdma_l2_cycles == 0 else str(
            Decimal((gdma_l2_datasizes[-1] / gdma_l2_cycles * frequency / 1000)).
            quantize(Decimal("0.00"))))
        gdma_ddr_avg_bls.append(0 if gdma_xact_sum == 0 else str(
            Decimal((gdma_bl_sum / gdma_xact_sum)).
            quantize(Decimal("0.00"))))
        sdma_cycles.append(max(sdma_cycles))
        sdma_ddr_datasizes.append(sum(sdma_ddr_datasizes))
        sdma_ddr_avg_bds.append(0 if sdma_ddr_cycles == 0 else str(
            Decimal((sdma_ddr_datasizes[-1] / sdma_ddr_cycles * frequency / 1000)).
            quantize(Decimal("0.00"))))
        sdma_ddr_avg_bls.append(0 if sdma_xact_sum == 0 else str(
            Decimal((sdma_bl_sum / sdma_xact_sum)).
            quantize(Decimal("0.00"))))
        cdma_cycles.append(max(cdma_cycles))
        cdma_ddr_datasizes.append(sum(cdma_ddr_datasizes))
        cdma_l2_datasizes.append(sum(cdma_l2_datasizes))
        cdma_ddr_avg_bds.append(0 if cdma_ddr_cycles == 0 else str(
            Decimal((cdma_ddr_datasizes[-1] / cdma_ddr_cycles * frequency / 1000)).
            quantize(Decimal("0.00"))))
        cdma_l2_avg_bds.append(0 if cdma_l2_cycles == 0 else str(
            Decimal((cdma_l2_datasizes[-1] / cdma_l2_cycles * frequency / 1000)).
            quantize(Decimal("0.00"))))
        cdma_ddr_avg_bls.append(0 if cdma_xact_sum == 0 else str(
            Decimal((cdma_bl_sum / cdma_xact_sum)).
            quantize(Decimal("0.00"))))
        frequency = int(chip_arch['Frequency(MHz)'])
        for idx in [len(sim_cycles) - 1]:
            sim_cycles[idx] = str(
                (Decimal(sim_cycles[idx] / frequency)).quantize(Decimal("0.00"))) + 'us'
            tiu_cycles[idx] = str(
                (Decimal(tiu_cycles[idx] / frequency)).quantize(Decimal("0.00"))) + 'us'
            alg_cycles[idx] = str(
                (Decimal(alg_cycles[idx] / frequency)).quantize(Decimal("0.00"))) + 'us'
            gdma_cycles[idx] = str(
                (Decimal(gdma_cycles[idx] / frequency)).quantize(Decimal("0.00"))) + 'us'
            sdma_cycles[idx] = str(
                (Decimal(sdma_cycles[idx] / frequency)).quantize(Decimal("0.00"))) + 'us'
            gdma_ddr_datasizes[idx] = str(
                (Decimal(gdma_ddr_datasizes[idx] / math.pow(2, 20))).quantize(Decimal("0.00"))) + 'MiB'
            gdma_l2_datasizes[idx] = str(
                (Decimal(gdma_l2_datasizes[idx] / math.pow(2, 20))).quantize(Decimal("0.00"))) + 'MiB'
            sdma_ddr_datasizes[idx] = str(
                (Decimal(sdma_ddr_datasizes[idx] / math.pow(2, 20))).quantize(Decimal("0.00"))) + 'MiB'
            cdma_cycles[idx] = str(
                (Decimal(cdma_cycles[idx] / frequency)).quantize(Decimal("0.00"))) + 'us'
            cdma_ddr_datasizes[idx] = str(
                (Decimal(cdma_ddr_datasizes[idx] / math.pow(2, 20))).quantize(Decimal("0.00"))) + 'MiB'
            cdma_l2_datasizes[idx] = str(
                (Decimal(cdma_l2_datasizes[idx] / math.pow(2, 20))).quantize(Decimal("0.00"))) + 'MiB'
        self.data = transpose([core_ids, tiu_work_ratios, prallelisms, concurrencys, sim_cycles, tiu_cycles, alg_cycles,
                            alg_opss, uArch_opss, uArch_rates, gdma_cycles,
                            gdma_ddr_datasizes, gdma_l2_datasizes, gdma_ddr_avg_bds,
                            gdma_l2_avg_bds, gdma_ddr_avg_bls, sdma_cycles, sdma_ddr_datasizes,
                            sdma_ddr_avg_bds, sdma_ddr_avg_bls, cdma_cycles,
                            cdma_ddr_datasizes, cdma_l2_datasizes, cdma_ddr_avg_bds,
                            cdma_l2_avg_bds, cdma_ddr_avg_bls]).tolist()

    def write(self):
        """
        Write summary information to Excel.
        :return: None
        """
        df = pd.DataFrame(self.data, columns=self.columns, index=None)
        df.to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=4, engine='xlsxwriter')
        para_desc = '(SimTiuCycle + SimGdmaCycle + SimSdmaCycle) / SimTotalCycle, presents the parallelism among all engines.'\
        if self.act_core_num > 1 else '(SimTiuCycle + SimGdmaCycle) / SimTotalCycle, presents the parallelism among all engines.'
        kpi_desc = pd.DataFrame(
                {
                    'Field': [
                        'Tiu Working Ratio',
                        'Parallelism(%)',
                        'Concurrency(%)',
                        'Total Alg Cycle',
                        'Total Alg Ops',
                        'Total uArch Ops',
                        'uArch Urate'],
                    'Description': [
                        'SimTiuCycle / SimTotalCycle, indicates the percentage of time tiu execution takes.',
                        para_desc,
                        '(SimTiuCycle + SimGdmaCycle - SimTotalCycle) / min(SimTiuCycle, SimGdmaCycle), indicates the concurrency between tiu and gdma.',
                        'The time required to execute tiu instructions theoretically.',
                        'The theoretical OPs required to execute tiu instructions.',
                        'The actual number of OPs accounting for microarchitecture.',
                        'totalAlgOps / totalUArchOps, since the shape of tensor needs to be aligned in '
                        'micro-architecture, the actual ops will be greater than the algorithm value. '
                        "It's better to closer to 100%."
                        ]
                },
                index=None
            )
        kpi_desc.to_excel(self.writer, index=False, sheet_name=self.sheet_name, startrow=9+self.act_core_num, startcol=0,
                                        engine='xlsxwriter')

    @classmethod
    def set_style(cls, file_path, chip_arch):
        """
        Set style for output Excel.
        :param file_path: the output Excel path.
        :return: None
        """
        wb = load_workbook(file_path)
        if cls.sheet_name not in wb.sheetnames:
            return
        df = pd.read_excel(file_path, cls.sheet_name)
        ws = wb[cls.sheet_name]
        ws.cell(1, 1).value = 'Platform: ' + chip_arch['Chip Arch']
        ws.cell(1, 1).fill = DetailsStyle.title_pattern
        ws.cell(1, 1).font = DetailsStyle.title_font
        ws.merge_cells('A1:B1')
        # set header style
        ws.cell(4, 8).value = 'TIU'
        ws.cell(4, 14).value = 'GDMA'
        ws.cell(4, 19).value = 'SDMA'
        ws.cell(4, 24).value = 'CDMA'
        for h, w in zip([4, 4, 4, 4], [8, 14, 19, 24]):
            ws.cell(h, w).fill = DetailsStyle.title_pattern
            ws.cell(h, w).font = DetailsStyle.title_font
        for w in range(6, 11):
            ws.cell(4, w).fill = DetailsStyle.tiu_pattern
        for w in range(11, 17):
            ws.cell(4, w).fill = DetailsStyle.gdma_pattern
        for w in range(17, 21):
            ws.cell(4, w).fill = DetailsStyle.sdma_pattern
        for w in range(21, 27):
            ws.cell(4, w).fill = DetailsStyle.cdma_pattern
        # set title style
        for cell in ws['5:5']:
            cell.fill = DetailsStyle.content_pattern
            cell.font = DetailsStyle.title_header_font
        # set content style
        content_start_rows = 6
        content_end_rows = len(df) + content_start_rows - 4
        for w in range(len(df.columns)):
            for h in range(content_start_rows, len(df) + content_start_rows):
                ws.cell(h, w + 1).font = DetailsStyle.title_font
        # set highlight style
        ratio_pos = 2
        parallelismPos = 3
        uRate_pos = 10
        total_uArch_ops = 9
        gdma_ddr_bd_pos = 14
        gdma_l2_bd_pos = 15
        gdma_ddr_avg_bl_pos = 16
        sdma_ddr_bd_pos = 19
        sdma_ddr_avg_bl_pos = 20
        cdma_ddr_bd_pos = 24
        cdma_l2_bd_pos = 25
        cdma_ddr_avg_bl_pos = 26
        chart_len = len(df) - 8
        for h in range(content_start_rows, chart_len):
            if int(ws.cell(h, total_uArch_ops).value) > 0:
                # ratio_pos
                if float(ws.cell(h, ratio_pos).value[:-1]) < 15:
                    ws.cell(h, ratio_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, ratio_pos).value[:-1]) < 30:
                    ws.cell(h, ratio_pos).fill = DetailsStyle.yellow_light
                # uRate_pos
                if float(ws.cell(h, uRate_pos).value[:-1]) < 50:
                    ws.cell(h, uRate_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, uRate_pos).value[:-1]) < 75:
                    ws.cell(h, uRate_pos).fill = DetailsStyle.yellow_light
            # parallelismPos
            if chip_arch['Chip Arch'].lower() in ["sg2260", 'a2']:
                if float(ws.cell(h, parallelismPos).value[:-1]) < 205:
                    ws.cell(h, parallelismPos).fill = DetailsStyle.yellow_light
            elif chip_arch['Chip Arch'].lower() in ['cv186x', 'bm1684x', 'mars3', 'bm1688']:
                if float(ws.cell(h, parallelismPos).value[:-1]) < 105:
                    ws.cell(h, parallelismPos).fill = DetailsStyle.red_light
            else:
                print('Not support chip arch')
                assert(0)
            # gdma_ddr_bd_pos
            ddr_max_bd, l2_max_bd, ddr_max_bl = float(chip_arch['DDR Max BW(GB/s)']), float(chip_arch['L2 Max BW(GB/s)']), float(chip_arch['Bus Max Burst'])
            if float(ws.cell(h, gdma_ddr_bd_pos).value) > 0:
                if float(ws.cell(h, gdma_ddr_bd_pos).value) < (ddr_max_bd * 0.5):
                    ws.cell(h, gdma_ddr_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, gdma_ddr_bd_pos).value) < (ddr_max_bd * 0.75):
                    ws.cell(h, gdma_ddr_bd_pos).fill = DetailsStyle.yellow_light
            # gdma_l2_bd_pos
            if float(ws.cell(h, gdma_l2_bd_pos).value) > 0:
                if float(ws.cell(h, gdma_l2_bd_pos).value) < (l2_max_bd * 0.5):
                    ws.cell(h, gdma_l2_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, gdma_l2_bd_pos).value) < (l2_max_bd * 0.75):
                    ws.cell(h, gdma_l2_bd_pos).fill = DetailsStyle.yellow_light
                # gdma_ddr_avg_bl_pos
                if float(ws.cell(h, gdma_ddr_avg_bl_pos).value) < ddr_max_bl:
                    ws.cell(h, gdma_ddr_avg_bl_pos).fill = DetailsStyle.red_light
            # sdma_ddr_bd_pos
            if float(ws.cell(h, sdma_ddr_bd_pos).value) > 0:
                if float(ws.cell(h, sdma_ddr_bd_pos).value) < (ddr_max_bd * 0.5):
                    ws.cell(h, sdma_ddr_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, sdma_ddr_bd_pos).value) < (ddr_max_bd * 0.75):
                    ws.cell(h, sdma_ddr_bd_pos).fill = DetailsStyle.yellow_light
                # sdma_ddr_avg_bl_pos
                if float(ws.cell(h, sdma_ddr_avg_bl_pos).value) < ddr_max_bl:
                    ws.cell(h, sdma_ddr_avg_bl_pos).fill = DetailsStyle.red_light
            # cdma_l2_bd_pos
            if ws.cell(h, cdma_l2_bd_pos).value and float(ws.cell(h, cdma_l2_bd_pos).value) > 0:
                if float(ws.cell(h, cdma_l2_bd_pos).value) < (l2_max_bd * 0.5):
                    ws.cell(h, cdma_l2_bd_pos).fill = DetailsStyle.red_light
                elif float(ws.cell(h, cdma_l2_bd_pos).value) < (l2_max_bd * 0.75):
                    ws.cell(h, cdma_l2_bd_pos).fill = DetailsStyle.yellow_light
                # cdma_ddr_avg_bl_pos
                if float(ws.cell(h, cdma_ddr_avg_bl_pos).value) < ddr_max_bl:
                    ws.cell(h, cdma_ddr_avg_bl_pos).fill = DetailsStyle.red_light
        content_end_rows = chart_len
        if chip_arch['Chip Arch'].lower() in ["sg2260", 'a2']:
            ws.cell(content_end_rows, parallelismPos).value = 'Parallelism<205%'
        elif chip_arch['Chip Arch'].lower() in ['cv186x', 'bm1684x', 'mars3', 'bm1688']:
            ws.cell(content_end_rows, parallelismPos).value = 'Parallelism<105%'
        else:
            print('Not support chip arch')
            assert(0)
        ws.cell(content_end_rows, ratio_pos).value = 'Ratio<30%'
        ws.cell(content_end_rows + 1, ratio_pos).value = 'Ratio<15%'
        ws.cell(content_end_rows, uRate_pos).value = 'Urate<75%'
        ws.cell(content_end_rows + 1, uRate_pos).value = 'Urate<50%'
        ws.cell(content_end_rows, gdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, gdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*50%'
        ws.cell(content_end_rows, gdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, gdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*50%'
        ws.cell(content_end_rows, gdma_ddr_avg_bl_pos).value = 'Length<' + str(ddr_max_bl)
        ws.cell(content_end_rows, sdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, sdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*50%'
        ws.cell(content_end_rows, sdma_ddr_avg_bl_pos).value = 'Length<' + str(ddr_max_bl)
        ws.cell(content_end_rows, cdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, cdma_ddr_bd_pos).value = 'BW<' + str(ddr_max_bd) + '*50%'
        ws.cell(content_end_rows, cdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*75%'
        ws.cell(content_end_rows + 1, cdma_l2_bd_pos).value = 'BW<' + str(l2_max_bd) + '*50%'
        ws.cell(content_end_rows, cdma_ddr_avg_bl_pos).value = 'Length<' + str(ddr_max_bl)
        for h, w in zip([content_end_rows, content_end_rows, content_end_rows, content_end_rows,
                         content_end_rows, content_end_rows, content_end_rows, content_end_rows],
                        [ratio_pos, parallelismPos, uRate_pos, gdma_ddr_bd_pos, gdma_l2_bd_pos, sdma_ddr_bd_pos, cdma_ddr_bd_pos, cdma_l2_bd_pos]):
            ws.cell(h, w).fill = DetailsStyle.yellow_light
            ws.cell(h, w).font = DetailsStyle.title_font
            ws.cell(h, w).alignment = DetailsStyle.center_align
        for h, w in zip([content_end_rows + 1, content_end_rows + 1, content_end_rows + 1, content_end_rows + 1, content_end_rows,
                        content_end_rows + 1, content_end_rows, content_end_rows + 1, content_end_rows + 1, content_end_rows],
                        [ratio_pos, uRate_pos, gdma_ddr_bd_pos, gdma_l2_bd_pos, gdma_ddr_avg_bl_pos, sdma_ddr_bd_pos, sdma_ddr_avg_bl_pos,
                         cdma_ddr_bd_pos, cdma_l2_bd_pos, cdma_ddr_avg_bl_pos]):
            ws.cell(h, w).fill = DetailsStyle.red_light
            ws.cell(h, w).font = DetailsStyle.title_font
            ws.cell(h, w).alignment = DetailsStyle.center_align
        # set border style
        header_start_row = 5
        for h in range(header_start_row, chart_len + header_start_row - 5):
            for w in range(1, len(df.columns) + 1):
                ws.cell(h, w).border = DetailsStyle.border
                ws.cell(h, w).alignment = DetailsStyle.center_align
        # set auto columns width
        for col in df.columns:
            index = list(df.columns).index(col)
            letter = get_column_letter(index + 1)
            collen = len(col)
            for h in range(1, chart_len + 4):
                if ws.cell(h, index + 1).value != None:
                    collen = max(collen, len(str(ws.cell(h, index + 1).value)))
            ws.column_dimensions[letter].width = collen * 1.05
        ws.sheet_properties.tabColor = 'FFC7CE'
        wb.save(file_path)
