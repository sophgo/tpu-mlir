#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
# @Time    : 2023/8/7 11:23
# @Author  : chongqing.zeng@sophgo.com
# @Project: PerfAI

"""
Note: The functionality to generate summary and layer information
    is temporarily not maintained due to the incomplete state of the
    global.profile file. Once the global.profile file is updated or
    completed, this feature can be revisited and implemented accordingly.
    Please refer to the global.profile file for more details on its
    current status and any updates that may be needed for the
    generation of summary and layer information.
"""

from enum import Enum

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from definition.bm1684x_defs import Arch, DataType
from definition.style import SummaryStyle
from utils.utils import load_arch_lib, get_ratio_str_3f, cycle_to_us, ops_to_gops, cycle_to_fps


class GlobalInfo:
    def __init__(self):
        self.subnet_list = []
        self.arch = Arch.UNKNOWN
        self.mem_info = []
        self.archlib = None
        self.freq = None
        self.net_name = None
        self.flops = 0
        self.no_perf_data = False
        self.quant_type = None

    def set_arch(self, arch):
        self.arch = arch
        if self.archlib is None:
            self.archlib = load_arch_lib(arch)
        assert self.archlib is not None


class SubnetInfo:
    def __init__(self):
        self.subnet_id = -1
        self.layer_list = []
        self.command_info = None
        self.gdma_nodes = []
        self.bd_nodes = []
        self.sim_info = None


class TensorInfo:
    def __init__(self):
        self.tensor_id = -1
        self.name = None
        self.shape = None
        self.dtype = DataType.UNKNOWN
        self.is_const = False
        self.address = -1
        self.gsize = 0
        self.loffset = -1
        self.nslice = 0
        self.hslice = 0
        self.in_layer = None
        self.out_layers = []


class StaticRunNode:
    __run_id = -1

    def __init__(self):
        self.__class__.__run_id += 1
        self.run_id = self.__class__.__run_id
        self.type = None
        self.core_id = -1
        self.bd_id = -1
        self.gdma_id = -1
        self.gdma_dir = None
        self.gdma_func = None
        self.bd_func = None
        self.layer = None
        self.command = None
        self.sim_info = None
        self.pmu_info = None


class SubnetType(Enum):
    UNKNOWN = -1
    TPU = 0
    CPU = 1
    MERGE = 2
    SWITCH = 3


class SummaryInfo:
    def __init__(self):
        self.layer_id = -1
        self.layer_type = None
        self.layer_name = ""
        self.is_local = False
        self.in_tensors = []
        self.out_tensors = []
        self.group_id = -1
        self.total_size = 0
        self.feature_size = 0
        self.weight_size = 0
        self.gdma_op = None
        self.gdma_tensor = None
        self.begin_usec = None
        self.end_usec = None
        self.gdma_nodes = []
        self.bd_nodes = []


class jsonObj:
    def __init__(self):
        self.flie_line = -1
        self.subnet_id = 0
        self.core_id = 0
        self.opcode = None
        self.bd_ids = None # (start_bd_id, end_bd_id]
        self.dma_ids = None # (start_gdma_id, end_gdma_id]
        self.operands = []
        self.results = []


class Summary:
    def __init__(self, writer):
        self.layer_summary_map = dict()
        self.layer_summary_header = ["Function", "Algorithm OPs", "Algorithm OPs Ratio", "Weight Size(B)(不准确)",
                                     "uArch OPs",
                                     "uArch URate", "uArch GOPs Ratio", "ASIC Cycles", "ASIC Time(us)",
                                     "Time Ratio", "ASIC FPS", "Operations"]
        self.mac_util_profile_header = ['Case', 'ReducedTime(us)', 'CurrentTotalTime(us)', 'Concurrency(%)',
                                        'macUtil(%)', 'Remark']
        self.layer_summary_row_map = {
            # Active
            "active": "Active",
            "leakyrelu": "Active",
            "prelu": "Active",
            # Attention
            "attention": "Attention",
            "fattention": "Attention",
            # BatchNorm
            # "batchnorm": "BatchNorm",
            # "batch_norm": "BatchNorm",
            # Conv
            "conv": "Conv",
            "conv2d": "Conv",
            "conv3d": "Conv",
            "deconv": "Conv",
            "deconv3d": "Conv",
            # Cast
            "cast": "Cast",
            # Eltwise
            "addconst": "Eltwise",
            "mulconst": "Eltwise",
            "subconst": "Eltwise",
            "maxconst": "Eltwise",
            "minconst": "Eltwise",
            "add": "Eltwise",
            "mul": "Eltwise",
            "sub": "Eltwise",
            "div": "Eltwise",
            "max": "Eltwise",
            "min": "Eltwise",
            "broadcast_binary": "Eltwise",
            "eltwise": "Eltwise",
            "eltwise_binary": "Eltwise",
            "binaryshift": "Eltwise",
            "binaryconstshift": "Eltwise",
            "clip": "Eltwise",
            "compare": "Eltwise",
            "compareconst": "Eltwise",
            # FC
            "fc": "FC",
            # MulShift
            "mulshift": "MulShift",
            # LayerNorm
            "groupnorm": "LayerNorm",
            "layernorm": "LayerNorm",
            "group_norm": "LayerNorm",
            "layer_norm": "LayerNorm",
            "instancenorm": "LayerNorm",
            "pixelnorm": "LayerNorm",
            # Lut
            "lut": "Lut",
            # MatMul
            "a16matmul": "MatMul",
            "batch_matmul": "MatMul",
            "matmul": "MatMul",
            # Pool
            "pool": "Pool",
            "pool2d": "Pool",
            # Reduce
            "reduce": "Reduce",
            # Requant
            # "requantfp": "Requant",
            # "requantint": "Requant",
            # "requantfpaxis": "Requant",
            # "requantintaxis": "Requant",
            # Slice
            # "slice": "StrideSlice",
            # "strideslice": "StrideSlice",
            # Softmax
            "softmax": "Softmax",
        }
        self.layer_summary_rows = [
            "Conv",
            "Pool",
            "Matmul",
            "Softmax",
            "BatchNorm",
            "LayerNorm",
            "Eltwise",
            "Others",
        ]
        self.sheet_name = 'Summary'
        self.data_rows = []
        self.mac_util_rows = []
        self.writer = writer

    def _get_layer_peak_tops(self, layer_name, dtype, platform, tpu_freq):
        platform = platform.lower()
        if layer_name in ['Conv', 'MatMul', 'Attention']:
            if platform == 'bm1684x':
                cu_num = 64 * 64 * 4 # int8 compute units number of BM1684X
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 64 * 16
            elif platform == 'a2':
                cu_num = 32 * 32 * 4
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 4
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 32 * 4
            elif platform == 'sg2260':
                cu_num = 64 * 64 * 4 * 8
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 64 * 16 * 8
            return cu_num * 2 * tpu_freq / 1e6
        elif layer_name in ['LayerNorm', 'Pool', 'Reduce', 'FC']:
            if platform == 'bm1684x':
                cu_num = 64 * 64 # int8 compute units number of BM1684X
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 64 * 16
            elif platform == 'a2':
                cu_num = 32 * 16
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 32 * 4
            elif platform == 'sg2260':
                cu_num = 64 * 64 * 8
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 64 * 16 * 8
            return cu_num * 2 * tpu_freq / 1e6
        elif layer_name in ['Lut']:
            if platform == 'bm1684x':
                cu_num = 64
            elif platform == 'a2':
                cu_num = 32
            elif platform == 'sg2260':
                cu_num = 64 * 8
            return cu_num * tpu_freq / 1e6
        else:
            if platform == 'bm1684x':
                cu_num = 64 * 64 # int8 compute units number of BM1684X
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 64 * 16
            elif platform == 'a2':
                cu_num = 32 * 16
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 32 * 4
            elif platform == 'sg2260':
                cu_num = 64 * 64 * 8
                if dtype in [DataType.FP16, DataType.BF16, 'f16', 'bf16']:
                    cu_num = cu_num / 2
                elif dtype in [DataType.FP32, 'f32']:
                    cu_num = 64 * 16 * 8
            return cu_num * tpu_freq / 1e6

    def _get_peak_tops(self, dtype, platform, tpu_freq):
        return self._get_layer_peak_tops('Conv', dtype, platform, tpu_freq)

    def _get_ddr_bw(self, tpu_freq):
        pass

    def _profile_with_peak_tops(self, layer_name, layer_tiu_us, layer_alg_ops, dtype, platform, tpu_freq, current_time, model_theo_time):
        peak_tops = self._get_peak_tops(dtype, platform, tpu_freq)
        tiu_theo_us = layer_alg_ops / peak_tops / 1e6
        reduced_us = layer_tiu_us - tiu_theo_us
        cur_time = current_time - reduced_us
        mac_util = round(model_theo_time/cur_time * 100, 3)
        return [layer_name + f' tiuTime: {layer_tiu_us:.2f} us -> {tiu_theo_us:.2f} us',
                reduced_us, cur_time, 100.0, mac_util,
                f'{layer_name}的耗时用ModelPeakTops得到的理论耗时替换']

    def _profile_with_layer_peak_tops(self, layer_name, layer_tiu_us, layer_alg_ops, dtype, platform, tpu_freq, current_time, model_theo_time):
        peak_tops = self._get_layer_peak_tops(layer_name, dtype, platform, tpu_freq)
        tiu_theo_us = layer_alg_ops / peak_tops / 1e6
        reduced_us = layer_tiu_us - tiu_theo_us
        cur_time = current_time - reduced_us
        mac_util = round(model_theo_time/cur_time * 100, 3)
        return [layer_name + f' tiuTime: {layer_tiu_us:.2f} us -> {tiu_theo_us:.2f} us',
                reduced_us, cur_time, 100.0, mac_util,
                f'{layer_name}的耗时用LayerPeakTops得到的理论耗时替换']
    def _profile_with_uarch_rate(self):
        pass

    def load(self, layer_infos, chip_arch):
        quant_type = chip_arch['quant_type']
        platform = chip_arch['Chip Arch']
        tpu_freq = float(chip_arch['TIU Frequency(MHz)'])
        s = dict()
        for i in layer_infos:
            func_name = self.layer_summary_row_map.get(i.layer_name.lower(), "Others")
            if func_name in s:
                s[func_name].add(i.layer_name)
            else:
                s[func_name] = {i.layer_name}
        self.layer_summary_rows = sorted(list(s.keys()))
        self.layer_summary_rows.remove("Others")
        self.layer_summary_rows.append("Others")
        total_weight_size = 0
        total_alg_ops = 0
        total_arch_ops = 0
        total_sim_cycles = 0
        for k in self.layer_summary_rows:
            self.layer_summary_map[k] = [k, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, s[k]]
        for layer in layer_infos:
            layer_name = layer.layer_name
            row_name = self.layer_summary_row_map.get(layer_name.lower(), "Others")
            layer_alg_ops = layer.alg_ops
            layer_arch_ops = layer.uarch_ops
            weight_size = layer.weight_size
            layer_cycles = layer.sim_cycle
            total_weight_size += weight_size
            total_alg_ops += layer_alg_ops
            total_arch_ops += layer_arch_ops
            total_sim_cycles += layer_cycles
            item = [row_name, layer_alg_ops, 0, weight_size, layer_arch_ops, 0, 0, layer_cycles,
                    0, 0, layer_cycles]
            for i in range(1, len(item) - 1):
                self.layer_summary_map[row_name][i] += item[i]
        for row in self.layer_summary_rows:
            self.data_rows.append(self.layer_summary_map.get(row))
        for summary in self.data_rows:
            summary[2] = get_ratio_str_3f(summary[1], total_alg_ops)
            summary[6] = get_ratio_str_3f(summary[4], total_arch_ops)
            summary[5] = get_ratio_str_3f(summary[1], summary[4])
            summary[8] = cycle_to_us(summary[7], tpu_freq)
            summary[9] = get_ratio_str_3f(summary[7], total_sim_cycles)
            # summary[1] = ops_to_gops(summary[1])
            # summary[4] = ops_to_gops(summary[4])
            summary[10] = '-'
        total_arch_urate = get_ratio_str_3f(total_alg_ops, total_arch_ops)
        others_row = self.data_rows[-1]
        self.data_rows = sorted(self.data_rows[:-1], key=lambda x: x[7], reverse=True)
        self.data_rows.append(others_row)
        self.data_rows.append(
            ["Overall", total_alg_ops, "100%", total_weight_size, total_arch_ops,
             total_arch_urate, "100%",
             total_sim_cycles, cycle_to_us(total_sim_cycles, tpu_freq), "100%", cycle_to_fps(total_sim_cycles), "-"])

        tiu_time = float(self.data_rows[-1][8])
        total_time = float(chip_arch['total_time(us)'])
        model_theo_time = chip_arch['flops'] / self._get_peak_tops(quant_type, platform, tpu_freq) / 1e6
        mac_util = round(model_theo_time / total_time * 100, 3)
        self.mac_util_rows.append(
            ['origin', 0, total_time, float(chip_arch['concurrency'][:-1]), mac_util, '排除CPU耗时及输入输出在runtime空间和用户空间的搬运耗时']
        )
        self.mac_util_rows.append(
            ['100% Concurrency', total_time - tiu_time, tiu_time, 100.0, round(model_theo_time / tiu_time * 100, 3), 'TIU和GDMA的并行度100%']
        )
        current_time = self.mac_util_rows[1][2]
        for row in self.data_rows[:-1]:
            self.mac_util_rows.append(self._profile_with_peak_tops(row[0], float(row[8]), int(row[1]), quant_type, platform, tpu_freq, current_time, model_theo_time))
            current_time = self.mac_util_rows[-1][2]
        current_time = self.mac_util_rows[1][2]
        for row in self.data_rows[:-1]:
            self.mac_util_rows.append(self._profile_with_layer_peak_tops(row[0], float(row[8]), int(row[1]), quant_type, platform, tpu_freq, current_time, model_theo_time))
            current_time = self.mac_util_rows[-1][2]


    def write(self, chip_arch):
        network = chip_arch['network']
        flops = chip_arch['flops']
        quant_type = chip_arch['quant_type']
        platform = chip_arch['Chip Arch']
        if platform.lower() == 'sg2260':
            int8_ops = 256
            fp32_ops = 16
        elif platform.lower() == 'bm1684x':
            int8_ops = 32
            fp32_ops = 2.2
        elif platform.lower() == 'a2':
            int8_ops = 14.4
            fp32_ops = 0.45
        ddr_bw = round(float(chip_arch['DDR Max BW(GB/s/Core)']) * int(chip_arch['Core Num']), 2)
        tpu_freq = float(chip_arch['TIU Frequency(MHz)'])
        dma_freq = float(chip_arch['DMA Frequency(MHz)'])
        condition_dict = {
            'Network': [network],
            'Model GFLOPs': [flops / 1e9],
            'Quant Type': [quant_type],
            'platform': [platform],
            'TPU INT8 TOPS': [self._get_peak_tops(DataType.INT8, platform, tpu_freq)],
            'TPU FP16 TOPS': [self._get_peak_tops(DataType.FP16, platform, tpu_freq)],
            'TPU FP32 TOPS': [self._get_peak_tops(DataType.FP32, platform, tpu_freq)],
            'DDR BW(GB/s)': [ddr_bw],
            'TPU Frequency(MHz)': [tpu_freq],
            'DMA Frequency(MHz)': [dma_freq]
        }
        npu_num = int(chip_arch['NPU Num'])
        Cube_IC_Align = int(chip_arch['Cube IC Align(8bits)'])
        Conv_OHOW_Align = int(chip_arch['Cube OHOW Align(8bits)'])
        Vector_OHOW_Align = int(chip_arch['Vector OHOW Align(8bits)'])
        Conv_ops = int(npu_num * Cube_IC_Align * Conv_OHOW_Align * tpu_freq * 2 / 1000)
        Vector_ops = round(npu_num * Vector_OHOW_Align * tpu_freq / 1000, 2)
        pooling_ops = Conv_ops / 4
        detail_spec = {
            'NPU NUM (lane)': [npu_num],
            'IC Alignment (8bits)': [Cube_IC_Align],
            'Conv OHOW Align': [Conv_OHOW_Align],
            'Vector OHOW Align(8bits)': [Vector_OHOW_Align],
            'Conv Ops/s': [Conv_ops],
            'Vect Ops/s': [Vector_ops],
            'Pool OPs/s': [pooling_ops]
        }

        pd.DataFrame(condition_dict).to_excel(self.writer, index=False, sheet_name=self.sheet_name, engine='xlsxwriter', startrow=0, startcol=1, float_format='%g')
        pd.DataFrame(detail_spec).to_excel(self.writer, index=False, sheet_name=self.sheet_name, engine='xlsxwriter', startrow=3, startcol=1, float_format='%g')
        if len(self.data_rows) > 0:
            df = pd.DataFrame(self.data_rows)
            df.columns = self.layer_summary_header
            df.to_excel(self.writer, index=False, sheet_name=self.sheet_name, engine='xlsxwriter', startrow=8, startcol=1, float_format='%g')
            next_start_row = 8 + len(df) + 3

        if len(self.mac_util_rows) > 0:
            df = pd.DataFrame(self.mac_util_rows)
            df.columns = self.mac_util_profile_header
            df.to_excel(self.writer, index=False, sheet_name=self.sheet_name, engine='xlsxwriter', startrow=next_start_row, startcol=1, float_format='%g')

    @classmethod
    def set_style(cls, file_path):
        """
        Set style for Excel, and highlight some fields we care about.
        :param frozen: freeze some statistics above the sheet
        :param file_path: Excel output path
        :param core_id: the id of current core
        :return: None
        """
        wb = load_workbook(file_path)
        summary_style = SummaryStyle()
        sheet_name = 'Summary'
        df = pd.read_excel(file_path, sheet_name)
        layer_num = (len(df) - 12) // 3
        ws = wb[sheet_name]
        ws.cell(1, 1).value = 'Condition'
        ws.cell(4, 1).value = 'Detail Spec'
        ws.cell(8, 1).value = 'Performance'
        ws.cell(9, 1).value = 'Summary'
        ws.cell(13+layer_num, 1).value = 'mac_util'
        ws.cell(14+layer_num, 1).value = 'analysis'
        for h, w in zip([1, 4, 8, 9, 13+layer_num, 14+layer_num], [1, 1, 1, 1, 1, 1]):
            ws.cell(h, w).fill = summary_style.title_pattern
            ws.cell(h, w).font = summary_style.title_font
        summary_start_cols = 2
        summary_end_cols = summary_start_cols + 10
        for w in range(summary_start_cols, summary_end_cols):
            for h in [1, 4]:
                if h == 1:
                    ws.cell(h, w).fill = summary_style.title_header_pattern
                    ws.cell(h, w).font = summary_style.title_header_font
                    ws.cell(h, w).border = SummaryStyle.border
                    ws.cell(h+1, w).fill = summary_style.title_content_pattern
                    ws.cell(h+1, w).font = summary_style.title_header_font
                    ws.cell(h+1, w).alignment = summary_style.center_align
                    ws.cell(h+1, w).border = SummaryStyle.border
                elif h == 4 and w < summary_end_cols - 3:
                    # summary title style
                    ws.cell(h, w).fill = summary_style.title_header_pattern
                    ws.cell(h, w).font = summary_style.title_header_font
                    ws.cell(h, w).border = SummaryStyle.border
                    # summary content style
                    ws.cell(h+1, w).fill = summary_style.title_content_pattern
                    ws.cell(h+1, w).font = summary_style.title_header_font
                    ws.cell(h+1, w).alignment = summary_style.center_align
                    ws.cell(h+1, w).border = SummaryStyle.border

        content_start_cols = 2
        content_end_cols = content_start_cols + 12
        content_header_row = 8
        for w in range(content_start_cols, content_end_cols):
            ws.cell(content_header_row+1, w).fill = summary_style.content_header2_pattern
            ws.cell(content_header_row+1, w).font = summary_style.content_header_font
            for h in range(content_header_row+2, content_header_row+2+layer_num+1):
                ws.cell(h, w).font = summary_style.title_header_font
                ws.cell(h, w).alignment = summary_style.center_align
                ws.cell(h, w).border = SummaryStyle.border
                ws.cell(h, w).fill = summary_style.content1_pattern
            ws.cell(content_header_row+2+layer_num, w).fill = summary_style.content2_pattern
        ws.cell(8, 3).value = 'Algorithm'
        ws.cell(8, 6).value = 'uArch'
        ws.cell(8, 9).value = 'Simulation'
        for w in range(3, 13):
            ws.cell(8, w).fill = summary_style.content_header1_pattern
        for h, w in zip([8, 8, 8], [3, 6, 9]):
            ws.cell(h, w).font = summary_style.title_font
            ws.cell(h, w).border = SummaryStyle.title_border
            ws.cell(h, w).alignment = summary_style.center_align
        ws.merge_cells('C8:E8')
        ws.merge_cells('F8:H8')
        ws.merge_cells('I8:L8')

        # mac_util analysis
        content_start_cols = 2
        content_end_cols = content_start_cols + 6
        content_header_row = 13+layer_num
        for w in range(content_start_cols, content_end_cols):
            ws.cell(content_header_row, w).fill = summary_style.content_header2_pattern
            ws.cell(content_header_row, w).font = summary_style.content_header_font
            # origin, 100% concurrency
            for h in range(content_header_row+1, content_header_row+3):
                ws.cell(h, w).font = summary_style.title_header_font
                ws.cell(h, w).border = SummaryStyle.border
                ws.cell(h, w).fill = summary_style.content1_pattern
            # Analysis with ModelPeakTops
            for h in range(content_header_row+3, content_header_row+3+layer_num):
                ws.cell(h, w).font = summary_style.title_header_font
                ws.cell(h, w).border = SummaryStyle.border
                ws.cell(h, w).fill = summary_style.content2_pattern
            # Analysis with LayerPeakTops
            for h in range(content_header_row+3+layer_num, content_header_row+3+layer_num*2):
                ws.cell(h, w).font = summary_style.title_header_font
                ws.cell(h, w).border = SummaryStyle.border
                ws.cell(h, w).fill = summary_style.content1_pattern

        for col in df.columns:
            # auto set columns width
            index = list(df.columns).index(col)
            collen = 0
            letter = get_column_letter(index + 1)
            for h in range(1, len(df) + 8):
                if h > 50:
                    break
                if ws.cell(h, index + 1).value is not None:
                    if len(str(ws.cell(h, index + 1).value)) > 35:
                        continue
                    collen = max(collen, len(str(ws.cell(h, index + 1).value)))
            ws.column_dimensions[letter].width = collen * 1.05
        ws.sheet_properties.tabColor = '31869B'
        wb.save(file_path)
