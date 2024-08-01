#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/8/7 11:49
# @Author  : chongqing.zeng@sophgo.com
# @Project: PerfAI
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


class DetailsStyle:
    title_pattern = PatternFill(fgColor='ffffff', fill_type="solid")
    title_header_pattern = PatternFill(fgColor='548235', fill_type="solid")
    title_content_pattern = PatternFill(fgColor='F2F2F2', fill_type="solid")
    title_font = Font(u'等线', size=10, bold=True, italic=False, strike=False, color='000000')
    title_header_font = Font(u'Calibri', size=10, bold=True, italic=False, strike=False, color='ffffff')

    content_pattern = PatternFill(fgColor='305496', fill_type="solid")
    key_content_pattern = PatternFill(fgColor='FCE4D6', fill_type="solid")
    red_light = PatternFill(fgColor='FF0000', fill_type="solid")
    yellow_light = PatternFill(fgColor='FFFF00', fill_type="solid")

    center_align = Alignment(horizontal='center', vertical='top')
    right_align = Alignment(horizontal='right', vertical='top')
    left_align = Alignment(horizontal='left', vertical='top')
    side = Side(border_style='thin', color="000000")
    border = Border(bottom=side, left=side, right=side, top=side)

    tiu_pattern = PatternFill(fgColor='008000', fill_type='solid')
    gdma_pattern = PatternFill(fgColor='FFA500', fill_type='solid')
    sdma_pattern = PatternFill(fgColor='D0CECE', fill_type='solid')
    cdma_pattern = PatternFill(fgColor='C0504D', fill_type='solid')

class LayerStyle:
    title_pattern = PatternFill(fgColor='ffffff', fill_type="solid")
    title_header_pattern = PatternFill(fgColor='B8CCE4', fill_type="solid")
    title_content_pattern = PatternFill(fgColor='F2F2F2', fill_type="solid")
    title_font = Font(u'等线', size=10, bold=True, italic=False, strike=False, color='000000')
    title_header_font = Font(u'Calibri', size=10, bold=True, italic=False, strike=False, color='000000')

    content_header1_pattern = PatternFill(fgColor='95B3D7', fill_type="solid")
    content_header2_pattern = PatternFill(fgColor='31869B', fill_type="solid")
    uarch_pattern = PatternFill(fgColor='538FD5', fill_type="solid")
    alg_pattern = PatternFill(fgColor='B8CCE4', fill_type="solid")
    sim_pattern = PatternFill(fgColor='95B3D7', fill_type="solid")
    content_header_font = Font(u'Calibri', size=10, bold=True, italic=False, strike=False, color='ffffff')

    red_light = PatternFill(fgColor='FF0000', fill_type="solid")
    yellow_light = PatternFill(fgColor='FFFF00', fill_type="solid")

    center_align = Alignment(horizontal='center', vertical='top')
    right_align = Alignment(horizontal='right', vertical='top')
    left_align = Alignment(horizontal='left', vertical='top')
    side = Side(border_style='thin', color="000000")
    border = Border(bottom=side, left=side, right=side, top=side)
    side2 = Side(border_style='medium', color="000000")
    title_border = Border(bottom=side2, left=side2, right=side2, top=side2)
    tab_pattern = PatternFill(fgColor='0070C0', fill_type='solid')


class SummaryStyle:
    title_pattern = PatternFill(fgColor='ffffff', fill_type="solid")
    title_header_pattern = PatternFill(fgColor='B8CCE4', fill_type="solid")
    title_content_pattern = PatternFill(fgColor='F4DCDB', fill_type="solid")
    title_font = Font(u'等线', size=10, bold=True, italic=False, strike=False, color='000000')
    title_header_font = Font(u'Calibri', size=10, bold=True, italic=False, strike=False, color='000000')

    # content0_pattern = PatternFill(fgColor='F2F2F2', fill_type="solid")
    content1_pattern = PatternFill(fgColor='FDF5E6', fill_type="solid")
    content2_pattern = PatternFill(fgColor='F5DEB3', fill_type="solid")
    content_header1_pattern = PatternFill(fgColor='C6E0B4', fill_type="solid")
    content_header2_pattern = PatternFill(fgColor='548235', fill_type="solid")
    content_header_font = Font(u'Calibri', size=10, bold=True, italic=False, strike=False, color='ffffff')

    red_light = PatternFill(fgColor='FF0000', fill_type="solid")
    yellow_light = PatternFill(fgColor='FFFF00', fill_type="solid")

    center_align = Alignment(horizontal='center', vertical='top')
    right_align = Alignment(horizontal='right', vertical='top')
    left_align = Alignment(horizontal='left', vertical='top')
    side = Side(border_style='thin', color="000000")
    border = Border(bottom=side, left=side, right=side, top=side)
    side2 = Side(border_style='medium', color="000000")
    title_border = Border(bottom=side2, left=side2, right=side2, top=side2)
    tab_pattern = PatternFill(fgColor='31869B', fill_type='solid')
