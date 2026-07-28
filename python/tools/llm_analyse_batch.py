#!/usr/bin/env python3
# ==============================================================================
#
# Copyright (C) 2025 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
"""Batch LLM analysis driver.

Loops :mod:`llm_analyse` over a grid of ``seq_length`` and
``max_input_length`` values (other parameters -- compute power, bandwidth,
utilizations -- are fixed for the whole batch), collects every generated
workbook into one folder, and optionally prints / writes a summary table of
TTFT and Tokens/s.

Four grid modes are selected with ``--mode``:

``auto``    each seq auto-sweeps max_input_length over powers of two from 256 to
           next_pow2(seq/2) (i.e. ~half the sequence, snapped up to a power of two).
``cross``   full cross product of seq x max_input_length (skips mil > seq).
``paired``  zip seq_length and max_input_length (equal-length lists).
``dedup``   same report grid as auto/cross, but only runs the minimal set of
            configurations needed to measure every distinct prefill
            (max_input_length) and every distinct decode (seq_length) once --
            prefill_tps depends only on mil and decode_tps only on seq (for the
            standard no-history-KV prefill), so the remaining grid points are
            reconstructed in the summary from those measurements.

Per-run artefacts land in a private sub-directory (``tmp_s<seq>_mil<mil>``)
under the output folder so the lowered MLIR of different shapes never collide.
With ``--only_result`` those sub-directories (including ``tmp_mlir_analyse``)
are removed once the workbook has been collected, leaving only the renamed
``.xlsx`` files behind.

Example::

    llm_analyse_batch.py -m /workspace/Qwen3.5-4B -s 2048 4096 8192 \\
        --max_input_length 512 1024 2048 -t 20 -b 128 -c bm1684x \\
        -o qwen35_batch --mode dedup --only_result
"""
import os
import sys
import csv
import argparse
import shutil
import subprocess
from typing import Optional

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

LLM_ANALYSE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "llm_analyse.py")

# ---------------------------------------------------------------------------
# Roofline recompute (workbooks store formulas with no cached values)
# ---------------------------------------------------------------------------
KEY_OPS = {
    "MatMul", "Conv", "Conv2D", "FAttention", "ChunkGatedDeltaRule", "A16MatMul", "Fp8MatMul", "Mlp"
}
PREFILL = {"vit", "embedding", "block", "block(LinearAttention)", "block(FullAttention)", "lm_head"}
DECODE = {
    "embedding_cache", "block_cache", "block_cache(LinearAttention)", "block_cache(FullAttention)",
    "lm_head"
}
RATIO_LABELS = {
    "FAttention (Prefill)": "fa_prefill",
    "FAttention (Decode)": "fa_decode",
    "Gather": "gather",
    "Permute/Concat": "permute_concat",
    "Mlp": "mlp",
    "ChunkGatedDeltaRule": "chunk",
    "RecurrentGatedDeltaRule": "recurrent",
}


def _ctype(opn: str) -> str:
    if opn == "Fp8MatMul":
        return "fp8"
    if opn in KEY_OPS:
        return "fp16"
    return "vector"


def recompute_metrics(path: str) -> Optional[dict]:
    """Recompute per-phase times from a workbook by replaying the roofline.

    Returns dict with: ``prefill_time`` (s, == TTFT, includes preproc if vit),
    ``decode_time`` (s, per-token), ``vit_time`` (s, vit module TPU time or
    None), ``preproc`` (s, 0 if no vit).
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(path, data_only=False)
    except Exception:
        return None
    if "Overview" not in wb.sheetnames:
        return None
    ws0 = wb["Overview"]
    params, ratios, mods = {}, {}, []
    for r in range(1, ws0.max_row + 1):
        label = ws0.cell(row=r, column=1).value
        val = ws0.cell(row=r, column=2).value
        if label is None:
            continue
        if label in RATIO_LABELS and isinstance(val, (int, float)):
            ratios[RATIO_LABELS[label]] = val
        elif isinstance(val, (int, float)):
            params[label] = val
        phase = ws0.cell(row=r, column=7).value
        cnt = ws0.cell(row=r, column=2).value
        if phase in ("Prefill", "Decode", "Both") and label and isinstance(cnt, (int, float)):
            mods.append((label, int(cnt), phase))
    try:
        uarch = params["uArch Rate"]
        bwutil = params["Bandwidth Utilization"]
        serial = 1.0 - params["Parallelism"]
        cpu_call = params["CPU Call"]
        preproc = params.get("Preprocess Time", 0.0)
        bw = params["Chip Bandwidth"]
    except KeyError:
        return None

    def tops(ctype):
        if ctype == "fp8":
            return params["FP8 Compute Power"]
        if ctype == "fp16":
            return params["FP16 Compute Power"]
        if ctype == "int8":
            return params["INT8 Compute Power"]
        return params["Vector Compute Power"]

    prefill_time = decode_time = vit_time = 0.0
    has_vit = False
    for name, count, phase in mods:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        is_decode = name in DECODE and name not in PREFILL
        mod_us = 0.0
        for r in range(2, ws.max_row + 1):
            opn = ws.cell(row=r, column=1).value
            gops = ws.cell(row=r, column=4).value
            io_mb = ws.cell(row=r, column=7).value
            if opn is None or not isinstance(gops, (int, float)):
                continue
            io_mb = io_mb if isinstance(io_mb, (int, float)) else 0.0
            t = tops(_ctype(opn))
            c_us = gops / (t * uarch) * 1000 if t else 0.0
            m_us = io_mb / (bw * bwutil) * 1000 if bw else 0.0
            base = max(c_us, m_us) + serial * min(c_us, m_us)
            ratio = 1.0
            if opn == "FAttention":
                ratio = ratios.get("fa_decode" if is_decode else "fa_prefill", 1.0)
            elif opn == "Gather":
                ratio = ratios.get("gather", 1.0)
            elif opn in ("Permute", "Concat"):
                ratio = ratios.get("permute_concat", 1.0)
            elif opn == "Mlp":
                ratio = ratios.get("mlp", 1.0)
            elif opn == "ChunkGatedDeltaRule":
                ratio = ratios.get("chunk", 1.0)
            elif opn == "RecurrentGatedDeltaRule":
                ratio = ratios.get("recurrent", 1.0)
            mod_us += base * ratio
        total_s = mod_us / 1e6 * count + count * cpu_call / 1e6
        if name == "vit":
            has_vit = True
            vit_time = total_s
        if "Prefill" in phase or phase == "Both":
            prefill_time += total_s
        if "Decode" in phase or phase == "Both":
            decode_time += total_s
    if has_vit:
        prefill_time += preproc
    return dict(
        prefill_time=prefill_time,
        decode_time=decode_time,
        vit_time=vit_time if has_vit else None,
        preproc=preproc if has_vit else 0.0,
    )


# ---------------------------------------------------------------------------
# Batch driver
# ---------------------------------------------------------------------------
def build_analyse_cmd(args, seq: int, mil: int, run_out_dir: str) -> str:
    parts = [
        sys.executable, LLM_ANALYSE, "-m", args.model_path, "-s",
        str(seq), "-t",
        str(args.tops), "-b",
        str(args.bandwidth), "-q", args.quantize, "-c", args.chip, "-o", run_out_dir
    ]
    if args.int8_tops is not None:
        parts += ["--int8_tops", str(args.int8_tops)]
    if args.fp8_tops is not None:
        parts += ["--fp8_tops", str(args.fp8_tops)]
    if args.vector_tops is not None:
        parts += ["--vector_tops", str(args.vector_tops)]
    if args.uarch_rate is not None:
        parts += ["--uarch_rate", str(args.uarch_rate)]
    if args.bw_util is not None:
        parts += ["--bw_util", str(args.bw_util)]
    if args.parallelism is not None:
        parts += ["--parallelism", str(args.parallelism)]
    if args.max_pixels:
        parts += ["--max_pixels", args.max_pixels]
    if mil > 0:
        parts += ["--max_input_length", str(mil)]
    return parts


def cleanup_run_dir(run_out_dir: str) -> None:
    """Remove per-run intermediates under ``run_out_dir`` and drop the
    now-empty directory.

    ``tmp_mlir_analyse/`` holds the lowered MLIR and ``config/`` holds the
    copied model config (tokenizer, config.json, embedding.bin, ...); both
    are test-time intermediates, not results. Anything else left in the run
    dir is kept in place so the user can inspect it rather than being
    silently deleted.
    """
    if not os.path.isdir(run_out_dir):
        return
    for sub in ("tmp_mlir_analyse", "config"):
        sub_path = os.path.join(run_out_dir, sub)
        if os.path.isdir(sub_path):
            shutil.rmtree(sub_path)
    if os.path.isdir(run_out_dir) and not os.listdir(run_out_dir):
        os.rmdir(run_out_dir)


# Default auto-sweep: powers of two from 256 up to next_pow2(seq/2). Capping
# at ~half the sequence leaves the other half for generation (a prompt that
# fills the entire context cannot produce output); snapping the cap up to a
# power of two keeps every mil value shared across seqs so dedup reuses
# measurements. next_pow2(seq/2) is always < seq, so the degenerate mil=seq
# is never produced.
AUTO_MIL_FLOOR = 256


def _next_pow2(x: int) -> int:
    """Smallest power of two >= x (returns 1 for x <= 0)."""
    if x <= 0:
        return 1
    p = 1
    while p < x:
        p <<= 1
    return p


def auto_max_input_lengths(seq: int) -> list:
    """Default max_input_length sweep for one seq_length.

    Powers of two from 256 up to ``next_pow2(seq // 2)``. The ~half-seq cap
    leaves room for generation, and snapping it up to a power of two keeps
    every value shared across seqs so dedup reuses measurements instead of
    re-running (e.g. seq=3000 -> [256,512,1024,2048]). For a seq too small
    to fit even one power-of-two prompt at the floor, falls back to testing
    its own cap.
    """
    cap = _next_pow2(seq // 2)
    vals = []
    v = AUTO_MIL_FLOOR
    while v <= cap:
        vals.append(v)
        v *= 2
    if not vals and cap > 0:
        vals.append(cap)
    return vals


def build_report_grid(args) -> list:
    """The full grid of (seq, mil) rows to report in the summary.

    dedup reuses the auto/cross grid construction -- it only changes which of
    these rows are actually run vs reconstructed.
    """
    mode = args.mode
    if mode in ("auto", "dedup") and args.max_input_length is None:
        per_seq = {s: auto_max_input_lengths(s) for s in args.seq_length}
        return [(s, mil) for s in args.seq_length for mil in per_seq[s]]
    if mode == "paired":
        if len(args.seq_length) != len(args.max_input_length):
            print(
                "Error: --mode paired requires seq_length and max_input_length of "
                "equal length.",
                file=sys.stderr)
            sys.exit(1)
        return list(zip(args.seq_length, args.max_input_length))
    # cross, or dedup with explicit max_input_length
    full = len(args.seq_length) * len(args.max_input_length)
    grid = [(s, mil) for s in args.seq_length for mil in args.max_input_length
            if mil == 0 or mil <= s]
    skipped = full - len(grid)
    if skipped > 0:
        print(f"Skipping {skipped} configuration(s) where max_input_length > "
              f"seq_length (invalid prefill).")
    return grid


def build_dedup_runs(grid: list) -> list:
    """Minimal run set covering every distinct mil>0 (prefill) and every
    distinct seq (decode).

    For the standard no-history-KV prefill, prefill_time depends only on mil
    and decode_time only on seq, so one measurement per distinct mil and per
    distinct seq suffices. mil==0 (sentinel: prefill the whole seq) is NOT a
    shareable mil -- its prefill length is the seq -- so those rows are run
    individually (their decode_time(seq) is still shareable).

    Returns at most ~max(|seqs|, |mils|) runs (plus the mil==0 sentinels).
    """
    sentinel = [(s, m) for (s, m) in grid if m == 0]
    mils = sorted({m for _, m in grid if m > 0})
    seqs = sorted({s for s, _ in grid})
    runs = list(sentinel)
    used_mils, used_seqs = set(), set()

    def fresh_mil(s):
        return next((m for m in reversed(mils) if m not in used_mils and m <= s), None)

    def fresh_seq(m):
        return next((s for s in reversed(seqs) if s not in used_seqs and s >= m), None)

    def any_mil(s):
        return next((m for m in mils if m <= s), None)

    def any_seq(m):
        return next((s for s in seqs if s >= m), None)

    if len(seqs) >= len(mils):
        # seqs drive: one run per seq, pair a fresh mil where possible.
        for s in reversed(seqs):
            m = fresh_mil(s)
            if m is None:
                m = any_mil(s)  # reuse an already-measured mil
            if m is None:
                continue  # seq only appears via mil==0 sentinel -> decode covered there
            used_mils.add(m)
            used_seqs.add(s)
            runs.append((s, m))
        for m in mils:
            if m not in used_mils:  # safety net
                s = any_seq(m)
                if s is not None:
                    runs.append((s, m))
                    used_mils.add(m)
    else:
        # mils drive: one run per mil, pair a fresh seq where possible.
        for m in reversed(mils):
            s = fresh_seq(m)
            if s is None:
                s = any_seq(m)
            if s is None:
                continue
            used_seqs.add(s)
            used_mils.add(m)
            runs.append((s, m))
        for s in seqs:
            if s not in used_seqs:  # safety net
                m = any_mil(s)
                if m is not None:
                    runs.append((s, m))
                    used_seqs.add(s)

    # dedup while preserving order
    seen, out = set(), []
    for r in runs:
        if r not in seen:
            seen.add(r)
            out.append(r)
    return out


def reconstruct_metrics(seq: int, mil: int, lookups: dict) -> dict:
    """Reconstruct a row's metrics from the dedup measurement lookups."""
    tp = lookups["text_prefill_by_mil"].get(mil) if mil > 0 else None
    dt = lookups["decode_by_seq"].get(seq)
    vit = lookups["vit_time"]
    preproc = lookups["preproc"]
    prefill_tps = (mil / tp) if (mil > 0 and tp is not None and tp > 0) else None
    if tp is not None:
        ttft = tp + (vit or 0.0) + (preproc if vit else 0.0)
    else:
        ttft = None
    decode_tps = (1.0 / dt) if (dt is not None and dt > 0) else None
    vit_fps = (1.0 / vit) if vit else None
    return dict(
        ttft=ttft,
        decode_time=dt,
        vit_time=vit,
        prefill_tps=prefill_tps,
        decode_tps=decode_tps,
        vit_fps=vit_fps,
        reconstructed=True,
    )


def main():
    parser = argparse.ArgumentParser(
        description="Batch-run llm_analyse over a grid of seq_length / "
        "max_input_length and collect the workbooks into one folder.")
    parser.add_argument('-m',
                        '--model_path',
                        type=str,
                        required=True,
                        help='original weight, like ./Qwen2-7B-Instruct')
    parser.add_argument('-s',
                        '--seq_length',
                        type=int,
                        nargs="+",
                        required=True,
                        help='seq_length value(s) to sweep')
    parser.add_argument('--max_input_length',
                        type=int,
                        nargs="+",
                        default=None,
                        help='max_input_length value(s); required for cross/paired, '
                        'optional for dedup (auto grid when omitted), ignored for auto')
    parser.add_argument('-t',
                        '--tops',
                        type=float,
                        required=True,
                        help='FP16 compute power in TOPS (fixed for the batch)')
    parser.add_argument('-b',
                        '--bandwidth',
                        type=float,
                        required=True,
                        help='Chip memory bandwidth in GB/s (fixed for the batch)')
    parser.add_argument('-q',
                        '--quantize',
                        default="auto",
                        choices=["auto", "f16", "w8f16", "w4f16", "bf16", "w8bf16", "w4bf16"],
                        help='Quantization mode (default: auto)')
    parser.add_argument('-c',
                        '--chip',
                        default="bm1684x",
                        choices=["bm1684x", "bm1688", "cv186x", "bm1690", "bm1684x2"],
                        help='Chip type (default: bm1684x)')
    parser.add_argument('--int8_tops',
                        type=float,
                        default=None,
                        help='INT8 TOPS (default: 2 * tops)')
    parser.add_argument('--fp8_tops', type=float, default=None, help='FP8 TOPS (default: 2 * tops)')
    parser.add_argument('-v',
                        '--vector_tops',
                        type=float,
                        default=None,
                        help='Vector TOPS (default: tops/8)')
    parser.add_argument('-r',
                        '--uarch_rate',
                        type=float,
                        default=None,
                        help='uArch Rate (default: 0.8)')
    parser.add_argument('-u',
                        '--bw_util',
                        type=float,
                        default=None,
                        help='Bandwidth utilization (default: 0.8)')
    parser.add_argument('-p',
                        '--parallelism',
                        type=float,
                        default=None,
                        help='Parallelism ratio (default: 0.5)')
    parser.add_argument('--max_pixels',
                        type=str,
                        default="",
                        help='max input pixels for vision models')
    parser.add_argument('-o',
                        '--out_dir',
                        required=True,
                        help='folder to collect the renamed .xlsx workbooks into')
    parser.add_argument('--prefix',
                        default="",
                        help='filename prefix for collected workbooks (default: out_dir basename)')
    parser.add_argument('--mode',
                        choices=["auto", "cross", "paired", "dedup"],
                        default="auto",
                        help='grid mode: auto=per-seq sweep, cross=full cross product, '
                        'paired=zip seq/mil, dedup=minimal runs + reconstruct the rest '
                        '(default: auto)')
    parser.add_argument(
        '--only_result',
        action='store_true',
        help='delete per-run tmp_mlir_analyse (and run dir) after collecting the xlsx')
    parser.add_argument('--no_summary',
                        action='store_true',
                        help='skip recomputing TTFT/Tokens/s and writing the summary CSV')
    args = parser.parse_args()

    if args.mode in ("cross", "paired") and args.max_input_length is None:
        print(f"Error: --mode {args.mode} requires --max_input_length.", file=sys.stderr)
        sys.exit(1)

    os.makedirs(args.out_dir, exist_ok=True)
    prefix = args.prefix or os.path.basename(os.path.normpath(args.out_dir))

    grid = build_report_grid(args)
    if not grid:
        print("Error: no valid configurations to run.", file=sys.stderr)
        sys.exit(1)

    if args.mode == "dedup":
        run_set = build_dedup_runs(grid)
        print(f"Mode dedup: report {len(grid)} row(s), run {len(run_set)} "
              f"(minimal cover: prefill by mil, decode by seq).")
    else:
        run_set = list(grid)
        if args.mode == "auto" and args.max_input_length is None:
            per_seq = {s: auto_max_input_lengths(s) for s in args.seq_length}
            print("Mode auto (max_input_length sweep 256 -> seq/2 per seq):")
            for s in args.seq_length:
                print(f"  seq={s:>6}  mil={per_seq[s]}")
        else:
            print(f"Mode {args.mode}: {len(grid)} configuration(s).")
    run_set_keys = set(run_set)

    print(f"Batch run: {len(run_set)} run(s) -> {args.out_dir}")
    rec_by_run = {}  # (seq, mil) -> rec dict
    dest_by_run = {}  # (seq, mil) -> xlsx path
    for idx, (seq, mil) in enumerate(run_set, 1):
        tag = f"{prefix}_s{seq}_mil{mil}"
        run_out_dir = os.path.join(args.out_dir, f"tmp_s{seq}_mil{mil}")
        if os.path.isdir(run_out_dir):
            shutil.rmtree(run_out_dir)
        os.makedirs(run_out_dir, exist_ok=True)
        cmd = build_analyse_cmd(args, seq, mil, run_out_dir)
        print(f"\n[{idx}/{len(run_set)}] seq={seq} max_input_length={mil}")
        print("  $ " + " ".join(cmd))
        try:
            subprocess.check_call(cmd)
        except subprocess.CalledProcessError as e:
            print(f"  FAILED (exit {e.returncode}); skipping.", file=sys.stderr)
            if args.only_result:
                cleanup_run_dir(run_out_dir)
            continue

        generated = os.path.join(run_out_dir, f"{os.path.basename(run_out_dir)}.xlsx")
        dest = os.path.join(args.out_dir, f"{tag}.xlsx")
        if not os.path.isfile(generated):
            print(f"  WARNING: expected workbook not found at {generated}", file=sys.stderr)
            dest = None
        else:
            shutil.move(generated, dest)
            print(f"  saved: {dest}")
        dest_by_run[(seq, mil)] = dest

        rec = None
        if not args.no_summary and dest and os.path.isfile(dest):
            m = recompute_metrics(dest)
            if m is not None:
                # prefill processes (max_input_length or seq_length) text tokens;
                # vit is accounted separately, so exclude it (and its CPU preproc)
                # from the text-prefill time used for prefill TPS.
                prefill_tokens = mil if mil > 0 else seq
                ttft = m["prefill_time"]
                vit_time = m["vit_time"]
                text_prefill_time = ttft - (vit_time or 0.0) - (m["preproc"] if vit_time else 0.0)
                prefill_tps = prefill_tokens / text_prefill_time if text_prefill_time > 0 else 0.0
                decode_tps = 1.0 / m["decode_time"] if m["decode_time"] > 0 else 0.0
                vit_fps = 1.0 / vit_time if vit_time else None
                rec = dict(
                    ttft=ttft,
                    decode_time=m["decode_time"],
                    vit_time=vit_time,
                    preproc=m["preproc"],
                    text_prefill_time=text_prefill_time,
                    prefill_tps=prefill_tps,
                    decode_tps=decode_tps,
                    vit_fps=vit_fps,
                    reconstructed=False,
                )
                vit_str = "-" if vit_fps is None else f"{vit_fps:.2f}"
                print(f"  TTFT={ttft:.6f}s  prefill={prefill_tps:.2f} tok/s  "
                      f"decode={decode_tps:.3f} tok/s  vit={vit_str} fps")
        rec_by_run[(seq, mil)] = rec

        if args.only_result:
            cleanup_run_dir(run_out_dir)

    # ---- Build dedup lookups from the measured runs ----
    lookups = {"text_prefill_by_mil": {}, "decode_by_seq": {}, "vit_time": None, "preproc": 0.0}
    for (s, m), rec in rec_by_run.items():
        if rec is None:
            continue
        if m > 0 and m not in lookups["text_prefill_by_mil"]:
            lookups["text_prefill_by_mil"][m] = rec["text_prefill_time"]
        if s not in lookups["decode_by_seq"]:
            lookups["decode_by_seq"][s] = rec["decode_time"]
        if lookups["vit_time"] is None and rec["vit_time"] is not None:
            lookups["vit_time"] = rec["vit_time"]
            lookups["preproc"] = rec["preproc"]

    # ---- Assemble per-row records over the full report grid ----
    rows = []
    for (seq, mil) in grid:
        if (seq, mil) in rec_by_run and rec_by_run[(seq, mil)] is not None:
            rec = rec_by_run[(seq, mil)]
            reconstructed = False
            dest = dest_by_run.get((seq, mil))
        elif (seq, mil) in rec_by_run:
            rec = None  # run failed
            reconstructed = False
            dest = dest_by_run.get((seq, mil))
        else:
            rec = reconstruct_metrics(seq, mil, lookups) if not args.no_summary else None
            reconstructed = rec is not None
            dest = None
        rows.append(dict(seq=seq, mil=mil, dest=dest, rec=rec, reconstructed=reconstructed))

    # ---- Effective hardware params (defaults applied; matches what the
    # workbooks actually used) ----
    eff_int8 = args.int8_tops if args.int8_tops is not None else args.tops * 2.0
    eff_fp8 = args.fp8_tops if args.fp8_tops is not None else args.tops * 2.0
    eff_vec = args.vector_tops if args.vector_tops is not None else args.tops / 8.0
    eff_uarch = args.uarch_rate if args.uarch_rate is not None else 0.8
    eff_bwutil = args.bw_util if args.bw_util is not None else 0.8
    eff_par = args.parallelism if args.parallelism is not None else 0.5

    # ---- Summary ----
    bar = "=" * 82
    print("\n" + bar)
    print(f"Collected {sum(1 for r in rows if r['dest'])} workbook(s) in {args.out_dir} "
          f"({sum(1 for r in rows if r['reconstructed'])} reconstructed)")
    print("Common parameters:")
    print(f"  chip={args.chip}  quant={args.quantize}  "
          f"fp16={args.tops}T int8={eff_int8}T fp8={eff_fp8}T vec={eff_vec}T")
    print(f"  bandwidth={args.bandwidth}GB/s  uarch={eff_uarch}  "
          f"bw_util={eff_bwutil}  parallelism={eff_par}")
    print(bar)
    if rows:

        def _fmt(v, p):
            return f"{v:.{p}f}" if v is not None else "-"

        print(f"{'seq':>6} {'max_in':>7} {'pre_TPS':>10} {'dec_TPS':>10} "
              f"{'vit_FPS':>9} {'TTFT(s)':>11} {'rec':>3}  workbook")
        for r in rows:
            rec = r["rec"]
            pt = _fmt(rec["prefill_tps"], 2) if rec else "-"
            dt = _fmt(rec["decode_tps"], 3) if rec else "-"
            vf = _fmt(rec["vit_fps"], 2) if rec and rec["vit_fps"] is not None else "-"
            tt = _fmt(rec["ttft"], 6) if rec else "-"
            rc = "*" if r["reconstructed"] else ""
            wb = os.path.basename(r["dest"]) if r["dest"] else "(reconstructed)"
            print(f"{r['seq']:>6} {r['mil']:>7} {pt:>10} {dt:>10} {vf:>9} {tt:>11} {rc:>3}  {wb}")

    if not args.no_summary and rows:
        csv_path = os.path.join(args.out_dir, f"{prefix}_summary.csv")
        with open(csv_path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow([
                "seq_length", "max_input_length", "prefill_tps", "decode_tps", "vit_fps", "ttft_s",
                "decode_time_s", "vit_time_s", "reconstructed", "chip", "quantize", "fp16_tops",
                "int8_tops", "fp8_tops", "vector_tops", "bandwidth_gbps", "uarch_rate", "bw_util",
                "parallelism", "xlsx"
            ])
            for r in rows:
                rec = r["rec"]
                w.writerow([
                    r["seq"], r["mil"],
                    f"{rec['prefill_tps']:.6f}" if rec and rec["prefill_tps"] is not None else "",
                    f"{rec['decode_tps']:.6f}" if rec and rec["decode_tps"] is not None else "",
                    f"{rec['vit_fps']:.6f}" if rec and rec["vit_fps"] is not None else "",
                    f"{rec['ttft']:.6f}" if rec and rec["ttft"] is not None else "",
                    f"{rec['decode_time']:.6f}" if rec and rec["decode_time"] is not None else "",
                    f"{rec['vit_time']:.6f}" if rec and rec["vit_time"] is not None else "",
                    "Y" if r["reconstructed"] else "N", args.chip, args.quantize, args.tops,
                    eff_int8, eff_fp8, eff_vec, args.bandwidth, eff_uarch, eff_bwutil, eff_par,
                    os.path.basename(r["dest"]) if r["dest"] else ""
                ])
        print(f"Summary CSV: {csv_path}")


if __name__ == "__main__":
    main()
