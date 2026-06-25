#!/usr/bin/env python3
"""
MLIR operator analysis tool (TPU-lowered IR only).

Parses a TPU-lowered MLIR file (``module.state = "TPU_LOWERED"``) and exports an
Excel spreadsheet with:
- FLOPs and data volume for each operator
- Estimated runtime based on chip compute power and bandwidth (Roofline model)
- Focus on MatMul/Conv/FAttention/ChunkGatedDeltaRule operators

Only the TPU dialect is supported (``tpu.*`` ops). The three data-supplying
``top.*`` ops ``top.Weight`` / ``top.Input`` / ``top.None`` are tolerated as
they only carry constants/inputs and produce no compute or I/O. Every other
``top.*`` op is rejected as unsupported.

Because the lowered IR already carries the real per-tensor storage dtypes
(bf16, ui8 for packed w4/w8 weights, f8E4M3FN, ...), data volumes are computed
from each tensor's own storage size — no external dtype substitution is needed.

Usage:
    python mlir_analyse.py block_0_bm1684x_w4bf16_tpu.mlir --tops 32 --bandwidth 64
    python mlir_analyse.py block_0_bm1684x_w4bf16_tpu.mlir --tops 32 --bandwidth 64 -o result.xlsx
"""

import re
import sys
import os
import argparse
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict

# ---------------------------------------------------------------------------
# Module-state contract
# ---------------------------------------------------------------------------

# The only module state this tool understands: the IR must already be lowered
# to the TPU dialect so that every tensor carries its real storage dtype.
REQUIRED_STATE = "TPU_LOWERED"

KEY_OPS = {
    "MatMul", "Conv", "Conv2D", "FAttention", "ChunkGatedDeltaRule", "A16MatMul", "Fp8MatMul", "Mlp"
}
SKIP_OPS = {"top.Weight", "top.None", "top.Input"}


def parse_module_state(filepath: str) -> Optional[str]:
    """Return the ``module.state`` attribute value (without quotes), or None."""
    with open(filepath, "r") as f:
        # The state lives on the ``module`` line, which is always near the top.
        for line in f:
            m = re.search(r'module\.state\s*=\s*"([^"]+)"', line)
            if m:
                return m.group(1)
            if "{" in line and "module" in line:
                # module attribute block may span multiple lines; keep reading
                # a bounded window.
                continue
    return None


def require_tpu_lowered(filepath: str) -> str:
    """Validate that ``filepath`` is a TPU-lowered MLIR module.

    Returns the resolved state on success; raises ``ValueError`` otherwise so
    callers can surface a clear message instead of silently mis-analysing a
    TOP_F32 / TPU_ADDRESSED file.
    """
    state = parse_module_state(filepath)
    if state is None:
        raise ValueError(f"{filepath}: could not find 'module.state' attribute. "
                         f"This tool only supports {REQUIRED_STATE} MLIR.")
    if state != REQUIRED_STATE:
        raise ValueError(f"{filepath}: module.state is '{state}', but this tool only "
                         f"supports '{REQUIRED_STATE}'. Lower the module first "
                         f"(e.g. via model_deploy.py).")
    return state


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class TensorInfo:
    shape: List[int]
    dtype: str

    @property
    def num_elements(self) -> int:
        r = 1
        for s in self.shape:
            r *= s
        return r

    @property
    def element_size(self) -> int:
        """Size in bytes."""
        sizes = {
            "f32": 4,
            "f16": 2,
            "bf16": 2,
            "f64": 8,
            "f8E4M3FN": 1,
            "f8E5M2": 1,
            "i8": 1,
            "i16": 2,
            "i32": 4,
            "i64": 8,
            "si32": 4,
            "ui8": 1,
            "ui16": 2,
            "ui32": 4,
            "ui64": 8,
        }
        return sizes.get(self.dtype, 4)

    @property
    def size_bytes(self) -> int:
        return self.num_elements * self.element_size

    def shape_str(self) -> str:
        return "x".join(str(d) for d in self.shape)


@dataclass
class MLIROp:
    results: List[str]
    op_type: str  # e.g. "top.MatMul"
    operands: List[str]
    attributes: Dict[str, str]
    input_types: List  # TensorInfo | None
    output_types: List  # TensorInfo | None
    loc_name: str
    line_num: int


# ---------------------------------------------------------------------------
# MLIR parsing helpers
# ---------------------------------------------------------------------------


def parse_tensor_type(s: str) -> Optional[TensorInfo]:
    s = s.strip()
    if s == "none":
        return None
    m = re.match(r"tensor<(.+)>", s)
    if not m:
        return None
    inner = m.group(1)
    parts = inner.split("x")
    # Last part(s) starting with a letter form the dtype
    dtype_idx = len(parts) - 1
    for i in range(len(parts) - 1, -1, -1):
        if any(c.isalpha() for c in parts[i]):
            dtype_idx = i
            break
    shape = [int(p) for p in parts[:dtype_idx]]
    dtype = "x".join(parts[dtype_idx:])
    return TensorInfo(shape=shape, dtype=dtype)


def _split_type_list(s: str) -> List[str]:
    """Split comma separated types respecting nested < >."""
    depth = 0
    parts: List[str] = []
    cur = ""
    for ch in s:
        if ch == "<":
            depth += 1
        elif ch == ">":
            depth -= 1
        if ch == "," and depth == 0:
            parts.append(cur.strip())
            cur = ""
        else:
            cur += ch
    if cur.strip():
        parts.append(cur.strip())
    return parts


def parse_type_list(s: str) -> List:
    s = s.strip()
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
    result = []
    for t in _split_type_list(s):
        t = t.strip()
        if t == "none":
            result.append(None)
        else:
            result.append(parse_tensor_type(t))
    return result


def _parse_int_attr(val: str, default: int = 0) -> int:
    """Parse '6144 : i64' -> 6144"""
    return int(val.split(":")[0].strip())


def parse_mlir_file(filepath: str) -> Tuple[List[MLIROp], Dict[str, str], Dict[str, str]]:
    with open(filepath, "r") as f:
        content = f.read()

    # Location definitions: #loc42 = loc("...")
    loc_map: Dict[str, str] = {}
    for m in re.finditer(r"#(loc\d+)\s*=\s*loc\(\"([^\"]+)\"\)", content):
        loc_map[m.group(1)] = m.group(2)

    ops: List[MLIROp] = []

    for line_num, line in enumerate(content.split("\n"), 1):
        stripped = line.strip()
        if not stripped or stripped.startswith("//") or stripped.startswith("#"):
            continue
        if stripped.startswith("module") or stripped.startswith("func") or stripped.startswith(
                "}") or stripped.startswith("return"):
            continue

        op_match = re.search(r'"((?:top|tpu)\.\w+)"', stripped)
        if not op_match:
            continue
        op_type = op_match.group(1)

        # Enforce TPU-lowered contract: only tpu.* ops are supported, except for
        # the three data-supplying top.* ops (Weight/Input/None).
        if op_type.startswith("top.") and op_type not in SKIP_OPS:
            raise ValueError(f"{filepath}:{line_num}: unsupported op '{op_type}'. This tool "
                             f"only supports tpu.* ops (plus {sorted(SKIP_OPS)}). "
                             f"The module is not fully lowered to the TPU dialect.")

        # ---------- results ----------
        eq_pos = stripped.find("=")
        quote_pos = stripped.find('"')
        results: List[str] = []
        if eq_pos != -1 and quote_pos != -1 and eq_pos < quote_pos:
            res_str = stripped[:eq_pos].strip()
            results = [r.strip().lstrip("%") for r in res_str.split(",") if r.strip()]

        # ---------- operands ----------
        after_op = stripped[stripped.index(op_match.group(0)) + len(op_match.group(0)):]
        op_paren = re.search(r"\(([^)]*)\)", after_op)
        operands = []
        if op_paren:
            operands = [o.strip().lstrip("%") for o in op_paren.group(1).split(",") if o.strip()]

        # ---------- attributes ----------
        attrs: Dict[str, str] = {}
        # Find { } block after operands, but be careful with nested brackets in arrays
        brace_start = stripped.find("{", stripped.index(op_match.group(0)))
        if brace_start != -1:
            depth = 0
            brace_end = brace_start
            for i in range(brace_start, len(stripped)):
                if stripped[i] == "{":
                    depth += 1
                elif stripped[i] == "}":
                    depth -= 1
                    if depth == 0:
                        brace_end = i
                        break
            attr_str = stripped[brace_start + 1:brace_end]
            # Parse key = value pairs (handle arrays and strings)
            for am in re.finditer(r"(\w+)\s*=\s*(\[[^\]]*\]|\"[^\"]*\"|[^,}]+)", attr_str):
                attrs[am.group(1)] = am.group(2).strip()

        # ---------- type signature ----------
        # Pattern: : (in_types) -> out_type(s) loc(...)
        input_types: List = []
        output_types: List = []

        # Match type signature after either ')' (no attrs) or '}' (with attrs)
        type_sig = re.search(r"[})]\s*:\s*\(([^)]*)\)\s*->\s*(.+?)\s*loc\(", stripped)
        if type_sig:
            input_types = parse_type_list(type_sig.group(1))
            out_str = type_sig.group(2).strip()
            if out_str.startswith("("):
                output_types = parse_type_list(out_str)
            else:
                ot = parse_tensor_type(out_str)
                output_types = [ot] if ot else ([None] if out_str == "none" else [])
        else:
            # Ops with no operands like top.Weight, top.None
            type_sig2 = re.search(r"\(\)\s*:\s*\(\)\s*->\s*(.+?)\s*loc\(", stripped)
            if type_sig2:
                out_str = type_sig2.group(1).strip()
                ot = parse_tensor_type(out_str)
                output_types = [ot] if ot else ([None] if out_str == "none" else [])

        # ---------- loc ----------
        loc_match = re.search(r"loc\(#(loc\d+)\)", stripped)
        loc_name = ""
        if loc_match:
            loc_name = loc_map.get(loc_match.group(1), loc_match.group(1))

        ops.append(
            MLIROp(
                results=results,
                op_type=op_type,
                operands=operands,
                attributes=attrs,
                input_types=input_types,
                output_types=output_types,
                loc_name=loc_name,
                line_num=line_num,
            ))

    # Build map: SSA result name -> defining op type
    ssa_op_map: Dict[str, str] = {}
    for op in ops:
        for r in op.results:
            ssa_op_map[r] = op.op_type

    return ops, loc_map, ssa_op_map


# ---------------------------------------------------------------------------
# FLOPs calculation per operator
# ---------------------------------------------------------------------------


def _valid_inputs(op: MLIROp) -> List[TensorInfo]:
    return [t for t in op.input_types if t is not None]


def calc_matmul_flops(op: MLIROp) -> int:
    """[..., M, K] x [..., K, N] -> [..., M, N]  =>  2*batch*M*K*N

    Respects left_transpose / right_transpose attributes:
      left_transpose=true  => lhs is [..., K, M]
      right_transpose=true => rhs is [..., N, K]
    """
    vi = _valid_inputs(op)
    if len(vi) < 2:
        return 0
    lhs = vi[0]
    out = op.output_types[0] if op.output_types else None
    if not out:
        raise ValueError(f"MatMul op at line {op.line_num} has no output type")
    left_trans = op.attributes.get("left_transpose", "false") == "true"
    K = lhs.shape[-2] if left_trans else lhs.shape[-1]
    # N is always the last dim of the output; this avoids ambiguity from packed
    # weights (e.g. A16MatMul w4 qweight shape [N, K/2]).
    M = out.shape[-2]
    N = out.shape[-1]
    batch = 1
    for d in out.shape[:-2]:
        batch *= d
    return 2 * batch * M * K * N


def calc_fp8matmul_flops(op: MLIROp) -> int:
    """FP8 block-wise quantized MatMul.

    Computation phases:
      1. Dynamic quantization per block: 2 * M * K
      2. MatMul and rescale: 4 * batch * M * K * N
    """
    vi = _valid_inputs(op)
    if len(vi) < 2:
        return 0
    lhs = vi[0]
    out = op.output_types[0] if op.output_types else None
    if not out:
        raise ValueError(f"Fp8MatMul op at line {op.line_num} has no output type")

    # Determine effective K, N from lhs shape
    K = lhs.shape[-1]
    M = out.shape[-2]
    N = out.shape[-1]
    batch = 1
    for d in out.shape[:-2]:
        batch *= d

    matmul_flops = 4 * batch * M * K * N
    dynamic_quantize_flops = 2 * M * K
    return matmul_flops + dynamic_quantize_flops


def calc_conv_flops(op: MLIROp) -> int:
    vi = _valid_inputs(op)
    if len(vi) < 2:
        return 0
    weight = vi[1]
    out = op.output_types[0] if op.output_types else None
    if not out:
        return 0
    N = out.shape[0]
    OC = out.shape[1]
    out_spatial = 1
    for d in out.shape[2:]:
        out_spatial *= d
    # kernel_ops = IC/G * KH * KW  (encoded in weight shape[1:])
    kernel_ops = 1
    for d in weight.shape[1:]:
        kernel_ops *= d
    return 2 * N * OC * out_spatial * kernel_ops


def calc_fattention_flops(op: MLIROp) -> int:
    """Flash / standard attention:  Q[B,H,T,D] x K^T -> [B,H,T,T], then x V
       FLOPs ≈ 4 * B * H * T^2 * D  (two matmuls)
    """
    vi = _valid_inputs(op)
    if len(vi) < 3:
        return 0
    q = vi[0]  # [B, H, T, D]
    k = vi[1]
    if len(q.shape) < 4:
        return 0
    batch, M_q, q_head, d = q.shape[0], q.shape[1], q.shape[2], q.shape[3]
    M_k = k.shape[1]

    return batch * M_q * q_head * d * M_k * 4


def calc_chunk_gated_delta_rule_flops(op: MLIROp) -> int:
    """
    Inputs: q[B,H_k,T,D], k[B,H_k,T,D], v[B,H_v,T,D],
            alpha[B,H_v,T], beta[B,H_v,T], state[B,H_v,D,D], eye

    Per chunk (size C) per head per batch:
      Q @ K^T          : 2*C*D*C   (intra-chunk, per k-head)
      Attn @ V         : 2*C*C*D   (weighted values, per v-head)
      Q @ State        : 2*C*D*D   (query from state, per v-head)
      K^T @ V (update) : 2*D*C*D   (state update, per v-head)
      Gate + update    : ~4*D*D    (element-wise, per v-head)
    """
    vi = _valid_inputs(op)
    if len(vi) < 3:
        return 0
    v = vi[2]  # [B, S, H_v, D] (batch, seq, num_v_heads, head_dim)
    if len(v.shape) < 4:
        return 0
    B, S, H_v, D = v.shape[0], v.shape[1], v.shape[2], v.shape[3]
    C = _parse_int_attr(op.attributes.get("chunk_size", "64"), 64)
    # Q@K^T runs per k-head; the rest run per v-head (state is [B, H_v, D, D]).
    H_k = _parse_int_attr(op.attributes.get("num_k_heads", str(H_v)), H_v)
    num_chunks = (S + C - 1) // C
    per_chunk_k = 2 * C * D * C  # Q @ K^T (intra-chunk attention)
    per_chunk_v = (
        2 * C * C * D +  # Attn @ V
        2 * C * D * D +  # Q @ State
        2 * D * C * D +  # K^T @ V (state update)
        4 * D * D  # gating element-wise
    )
    return B * num_chunks * (H_k * per_chunk_k + H_v * per_chunk_v)


def calc_recurrent_gated_delta_rule_flops(op: MLIROp) -> int:
    """Recurrent (token-by-token) gated delta rule.

    Inputs: q[B,H,T,D], k[B,H,T,D], v[B,H,T,D],
            alpha[B,H,T], beta[B,H,T], state[B,H,D,D]

    Per token per head per batch:
      k @ State        : 2*D*D   (predicted value, matmul)
      delta = v - kS   :   D     (element-wise sub)
      beta * delta     :   D     (element-wise mul)
      outer k^T*delta  :   D*D   (rank-1 outer product update)
      gate * State     :   D*D   (alpha gating, element-wise)
      State += update  :   D*D   (element-wise add)
      q @ State        : 2*D*D   (output projection, matmul)
    Total = 7*D*D + 2*D per step (two 2*D*D matmuls, three D*D element-wise
    matrix ops, and two D-length element-wise ops).
    """
    vi = _valid_inputs(op)
    if len(vi) < 3:
        return 0
    v = vi[2]
    H = _parse_int_attr(op.attributes.get("num_v_heads", "0"), 0)
    D = _parse_int_attr(op.attributes.get("d", "0"), 0)
    # Decode path: v is 2D [B, H*D] (one token). Prefer the recurrent_state
    # tensor vi[5] [B, H, D, D] for B/H/D when available; fall back to attrs.
    if len(vi) >= 6 and vi[5].shape and len(vi[5].shape) >= 4:
        st = vi[5]
        B = st.shape[0]
        H = st.shape[1] if H == 0 else H
        D = st.shape[2] if D == 0 else D
    elif v.shape and len(v.shape) >= 4:
        B, _, H_v, D_v = v.shape[0], v.shape[1], v.shape[2], v.shape[3]
        H = H if H else H_v
        D = D if D else D_v
    else:
        # 2D decode value [B, H*D]; T = 1.
        B = v.shape[0] if v.shape else 0
    if H == 0 or D == 0:
        return 0
    T = 1
    # k@State (2D^2) + q@State (2D^2) + outer k^T*delta (D^2) + gate*State (D^2)
    # + State+=update (D^2) + delta sub (D) + beta*delta (D) = 7*D^2 + 2*D
    per_step = 7 * D * D + 2 * D
    return B * H * T * per_step


def calc_topk_flops(op: MLIROp) -> int:
    """TopK selection. Operands: [x] -> (values, indices).

    Cost is dominated by the partial sort/scan over the last axis of size N to
    extract the K largest entries. We approximate it as K comparisons per
    element of the reduction axis, i.e. K * N * (output volume / N).

    For a 1x10240x256 input with K=8 this yields 8 * 256 * 10240 comparisons,
    which is a conservative order-of-magnitude estimate.
    """
    vi = _valid_inputs(op)
    if not vi:
        return 0
    inp = vi[0]
    if not inp.shape:
        return 0
    K = _parse_int_attr(op.attributes.get("K", "1"), 1)
    # num_elements excluding the reduced axis
    axis = _parse_int_attr(op.attributes.get("axis", "-1"), -1)
    rank = len(inp.shape)
    if axis < 0:
        axis += rank
    if axis < 0 or axis >= rank:
        return 0
    N = inp.shape[axis]
    outer = 1
    for i, d in enumerate(inp.shape):
        if i != axis:
            outer *= d
    # comparisons ~ K * N per outer element (partial selection cost)
    return K * N * outer


def calc_mlp_flops(op: MLIROp) -> int:
    """Fused MLP / MoE expert MLP.

    The op folds three GEMMs (gate, up, down) plus a SiLU and a Mul on the
    gate/up outputs:
      gate: x[M, K] @ W_gate -> [M, N]   (SiLU-gated)
      up:   x[M, K] @ W_up   -> [M, N]
      down: (silu(gate) * up)[M, N] @ W_down -> [M, K]

    For a GEMM x[M,K] @ W[K,N] the FLOPs are 2*M*K*N = 2*M*(K*N), i.e. twice
    the number of (logical) weight elements times M. We exploit this so we do
    not have to disambiguate which stored axis is K vs N (which depends on
    right_transpose_* and on w4 packing of the last axis).

    MoE (is_expert=true): weights carry a leading num_expert axis, but each
    token is dispatched to only num_expert_per_tok experts, so M already equals
    (tokens * num_expert_per_tok) taken from the output shape. We therefore
    drop the expert axis when counting per-GEMM weight elements.
    """
    # Operand order: input, weight_gate, scale_gate, zp_gate, bias_gate,
    # weight_up, scale_up, zp_up, bias_up, weight_down, scale_down, zp_down,
    # bias_down, expert_id, buffer. Some operands (bias_*, buffer) may be None,
    # so index op.input_types directly rather than _valid_inputs(), which drops
    # None entries and shifts indices.
    n_in = len(op.input_types)
    if n_in < 10:
        return 0
    x = op.input_types[0]
    gate_w = op.input_types[1]
    up_w = op.input_types[5]
    down_w = op.input_types[9]
    if not (x and gate_w and up_w and down_w and x.shape and gate_w.shape and up_w.shape
            and down_w.shape):
        return 0
    is_expert = op.attributes.get("is_expert", "false") == "true"
    is_quant = op.attributes.get("quantized", "false") == "true"
    wbits = _parse_int_attr(op.attributes.get("weight_bits", "4"), 4)
    # ui8 storage packs 2 w4 weights per byte; w8 is 1:1.
    pack = 2 if (is_quant and wbits == 4) else 1

    # M = routed token count; K = hidden size (output last dim).
    # For MoE the output is [tokens, num_expert_per_tok, K], so the leading
    # dims already encode the per-expert dispatch.
    out = op.output_types[0] if op.output_types else None
    if out and out.shape:
        M = 1
        for d in out.shape[:-1]:
            M *= d
        K_hidden = out.shape[-1]
    else:
        M = 1
        for d in x.shape[:-1]:
            M *= d
        K_hidden = x.shape[-1]

    def _logical_elems(w):
        # Logical weight elements per (expert) unit: drop the leading expert
        # axis for MoE weights, then unpack w4 (2 weights per stored byte).
        dims = list(w.shape)
        if is_expert and len(dims) > 2:
            dims = dims[1:]
        n = 1
        for d in dims:
            n *= d
        return n * pack

    gate_elems = _logical_elems(gate_w)
    up_elems = _logical_elems(up_w)
    down_elems = _logical_elems(down_w)

    flops = 0
    flops += 2 * M * gate_elems  # gate GEMM
    flops += 2 * M * up_elems  # up GEMM
    flops += 2 * M * down_elems  # down GEMM
    # Element-wise SiLU on gate output (~6 ops/elem) + Mul with up output
    # (~1 op/elem), operating on the [M, N_intermediate] intermediate tensor.
    # N_intermediate = gate weight logical elems / K_hidden.
    n_inter = gate_elems // K_hidden if K_hidden else 0
    flops += 7 * M * n_inter
    return flops


def _out_elements(op: MLIROp) -> int:
    """Return num_elements of the first output tensor, or 0."""
    out = op.output_types[0] if op.output_types else None
    return out.num_elements if out else 0


def _in_elements(op: MLIROp, idx: int = 0) -> int:
    """Return num_elements of input[idx], or 0."""
    vi = _valid_inputs(op)
    return vi[idx].num_elements if idx < len(vi) else 0


def _has_do_relu(op: MLIROp) -> bool:
    return op.attributes.get("do_relu", "false") == "true"


def calc_flops(op: MLIROp) -> int:
    name = op.op_type.split(".")[-1]

    # ---- Complex operators with dedicated calculators ----
    if name == "MatMul":
        return calc_matmul_flops(op)
    elif name == "A16MatMul":
        # Weight-only quantized MatMul; FLOPs same as regular MatMul.
        # lhs [..., M, K]; rhs qweight is [N, K] (w8) or [N, K/2] (w4, ui8 packed).
        # calc_matmul_flops uses lhs last dim as K and rhs[-2] as N (right_transpose),
        # which is correct for both w4 and w8.
        return calc_matmul_flops(op)
    elif name == "Fp8MatMul":
        # FP8 block-wise quantized MatMul.
        # Includes matmul FLOPs + weight dequantization overhead.
        return calc_fp8matmul_flops(op)
    elif name in ("Conv", "Conv2D"):
        return calc_conv_flops(op)
    elif name == "FAttention":
        return calc_fattention_flops(op)
    elif name == "ChunkGatedDeltaRule":
        return calc_chunk_gated_delta_rule_flops(op)
    elif name == "RecurrentGatedDeltaRule":
        return calc_recurrent_gated_delta_rule_flops(op)
    elif name == "Mlp":
        return calc_mlp_flops(op)
    elif name == "TopK":
        return calc_topk_flops(op)
    # ---- Normalization layers ----
    # RMSNorm: 3 * num_elements
    elif name == "RMSNorm":
        return 3 * _out_elements(op)
    # LayerNorm / GroupNorm / InstanceNorm / PixelNorm:
    #   num_elements * (10 + have_weight + have_bias)
    #   In MLIR, weight/bias are operands; we approximate as 10+2=12 when present.
    elif name in ("LayerNorm", "GroupNorm", "InstanceNorm", "PixelNorm"):
        vi = _valid_inputs(op)
        # inputs beyond the first are weight and bias
        have_weight = 1 if len(vi) > 1 else 0
        have_bias = 1 if len(vi) > 2 else 0
        return _out_elements(op) * (10 + have_weight + have_bias)
    # BatchNorm: 2 * num_elements
    elif name == "BatchNorm":
        return 2 * _out_elements(op)
    # Softmax: (5 + log?) * num_elements
    elif name == "Softmax":
        log_flag = 1 if op.attributes.get("log", "false") == "true" else 0
        return _in_elements(op) * (5 + log_flag)

    # ---- tpu.Active: fused activation selected by `mode` attr ----
    # mode = #tpu<active_mode NAME>; map NAME to the same multipliers used
    # for the corresponding top-dialect activation op below.
    elif name == "Active":
        mode = op.attributes.get("mode", "")
        m = re.search(r"active_mode\s+(\w+)", mode)
        act = m.group(1) if m else ""
        if act in ("SILU", "SWISH", "HARDSWISH", "GELU", "TGELU"):
            return 5 * _out_elements(op)
        if act in ("SIGMOID", "TANH", "EXP", "LOG", "SIN", "COS", "TAN", "SINH", "COSH", "ARCTANH",
                   "LOGB", "HARDSIGMOID", "MISH"):
            return 4 * _out_elements(op)
        if act == "SOFT_PLUS":
            return 3 * _out_elements(op)
        return _out_elements(op)

    # ---- Activation functions (multiplier per element) ----
    # 5x: SiLU, Swish, HardSwish, GELU
    elif name in ("SiLU", "Swish", "HardSwish", "GELU"):
        return 5 * _out_elements(op)
    # 4x: Sigmoid, Tanh, Exp, Log, Sin, Cos, Tan, Sinh, Cosh,
    #     Arccos, Arctanh, LogB, HardSigmoid, Mish, Rope
    elif name in ("Sigmoid", "Tanh", "Exp", "Log", "Sin", "Cos", "Tan", "Sinh", "Cosh", "Arccos",
                  "Arctanh", "LogB", "HardSigmoid", "Mish", "Rope"):
        return 4 * _out_elements(op)
    # 3x: Softplus, RMSNorm(above), Sign, Softsign
    elif name == "Softplus":
        return 3 * _in_elements(op)
    elif name in ("Sign", "Softsign"):
        return 3 * _in_elements(op)

    # ---- Binary / element-wise with do_relu ----
    # Add / Mul / Sub: num_elements * (num_inputs - 1 + do_relu)
    elif name in ("Add", "Mul", "Sub"):
        n_inputs = len(_valid_inputs(op))
        do_relu = 1 if _has_do_relu(op) else 0
        return _out_elements(op) * max(n_inputs - 1 + do_relu, 1)
    # AddConst / SubConst / MulConst / Reciprocal: num_elements * (1 + do_relu)
    elif name in ("AddConst", "SubConst", "MulConst", "Reciprocal"):
        do_relu = 1 if _has_do_relu(op) else 0
        return _out_elements(op) * (1 + do_relu)
    # Scale: num_elements * (2 + do_relu)
    elif name == "Scale":
        do_relu = 1 if _has_do_relu(op) else 0
        return _out_elements(op) * (2 + do_relu)
    # Upsample: num_elements * (2 if do_relu else 1)
    elif name == "Upsample":
        return _out_elements(op) * (2 if _has_do_relu(op) else 1)

    # ---- 2x per element ----
    # Rsqrt, Clip, Normalize, BinaryShift, BinaryConstShift, LogicalAnd,
    # BatchNorm(above), Variance, WeightReorder, MeanRstd
    elif name in ("Rsqrt", "Clip", "Normalize", "BinaryShift", "BinaryConstShift", "LogicalAnd",
                  "Variance", "WeightReorder"):
        return 2 * _out_elements(op)

    # ---- 1x per element (simple element-wise) ----
    elif name in ("Abs", "Relu", "Elu", "PRelu", "LeakyRelu", "Lut", "Sqrt", "Ceil", "Floor",
                  "Round", "Pow", "Pow2", "Pow3", "Erf", "Mod", "Cast", "Copy", "Div", "DivConst",
                  "Max", "Min", "MaxConst", "MinConst", "Compare", "CompareConst", "Where",
                  "MaskedFill", "Trilu", "Reverse", "Reduce", "QuantizeLinear", "DequantizeLinear",
                  "Csc", "DtypeCast", "SwapDimInner"):
        return _out_elements(op)

    # ---- Pooling: output_elements * kernel_size ----
    elif name in ("AvgPool", "MaxPool", "MaxPoolWithMask", "AdaptiveAvgPool"):
        out_elem = _out_elements(op)
        ks = _parse_int_attr(op.attributes.get("kernel_shape", ""), 0)
        if ks == 0:
            # Try to get kernel from attributes like kernel_h, kernel_w
            kh = _parse_int_attr(op.attributes.get("kernel_h", "1"), 1)
            kw = _parse_int_attr(op.attributes.get("kernel_w", "1"), 1)
            ks = kh * kw
        do_relu = 1 if _has_do_relu(op) else 0
        return out_elem * (ks + do_relu) if ks > 0 else out_elem

    # ---- Attention variants ----
    elif name == "Attention":
        return 0  # unimplemented in C++

    # ---- RoiAlign / RoiExtractor / PoolMask: 4 * num_elements ----
    elif name in ("RoiAlign", "RoiExtractor", "PoolMask"):
        return 4 * _out_elements(op)

    # ---- Interp: mode-dependent ----
    elif name == "Interp":
        mode = op.attributes.get("mode", "nearest")
        return _out_elements(op) * (1 if "nearest" in mode else 2)

    # ---- Arg: input elements ----
    elif name == "Arg":
        return _in_elements(op)

    # ---- Sort: n^2 ----
    elif name == "Sort":
        n = _in_elements(op)
        return n * n

    # ---- DequantInt / RequantInt: mode-dependent ----
    elif name in ("DequantInt", "RequantInt"):
        qmode = op.attributes.get("quant_mode", "Normal")
        mult = 3 if "Normal" in qmode else 5
        return _out_elements(op) * mult

    # ---- Zero FLOPs (shape manipulation / data movement) ----
    elif name in ("Reshape", "Permute", "Slice", "SliceAxis", "StridedSlice", "Weight", "Input",
                  "None", "Transpose", "Expand", "Repeat", "Concat", "Split", "Gather",
                  "GatherElements", "GatherND", "ScatterND", "ScatterElements", "Flatten",
                  "Squeeze", "Unsqueeze", "View", "Pad", "Tile", "Shape", "Size", "Range", "Arange",
                  "Depth2Space", "ShuffleChannel", "Pack", "Unpack", "ConstantFill", "RandnLike",
                  "MeshGrid", "Loop", "If", "List", "Custom", "Correlation", "ConcatSlice", "Nms",
                  "Einsum", "RequantFp", "IndexPut", "MaskRCNNGetBboxB"):
        return 0

    # ---- Fallback: 0 ----
    return 0


# ---------------------------------------------------------------------------
# Data volume (TPU-lowered, dtype-aware)
# ---------------------------------------------------------------------------

# No external dtype map is needed: in a TPU-lowered module every tensor already
# carries its real storage dtype (bf16, ui8 for packed w4/w8 weights, f8E4M3FN,
# ...), so each tensor's own ``size_bytes`` is the authoritative storage size.


def calc_data_volume(op: MLIROp, ssa_op_map: Optional[Dict[str, str]] = None) -> Tuple[int, int]:
    """Return ``(read_bytes, write_bytes)`` for a TPU-lowered op.

    The lowered IR carries the real per-tensor storage dtypes, so data volumes
    are read directly from each tensor's own ``size_bytes`` — no external
    ``dtype_mode`` substitution is required. ``ssa_op_map`` is retained for
    API compatibility but is no longer used (weight detection was only needed
    on the old top-dialect path).
    """
    opn = op.op_type.split(".")[-1]

    # --- write bytes ---
    wb = sum(t.size_bytes for t in op.output_types if isinstance(t, TensorInfo))

    # --- read bytes ---
    # Shape / data-movement ops that do not re-read their inputs from memory:
    if opn == "Concat":
        # only_merge=true means pure memory aliasing, no real I/O.
        if op.attributes.get("only_merge", "false") == "true":
            return 0, 0
        return 0, wb
    if opn in ("Reshape", "Slice"):
        return 0, 0
    if opn in ("Permute", "Gather"):
        return 0, wb

    rb = 0
    # MoE expert MLP: weights and their scales/zp carry a leading num_expert
    # axis, but each token is dispatched to only num_expert_per_tok of the
    # num_expert experts, so only that fraction of the weight store is read.
    # Scale those inputs by the dispatch ratio; the activation and expert_id
    # (no leading expert axis) are read in full.
    if opn == "Mlp" and op.attributes.get("is_expert", "false") == "true":
        num_expert = _parse_int_attr(op.attributes.get("num_expert", "1"), 1)
        num_expert_per_tok = _parse_int_attr(op.attributes.get("num_expert_per_tok", "1"), 1)
        ratio = num_expert_per_tok / num_expert if num_expert else 1.0
        for t in op.input_types:
            if not isinstance(t, TensorInfo):
                continue
            bytes_i = t.size_bytes
            if t.shape and t.shape[0] == num_expert:
                bytes_i = bytes_i * ratio
            rb += bytes_i
        return int(rb), wb

    for t in op.input_types:
        if isinstance(t, TensorInfo):
            rb += t.size_bytes
    return rb, wb


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------


def fmt_num(n: float) -> str:
    if n == 0:
        return "0"
    for thresh, suffix in [(1e12, "T"), (1e9, "G"), (1e6, "M"), (1e3, "K")]:
        if n >= thresh:
            return f"{n / thresh:.2f}{suffix}"
    return f"{n:.0f}"


def fmt_bytes(n: float) -> str:
    if n == 0:
        return "0 B"
    for thresh, suffix in [(1e9, "GB"), (1e6, "MB"), (1e3, "KB")]:
        if n >= thresh:
            return f"{n / thresh:.2f} {suffix}"
    return f"{n:.0f} B"


def fmt_time(us: float) -> str:
    if us >= 1e6:
        return f"{us / 1e6:.3f} s"
    if us >= 1e3:
        return f"{us / 1e3:.3f} ms"
    return f"{us:.3f} us"


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


def export_excel(ops: List[MLIROp],
                 chip_tops: float,
                 bw_gbps: float,
                 output_path: str,
                 ssa_op_map: Optional[Dict[str, str]] = None,
                 vector_tops: Optional[float] = None,
                 uarch_rate: float = 0.8,
                 bw_util: float = 0.7):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    if vector_tops is None:
        vector_tops = chip_tops / 8.0

    compute_ops = [op for op in ops if op.op_type not in SKIP_OPS]

    # Pre-compute per-op static data (FLOPs, data volume)
    rows_data = []
    total_flops = 0
    total_read = 0
    total_write = 0

    for op in compute_ops:
        opn = op.op_type.split(".")[-1]
        flops = calc_flops(op)
        rb, wb = calc_data_volume(op, ssa_op_map)
        total_io = rb + wb

        inp_shapes = ", ".join(t.shape_str() if t else "none" for t in op.input_types)
        out_shapes = ", ".join(t.shape_str() if t else "none" for t in op.output_types)

        rows_data.append(
            dict(
                opn=opn,
                loc=op.loc_name,
                inp_shapes=inp_shapes,
                out_shapes=out_shapes,
                flops=flops,
                rb=rb,
                wb=wb,
                total_io=total_io,
                is_key=opn in KEY_OPS,
            ))
        total_flops += flops
        total_read += rb
        total_write += wb

    # ---------- Styles ----------
    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    key_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    sum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    edit_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    bold = Font(bold=True)
    thin = Border(*(Side(style="thin"), ) * 4)
    center = Alignment(horizontal="center", wrap_text=True)

    def _write_header(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin

    def _set_col_widths(ws, widths):
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # ===================== Sheet 0: Overview (create FIRST for formula refs) ===
    ws0 = wb.create_sheet("Overview")
    ws0.cell(row=1, column=1, value="Parameter").font = bold
    ws0.cell(row=1, column=2, value="Value").font = bold
    ws0.cell(row=1, column=3, value="Unit").font = bold
    # Row 2: TOPS  (Overview!B2 — editable)
    ws0.cell(row=2, column=1, value="Chip Compute Power")
    c_tops = ws0.cell(row=2, column=2, value=chip_tops)
    c_tops.fill = edit_fill
    c_tops.font = Font(bold=True, color="006100")
    c_tops.number_format = "#,##0.##"
    ws0.cell(row=2, column=3, value="TOPS")
    # Row 3: Bandwidth  (Overview!B3 — editable)
    ws0.cell(row=3, column=1, value="Chip Bandwidth")
    c_bw = ws0.cell(row=3, column=2, value=bw_gbps)
    c_bw.fill = edit_fill
    c_bw.font = Font(bold=True, color="006100")
    c_bw.number_format = "#,##0.##"
    ws0.cell(row=3, column=3, value="GB/s")
    # Row 4: Vector TOPS  (Overview!B4 — editable)
    ws0.cell(row=4, column=1, value="Vector Compute Power")
    c_vector = ws0.cell(row=4, column=2, value=vector_tops)
    c_vector.fill = edit_fill
    c_vector.font = Font(bold=True, color="006100")
    c_vector.number_format = "#,##0.##"
    ws0.cell(row=4, column=3, value="TOPS")
    # Row 5: uArch Rate  (Overview!B5 — editable)
    ws0.cell(row=5, column=1, value="uArch Rate")
    c_cu = ws0.cell(row=5, column=2, value=uarch_rate)
    c_cu.fill = edit_fill
    c_cu.font = Font(bold=True, color="006100")
    c_cu.number_format = "0%"
    ws0.cell(row=5, column=3, value="")
    # Row 6: Bandwidth Utilization  (Overview!B6 — editable)
    ws0.cell(row=6, column=1, value="Bandwidth Utilization")
    c_bu = ws0.cell(row=6, column=2, value=bw_util)
    c_bu.fill = edit_fill
    c_bu.font = Font(bold=True, color="006100")
    c_bu.number_format = "0%"
    ws0.cell(row=6, column=3, value="")

    # Row 8: blank
    # Row 9+: static info
    info_row = 9
    static_items = [
        ("MLIR File", os.path.basename(output_path).replace("_analysis.xlsx", ".mlir")),
        ("Module State", REQUIRED_STATE),
        ("", ""),
        ("Total GOPs", total_flops / 1e9),
        ("Total Data Read (GB)", total_read / 1e9),
        ("Total Data Write (GB)", total_write / 1e9),
        ("Total Data Volume (GB)", (total_read + total_write) / 1e9),
        ("", ""),
        ("Note", "Modify B2-B6 (green cells) to update all performance estimates."),
        ("Note", "Green cells are editable parameters."),
        ("Note", "Roofline model: Est.Time = max(FLOPs/ComputePower, DataVolume/Bandwidth)"),
        ("Note", "Data volumes read the real per-tensor dtypes from the TPU-lowered IR."),
        ("Note", "ChunkGatedDeltaRule FLOPs are approximate (intra-chunk attn + state update)"),
    ]
    for r_off, (k, v) in enumerate(static_items):
        ws0.cell(row=info_row + r_off, column=1, value=k)
        cell = ws0.cell(row=info_row + r_off, column=2, value=v)
        if isinstance(v, (int, float)) and v > 1000:
            cell.number_format = "#,##0"
    ws0.column_dimensions["A"].width = 24
    ws0.column_dimensions["B"].width = 50
    ws0.column_dimensions["C"].width = 8

    # Formula building helpers — reference Overview!B2 (TOPS), B3 (BW), B4 (Vector TOPS), B5 (Compute Util), B6 (BW Util)
    # Cell E stores GOPs (= FLOPs / 1e9)
    # Cells F,G,H store MB (= bytes / 1e6)
    # Compute(us) = GOPs / (TOPS * ComputeUtil) * 1000
    # Memory(us) = IO_MB / (BW * BWUtil) * 1000
    TOPS_REF = "Overview!$B$2"
    BW_REF = "Overview!$B$3"
    VECTOR_TOPS_REF = "Overview!$B$4"
    CU_REF = "Overview!$B$5"
    BU_REF = "Overview!$B$6"

    def _compute_formula(gops_cell, tops_ref=TOPS_REF):
        return f"=IF({tops_ref}=0,0,{gops_cell}/({tops_ref}*{CU_REF})*1000)"

    def _memory_formula(io_cell):
        return f"=IF({BW_REF}=0,0,{io_cell}/({BW_REF}*{BU_REF})*1000)"

    def _est_time_formula(comp_cell, mem_cell):
        return f"=MAX({comp_cell},{mem_cell})"

    def _bottleneck_formula(comp_cell, mem_cell):
        return f'=IF({comp_cell}>={mem_cell},"Compute","Memory")'

    def _ai_formula(gops_cell, io_cell):
        return f"=IF({io_cell}=0,0,{gops_cell}*1000/{io_cell})"

    # ===================== Sheet 1: All Operators =====================
    ws1 = wb.active
    ws1.title = "All Operators"
    h1 = [
        "No.",
        "Op Type",
        "Input Shapes",
        "Output Shapes",
        "GOPs",
        "Read (MB)",
        "Write (MB)",
        "Total I/O (MB)",
        "Compute (us)",
        "Memory (us)",
        "Est. Time (us)",
        "Bottleneck",
        "Arith. Intensity (F/B)",
        "Name",
    ]
    _write_header(ws1, h1)

    # Column letters: E=GOPs, H=TotalIO, I=Compute, J=Memory, K=Est, L=Bottleneck, M=AI
    for idx, d in enumerate(rows_data, 1):
        r = idx + 1
        # Static columns A-H
        for c, v in enumerate([
                idx,
                d["opn"],
                d["inp_shapes"],
                d["out_shapes"],
                d["flops"] / 1e9,
                d["rb"] / 1e6,
                d["wb"] / 1e6,
                d["total_io"] / 1e6,
        ], 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin
            if d["is_key"]:
                cell.fill = key_fill
            if c in (5, 6, 7, 8):
                cell.number_format = "#,##0.000"

        # Formula columns I-M
        # I: Compute(us) — use TOPS for key ops, Vector TOPS for others
        tops_ref = TOPS_REF if d["is_key"] else VECTOR_TOPS_REF
        cell_i = ws1.cell(row=r, column=9)
        cell_i.value = _compute_formula(f"E{r}", tops_ref)
        cell_i.number_format = "#,##0.000"
        cell_i.border = thin
        # J: Memory(us)
        cell_j = ws1.cell(row=r, column=10)
        cell_j.value = _memory_formula(f"H{r}")
        cell_j.number_format = "#,##0.000"
        cell_j.border = thin
        # K: Est.Time = MAX(I, J)
        cell_k = ws1.cell(row=r, column=11)
        cell_k.value = _est_time_formula(f"I{r}", f"J{r}")
        cell_k.number_format = "#,##0.000"
        cell_k.border = thin
        # L: Bottleneck
        cell_l = ws1.cell(row=r, column=12)
        cell_l.value = _bottleneck_formula(f"I{r}", f"J{r}")
        cell_l.border = thin
        # M: Arith.Intensity
        cell_m = ws1.cell(row=r, column=13)
        cell_m.value = _ai_formula(f"E{r}", f"H{r}")
        cell_m.number_format = "#,##0.00"
        cell_m.border = thin
        # N: Name
        cell_name = ws1.cell(row=r, column=14, value=d["loc"])
        cell_name.border = thin

        if d["is_key"]:
            for c in range(9, 15):
                ws1.cell(row=r, column=c).fill = key_fill

    # Summary row
    sr = len(rows_data) + 3
    for c in range(1, len(h1) + 1):
        cell = ws1.cell(row=sr, column=c)
        cell.fill = sum_fill
        cell.font = bold
        cell.border = thin
    ws1.cell(row=sr, column=1, value="TOTAL")
    first_data = 2
    last_data = len(rows_data) + 1
    ws1.cell(row=sr, column=5,
             value=f"=SUM(E{first_data}:E{last_data})").number_format = "#,##0.000"
    ws1.cell(row=sr, column=6,
             value=f"=SUM(F{first_data}:F{last_data})").number_format = "#,##0.000"
    ws1.cell(row=sr, column=7,
             value=f"=SUM(G{first_data}:G{last_data})").number_format = "#,##0.000"
    ws1.cell(row=sr, column=8,
             value=f"=SUM(H{first_data}:H{last_data})").number_format = "#,##0.000"
    ws1.cell(row=sr, column=11,
             value=f"=SUM(K{first_data}:K{last_data})").number_format = "#,##0.000"
    ws1.freeze_panes = "A2"
    # No. OpType InShapes OutShapes GOPs Read Write TotalIO Comp Mem Est Bottleneck AI Name
    _set_col_widths(ws1, [5, 18, 22, 16, 10, 10, 10, 12, 12, 12, 12, 11, 10, 20])

    # ===================== Sheet 2: Key Operators =====================
    ws2 = wb.create_sheet("Key Operators")
    h2 = [
        "No.",
        "Op Type",
        "Input Shapes",
        "Output Shapes",
        "GOPs",
        "Read (MB)",
        "Write (MB)",
        "Total I/O (MB)",
        "Compute (us)",
        "Memory (us)",
        "Est. Time (us)",
        "Bottleneck",
        "Arith. Intensity",
        "% Total GOPs",
        "% Total Time",
        "Name",
    ]
    _write_header(ws2, h2)

    key_rows = [d for d in rows_data if d["is_key"]]
    for ki, d in enumerate(key_rows, 1):
        r = ki + 1
        # Static columns A-H
        for c, v in enumerate([
                ki,
                d["opn"],
                d["inp_shapes"],
                d["out_shapes"],
                d["flops"] / 1e9,
                d["rb"] / 1e6,
                d["wb"] / 1e6,
                d["total_io"] / 1e6,
        ], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.fill = key_fill
            if c in (5, 6, 7, 8):
                cell.number_format = "#,##0.000"

        # I: Compute(us) — key ops always use TOPS
        cell_i = ws2.cell(row=r, column=9)
        cell_i.value = _compute_formula(f"E{r}", TOPS_REF)
        cell_i.number_format = "#,##0.000"
        cell_i.border = thin
        cell_i.fill = key_fill
        # J: Memory(us)
        cell_j = ws2.cell(row=r, column=10)
        cell_j.value = _memory_formula(f"H{r}")
        cell_j.number_format = "#,##0.000"
        cell_j.border = thin
        cell_j.fill = key_fill
        # K: Est.Time
        cell_k = ws2.cell(row=r, column=11)
        cell_k.value = _est_time_formula(f"I{r}", f"J{r}")
        cell_k.number_format = "#,##0.000"
        cell_k.border = thin
        cell_k.fill = key_fill
        # L: Bottleneck
        cell_l = ws2.cell(row=r, column=12)
        cell_l.value = _bottleneck_formula(f"I{r}", f"J{r}")
        cell_l.border = thin
        cell_l.fill = key_fill
        # M: Arith.Intensity
        cell_m = ws2.cell(row=r, column=13)
        cell_m.value = _ai_formula(f"E{r}", f"H{r}")
        cell_m.number_format = "#,##0.00"
        cell_m.border = thin
        cell_m.fill = key_fill
        # N: % Total GOPs  (reference Sheet1 total)
        cell_n = ws2.cell(row=r, column=14)
        cell_n.value = f"=IF('All Operators'!E{sr}=0,0,E{r}/'All Operators'!E{sr}*100)"
        cell_n.number_format = "0.0\"%\""
        cell_n.border = thin
        cell_n.fill = key_fill
        # O: % Total Time  (reference Sheet1 total)
        cell_o = ws2.cell(row=r, column=15)
        cell_o.value = f"=IF('All Operators'!K{sr}=0,0,K{r}/'All Operators'!K{sr}*100)"
        cell_o.number_format = "0.0\"%\""
        cell_o.border = thin
        cell_o.fill = key_fill
        # P: Name
        cell_name = ws2.cell(row=r, column=16, value=d["loc"])
        cell_name.border = thin
        cell_name.fill = key_fill

    # Key ops summary rows
    k_first = 2
    k_last = len(key_rows) + 1
    sr2 = len(key_rows) + 3
    for row_num, label in [(sr2, "Key Ops Total"), (sr2 + 1, "Overall Total")]:
        for c in range(1, len(h2) + 1):
            cell = ws2.cell(row=row_num, column=c)
            cell.fill = sum_fill
            cell.font = bold
            cell.border = thin
        ws2.cell(row=row_num, column=1, value=label)

    # Key Ops Total row
    ws2.cell(row=sr2, column=5, value=f"=SUM(E{k_first}:E{k_last})").number_format = "#,##0.000"
    ws2.cell(row=sr2, column=11, value=f"=SUM(K{k_first}:K{k_last})").number_format = "#,##0.000"
    ws2.cell(row=sr2, column=14).value = \
        f"=IF('All Operators'!E{sr}=0,0,E{sr2}/'All Operators'!E{sr}*100)"
    ws2.cell(row=sr2, column=14).number_format = "0.0\"%\""
    ws2.cell(row=sr2, column=15).value = \
        f"=IF('All Operators'!K{sr}=0,0,K{sr2}/'All Operators'!K{sr}*100)"
    ws2.cell(row=sr2, column=15).number_format = "0.0\"%\""
    # Overall Total row
    ws2.cell(row=sr2 + 1, column=5, value=f"='All Operators'!E{sr}").number_format = "#,##0.000"
    ws2.cell(row=sr2 + 1, column=11, value=f"='All Operators'!K{sr}").number_format = "#,##0.000"
    ws2.cell(row=sr2 + 1, column=14, value="100.0%")
    ws2.cell(row=sr2 + 1, column=15, value="100.0%")

    ws2.freeze_panes = "A2"
    # No. OpType InShapes OutShapes GOPs Read Write TotalIO Comp Mem Est Bottleneck AI %GOPs %Time Name
    _set_col_widths(ws2, [5, 18, 22, 16, 10, 10, 10, 12, 12, 12, 12, 11, 10, 10, 10, 20])

    # Move Overview sheet to the front
    wb.move_sheet(ws0, offset=-1)

    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------


def print_summary(ops,
                  chip_tops,
                  bw_gbps,
                  ssa_op_map=None,
                  vector_tops=None,
                  uarch_rate=0.5,
                  bw_util=0.7):
    if vector_tops is None:
        vector_tops = chip_tops / 8.0
    chip_flops = chip_tops * uarch_rate * 1e12
    vector_flops = vector_tops * uarch_rate * 1e12
    bw_bytes = bw_gbps * bw_util * 1e9
    compute_ops = [op for op in ops if op.op_type not in SKIP_OPS]

    print("\n" + "=" * 110)
    print(
        f"  Chip: {chip_tops} TOPS | Vector: {vector_tops} TOPS | BW: {bw_gbps} GB/s | CU: {uarch_rate:.0%} | BU: {bw_util:.0%} | State: {REQUIRED_STATE}"
    )
    print("=" * 110)

    total_flops = 0
    total_time = 0.0

    print(
        f"\n{'No.':<5} {'Op Type':<25} {'FLOPs':>12} {'Data I/O':>14} {'Comp.(us)':>12} {'Mem.(us)':>12} {'Est.(us)':>12} {'Bound':>8}"
    )
    print("-" * 100)

    for idx, op in enumerate(compute_ops, 1):
        opn = op.op_type.split(".")[-1]
        flops = calc_flops(op)
        rb, wb = calc_data_volume(op, ssa_op_map)
        tio = rb + wb
        use_flops = chip_flops if opn in KEY_OPS else vector_flops
        ct = flops / use_flops * 1e6 if use_flops else 0
        mt = tio / bw_bytes * 1e6 if bw_bytes else 0
        et = max(ct, mt)
        bn = "Compute" if ct >= mt else "Memory"
        total_flops += flops
        total_time += et

        marker = " ***" if opn in KEY_OPS else ""
        print(
            f"{idx:<5} {opn:<25} {fmt_num(flops):>12} {fmt_bytes(tio):>14} {ct:>12.3f} {mt:>12.3f} {et:>12.3f} {bn:>8}{marker}"
        )

    print("-" * 100)
    print(
        f"{'':5} {'TOTAL':<25} {fmt_num(total_flops):>12} {'':>14} {'':>12} {'':>12} {total_time:>12.3f}"
    )
    print(f"\n  Total estimated time: {fmt_time(total_time)}")
    print(f"  (*** = key operators: MatMul / Conv / FAttention / ChunkGatedDeltaRule)\n")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Analyze TPU-lowered MLIR operators: FLOPs, data volume, "
        "estimated runtime (Roofline model). Only supports "
        "module.state = \"TPU_LOWERED\".")
    parser.add_argument("mlir_file", help="Path to a TPU-lowered MLIR file")
    parser.add_argument(
        "-t",
        "--tops",
        type=float,
        required=True,
        help="Chip compute power in TOPS (e.g. 32 for BM1684X INT8, or use FP TOPS)")
    parser.add_argument("-b",
                        "--bandwidth",
                        type=float,
                        required=True,
                        help="Chip memory bandwidth in GB/s (e.g. 64)")
    parser.add_argument("-v",
                        "--vector_tops",
                        type=float,
                        default=None,
                        help="Vector compute power in TOPS (default: tops/8)")
    parser.add_argument("-r",
                        "--uarch_rate",
                        type=float,
                        default=0.7,
                        help="uArch Rate (default: 0.7)")
    parser.add_argument("-u",
                        "--bw_util",
                        type=float,
                        default=0.7,
                        help="Bandwidth utilization ratio (default: 0.7)")
    parser.add_argument("-o",
                        "--output",
                        default=None,
                        help="Output Excel path (default: <mlir_name>_analysis.xlsx)")
    args = parser.parse_args()

    if args.output is None:
        base = os.path.splitext(os.path.basename(args.mlir_file))[0]
        args.output = f"{base}_analysis.xlsx"

    print(f"Parsing: {args.mlir_file}")
    try:
        state = require_tpu_lowered(args.mlir_file)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    print(f"Module state: {state}")
    ops, _, ssa_op_map = parse_mlir_file(args.mlir_file)
    print(
        f"Found {len(ops)} operations ({len([o for o in ops if o.op_type not in SKIP_OPS])} compute ops)"
    )

    print_summary(ops, args.tops, args.bandwidth, ssa_op_map, args.vector_tops, args.uarch_rate,
                  args.bw_util)
    export_excel(ops, args.tops, args.bandwidth, args.output, ssa_op_map, args.vector_tops,
                 args.uarch_rate, args.bw_util)


if __name__ == "__main__":
    main()
