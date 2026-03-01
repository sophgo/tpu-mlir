#!/usr/bin/env python3
"""
Parse MLIR files and export operator structure to Excel,
annotating each operator with FLOPs and data volume.

Usage:
    python mlir_to_excel.py <input.mlir> [-o output.xlsx]
"""

import re
import sys
import argparse
from pathlib import Path
from functools import reduce
from operator import mul

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── dtype sizes ───────────────────────────────────────────────────────────────
DTYPE_BYTES = {
    "f64": 8,
    "f32": 4,
    "f16": 2,
    "bf16": 2,
    "si64": 8,
    "si32": 4,
    "si16": 2,
    "si8": 1,
    "ui64": 8,
    "ui32": 4,
    "ui16": 2,
    "ui8": 1,
    "i64": 8,
    "i32": 4,
    "i16": 2,
    "i8": 1,
}


def prod(dims):
    """Product of a list of ints."""
    return reduce(mul, dims, 1)


def parse_tensor_type(type_str):
    """
    Parse 'tensor<1x1024x1024xbf16>' -> ([1,1024,1024], 'bf16', 2).
    Returns (shape, dtype, bytes_per_elem).  For 'none' returns None.
    """
    type_str = type_str.strip()
    if type_str == "none":
        return None
    m = re.match(r"tensor<([\dx]+)x(\w+)>", type_str)
    if not m:
        return None
    shape = [int(d) for d in m.group(1).split("x")]
    dtype = m.group(2)
    bpe = DTYPE_BYTES.get(dtype, 0)
    return (shape, dtype, bpe)


def tensor_bytes(parsed):
    """Total bytes for one parsed tensor type."""
    if parsed is None:
        return 0
    shape, _, bpe = parsed
    return prod(shape) * bpe


def tensor_elements(parsed):
    if parsed is None:
        return 0
    return prod(parsed[0])


# ── FLOPs estimation ─────────────────────────────────────────────────────────
def estimate_flops(op_type, inputs, outputs, attrs_str):
    """
    Return estimated FLOPs for the operator.
    inputs / outputs: list of parsed tensor types (or None entries).
    """
    op = op_type.split(".")[-1]  # strip dialect prefix

    out_elems = tensor_elements(outputs[0]) if outputs else 0

    if op == "MatMul":
        # [batch..., M, K] x [batch..., K, N] -> [batch..., M, N]
        # FLOPs = 2 * batch * M * K * N
        valid_in = [i for i in inputs if i is not None]
        if len(valid_in) >= 2:
            a_shape = valid_in[0][0]  # left matrix
            b_shape = valid_in[1][0]  # right matrix
            K = a_shape[-1]
            N = b_shape[-1]
            batch_M = prod(a_shape[:-1]) if len(a_shape) > 1 else 1
            return 2 * batch_M * K * N
        return 0

    if op == "FAttention":
        # Flash Attention:  Q*K^T + softmax + attn*V
        # attrs contain batch, q_head, kv_head, mq, mk, dim
        def _attr_int(name):
            m = re.search(rf"{name}\s*=\s*(\d+)", attrs_str)
            return int(m.group(1)) if m else 0

        batch = _attr_int("batch")
        q_head = _attr_int("q_head")
        mq = _attr_int("mq")
        mk = _attr_int("mk")
        dim = _attr_int("dim")
        # Q*K^T: 2*batch*q_head*mq*mk*dim
        # softmax: ~5*batch*q_head*mq*mk
        # attn*V: 2*batch*q_head*mq*dim*mk
        flops = (2 * batch * q_head * mq * mk * dim + 5 * batch * q_head * mq * mk +
                 2 * batch * q_head * mq * dim * mk)
        return flops

    if op == "RMSNorm":
        # per element: square, sum-reduce, rsqrt, multiply by weight  ~5N
        return 5 * out_elems

    if op in ("Add", "Mul"):
        return out_elems

    if op in ("Active", "Sigmoid"):
        # Sigmoid: exp + add + div  ~4 ops
        mode = ""
        m_mode = re.search(r"active_mode\s+(\w+)", attrs_str)
        if m_mode:
            mode = m_mode.group(1)
        if mode == "SILU" or op == "SiLU":
            return 5 * out_elems  # sigmoid + mul
        return 4 * out_elems

    if op == "SiLU":
        return 5 * out_elems

    if op == "Rope":
        # sin, cos lookups + 2 muls + add per element  ~6 ops
        return 6 * out_elems

    if op == "Softmax":
        return 5 * out_elems

    # Zero-FLOPs ops: Cast, Reshape, Permute, Gather, Slice, Weight, Input, None
    return 0


# ── MLIR parser ──────────────────────────────────────────────────────────────
LOC_DEF_RE = re.compile(r'#(loc\d*)\s*=\s*loc\("([^"]+)"\)')

# Matches an operation line like:
#   %9 = "tpu.MatMul"(%7, %8, ...) {attrs} : (in_types) -> out_type loc(#locN)
OP_RE = re.compile(r'(%\d+)\s*=\s*"([^"]+)"\(([^)]*)\)\s*'  # result, op, operands
                   r'(\{[^}]*\})?\s*'  # optional attrs
                   r':\s*\(([^)]*)\)\s*->\s*'  # input types
                   r'(tensor<[^>]+>(?:\s*,\s*tensor<[^>]+>)*|none)'  # output type(s)
                   r'\s*loc\(#(loc\d*)\)'  # location
                   )

# Also match ops with multiple return values wrapped in parens:
#   (%a, %b) = "op"(...) ...
MULTI_RET_RE = re.compile(r'\(([^)]+)\)\s*=\s*"([^"]+)"\(([^)]*)\)\s*'
                          r'(\{[^}]*\})?\s*'
                          r':\s*\(([^)]*)\)\s*->\s*'
                          r'\(([^)]+)\)'
                          r'\s*loc\(#(loc\d*)\)')


def parse_type_list(s):
    """Split a comma-separated list of tensor types / none."""
    results = []
    for t in re.findall(r'tensor<[^>]+>|none', s):
        results.append(parse_tensor_type(t))
    return results


def shape_str(parsed):
    if parsed is None:
        return "none"
    return "x".join(str(d) for d in parsed[0]) + "x" + parsed[1]


def parse_mlir(filepath):
    """
    Parse an MLIR file and return:
      - module_attrs: dict of module-level attributes
      - ops: list of dicts with operator info
    """
    text = Path(filepath).read_text()

    # 1. Build location map
    loc_map = {}
    for m in LOC_DEF_RE.finditer(text):
        loc_map[m.group(1)] = m.group(2)

    # 2. Extract module attributes
    module_attrs = {}
    m_mod = re.search(r'module\s+@(\w+)\s+attributes\s*\{([^}]+)\}', text)
    if m_mod:
        module_attrs["module_name"] = m_mod.group(1)
        for kv in re.finditer(r'module\.(\w+)\s*=\s*("?[^",}]+"?)', m_mod.group(2)):
            module_attrs[kv.group(1)] = kv.group(2).strip('"')

    # 3. Parse operations
    ops = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("module") or line.startswith(
                "func.func") or line.startswith("}") or line.startswith("return"):
            continue

        result_var = None
        op_type = None
        operands_str = None
        attrs_str = ""
        in_types_str = None
        out_types_str = None
        loc_ref = None

        m = OP_RE.search(line)
        if m:
            result_var = m.group(1)
            op_type = m.group(2)
            operands_str = m.group(3)
            attrs_str = m.group(4) or ""
            in_types_str = m.group(5)
            out_types_str = m.group(6)
            loc_ref = m.group(7)
        else:
            mm = MULTI_RET_RE.search(line)
            if mm:
                result_var = mm.group(1)
                op_type = mm.group(2)
                operands_str = mm.group(3)
                attrs_str = mm.group(4) or ""
                in_types_str = mm.group(5)
                out_types_str = mm.group(6)
                loc_ref = mm.group(7)
            else:
                continue

        inputs = parse_type_list(in_types_str)
        outputs = parse_type_list(out_types_str)
        name = loc_map.get(loc_ref, loc_ref or "")

        # Skip pure infrastructure ops
        if op_type in ("top.None", ):
            continue

        flops = estimate_flops(op_type, inputs, outputs, attrs_str)

        # Data volume: sum of all non-None input + output tensor bytes
        in_bytes = sum(tensor_bytes(t) for t in inputs)
        out_bytes = sum(tensor_bytes(t) for t in outputs)
        total_bytes = in_bytes + out_bytes

        in_shapes = ", ".join(shape_str(t) for t in inputs if t is not None)
        out_shapes = ", ".join(shape_str(t) for t in outputs if t is not None)

        ops.append({
            "var": result_var,
            "op_type": op_type,
            "name": name,
            "input_shapes": in_shapes,
            "output_shapes": out_shapes,
            "input_bytes": in_bytes,
            "output_bytes": out_bytes,
            "total_bytes": total_bytes,
            "flops": flops,
            "attrs": attrs_str.strip("{}").strip(),
        })

    return module_attrs, ops


# ── human-readable size ──────────────────────────────────────────────────────
def human_size(nbytes):
    for unit in ("B", "KB", "MB", "GB"):
        if abs(nbytes) < 1024:
            return f"{nbytes:.1f} {unit}"
        nbytes /= 1024
    return f"{nbytes:.1f} TB"


def human_flops(f):
    if f == 0:
        return "0"
    for unit in ("", "K", "M", "G", "T"):
        if abs(f) < 1000:
            return f"{f:.1f} {unit}FLOPs" if unit else f"{int(f)}"
        f /= 1000
    return f"{f:.1f} PFLOPs"


# ── Excel export ─────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
COMPUTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WEIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
DATA_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_excel(module_attrs, ops, output_path):
    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Module Attribute", "Value"])
    for k, v in module_attrs.items():
        ws_summary.append([k, v])

    total_flops = sum(o["flops"] for o in ops)
    total_data = sum(o["total_bytes"] for o in ops)
    ws_summary.append([])
    ws_summary.append(["Total FLOPs", total_flops])
    ws_summary.append(["Total FLOPs (readable)", human_flops(total_flops)])
    ws_summary.append(["Total Data Volume (bytes)", total_data])
    ws_summary.append(["Total Data Volume (readable)", human_size(total_data)])
    ws_summary.append(["Number of Operators", len(ops)])

    # Style summary header
    for cell in ws_summary[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 50

    # ── Sheet 2: Operators ───────────────────────────────────────────────
    ws_ops = wb.create_sheet("Operators")
    headers = [
        "#",
        "Variable",
        "Op Type",
        "Name",
        "Input Shapes",
        "Output Shapes",
        "FLOPs",
        "FLOPs (readable)",
        "Input Bytes",
        "Output Bytes",
        "Total Data Bytes",
        "Data Volume (readable)",
        "Attributes",
    ]
    ws_ops.append(headers)

    # Style header row
    for col_idx, cell in enumerate(ws_ops[1], 1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Classify ops for coloring
    compute_ops = {
        "MatMul", "FAttention", "RMSNorm", "Active", "Sigmoid", "SiLU", "Rope", "Softmax", "Add",
        "Mul"
    }

    for idx, op in enumerate(ops, 1):
        short_op = op["op_type"].split(".")[-1]
        row = [
            idx,
            op["var"],
            op["op_type"],
            op["name"],
            op["input_shapes"],
            op["output_shapes"],
            op["flops"],
            human_flops(op["flops"]),
            op["input_bytes"],
            op["output_bytes"],
            op["total_bytes"],
            human_size(op["total_bytes"]),
            op["attrs"],
        ]
        ws_ops.append(row)
        row_idx = idx + 1  # 1-indexed, header is row 1

        # Color rows by category
        fill = None
        if short_op in ("Weight", ):
            fill = WEIGHT_FILL
        elif short_op in compute_ops:
            fill = COMPUTE_FILL
        elif short_op in ("Input", "Cast", "Reshape", "Permute", "Gather", "Slice"):
            fill = DATA_FILL

        for col in range(1, len(headers) + 1):
            cell = ws_ops.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if fill:
                cell.fill = fill

    # Column widths
    col_widths = [5, 8, 18, 50, 40, 40, 16, 16, 14, 14, 16, 16, 60]
    for i, w in enumerate(col_widths, 1):
        ws_ops.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws_ops.freeze_panes = "A2"

    # ── Sheet 3: FLOPs by Op Type ────────────────────────────────────────
    ws_stats = wb.create_sheet("Stats by Op Type")
    ws_stats.append([
        "Op Type", "Count", "Total FLOPs", "FLOPs (readable)", "Total Data Bytes",
        "Data (readable)", "FLOPs %", "Data %"
    ])
    for cell in ws_stats[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Aggregate
    from collections import defaultdict
    agg = defaultdict(lambda: {"count": 0, "flops": 0, "data": 0})
    for op in ops:
        key = op["op_type"]
        agg[key]["count"] += 1
        agg[key]["flops"] += op["flops"]
        agg[key]["data"] += op["total_bytes"]

    sorted_agg = sorted(agg.items(), key=lambda x: -x[1]["flops"])
    for op_type, vals in sorted_agg:
        flops_pct = vals["flops"] / total_flops * 100 if total_flops else 0
        data_pct = vals["data"] / total_data * 100 if total_data else 0
        ws_stats.append([
            op_type,
            vals["count"],
            vals["flops"],
            human_flops(vals["flops"]),
            vals["data"],
            human_size(vals["data"]),
            f"{flops_pct:.1f}%",
            f"{data_pct:.1f}%",
        ])
    ws_stats.column_dimensions["A"].width = 25
    ws_stats.column_dimensions["B"].width = 10
    ws_stats.column_dimensions["C"].width = 18
    ws_stats.column_dimensions["D"].width = 18
    ws_stats.column_dimensions["E"].width = 18
    ws_stats.column_dimensions["F"].width = 18
    ws_stats.column_dimensions["G"].width = 12
    ws_stats.column_dimensions["H"].width = 12

    wb.save(output_path)
    print(f"Saved to {output_path}")
    print(f"  Operators: {len(ops)}")
    print(f"  Total FLOPs: {human_flops(total_flops)} ({total_flops})")
    print(f"  Total Data: {human_size(total_data)} ({total_data} bytes)")


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Export MLIR structure to Excel")
    parser.add_argument("input", help="Path to the MLIR file")
    parser.add_argument("-o", "--output", help="Output Excel file (default: <input>.xlsx)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: {input_path} does not exist", file=sys.stderr)
        sys.exit(1)

    output_path = args.output or str(input_path.with_suffix(".xlsx"))

    module_attrs, ops = parse_mlir(input_path)
    write_excel(module_attrs, ops, output_path)


if __name__ == "__main__":
    main()
