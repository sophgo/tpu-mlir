#!/usr/bin/env python3
# ==============================================================================
#
# Copyright (C) 2025 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
"""
LLM MLIR analysis tool.

Discovers and analyzes the TPU-lowered (*_tpu.mlir) files generated for an LLM
model, producing a single Excel workbook with:
  - Overview sheet: chip parameters, per-module summary, phase totals
  - Per-module sheets: detailed per-operator analysis (Roofline model)

Only TPU-lowered modules (``module.state = "TPU_LOWERED"``) are supported; the
tool analyses the ``tpu.*`` dialect ops only (the three data-supplying
``top.*`` ops ``top.Weight`` / ``top.Input`` / ``top.None`` are tolerated as
they carry no compute or I/O).

The *_tpu.mlir files carry the real per-tensor dtypes (bf16, ui8 for packed
w4/w8 weights, f8E4M3FN, ...), so data volumes are computed from each tensor's
own storage size — no external dtype substitution is needed.

Transformer blocks (block_0..N, block_cache_0..N) are grouped; the first
block is analyzed as representative and multiplied by the block count.

Directory structure expected:
    <model_dir>/<module_name>/<module_name>_<chip>_<quant>_tpu.mlir
    e.g. .../block_0/block_0_bm1684x_w4bf16_tpu.mlir

Usage:
    python llm_analyse.py -m <model> -c bm1684x -s 2048 -t 16 -b 64 -o <out_dir>
    # -t: FP16 TOPS; INT8/FP8 TOPS default to 2*t
"""

import os
import re
import sys
import argparse
import glob
from typing import List, Dict, Tuple, Optional
import subprocess

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mlir_analyse import (
    parse_mlir_file,
    require_tpu_lowered,
    calc_flops,
    calc_data_volume,
    KEY_OPS,
    SKIP_OPS,
    fmt_num,
    fmt_bytes,
    fmt_time,
)

# ---------------------------------------------------------------------------
# Phase classification
# ---------------------------------------------------------------------------
PREFILL_MODULES = {
    "vit", "embedding", "block", "block(LinearAttention)", "block(FullAttention)", "lm_head"
}
DECODE_MODULES = {
    "embedding_cache", "block_cache", "block_cache(LinearAttention)", "block_cache(FullAttention)",
    "lm_head"
}
MODULE_ORDER = [
    "vit",
    "embedding",
    "block",
    "lm_head",
    "embedding_cache",
    "block_cache",
]

# Qwen3.5 dense and MoE share the same hybrid linear/full-attention layout
# (3:1 ratio, layer 0 = linear, layer 3 = full), so both are split the same way.
QWEN3_5_MODEL_TYPES = {"qwen3_5_text", "qwen3_5_moe_text"}

# ---------------------------------------------------------------------------
# Module discovery
# ---------------------------------------------------------------------------


def _find_tpu_mlir(sub_dir: str, entry: str, chip: str) -> Optional[str]:
    """Return the TPU-lowered mlir path for ``entry`` under ``sub_dir``.

    The lowered file is named ``{entry}_{chip}_{quant}_tpu.mlir`` (e.g.
    ``block_0_bm1684x_w4bf16_tpu.mlir``). The ``{quant}`` part differs per
    module (blocks use the LLM quant like ``w4bf16``; vit/embedding/lm_head are
    usually ``bf16``/``f16``), so we glob on ``{entry}_{chip}_*_tpu.mlir``.
    Returns ``None`` if not found.
    """
    pattern = f"{entry}_{chip}_*_tpu.mlir"
    matches = glob.glob(os.path.join(sub_dir, pattern))
    if len(matches) == 0:
        return None
    path = matches[0]
    new_path = path.replace("_tpu.mlir", "_tpu_fix.mlir")
    subprocess.run(
        ["tpuc-opt", f"--strip-io-quant=quant_input=True quant_output=True", path, "-o", new_path],
        check=True)
    return new_path


def discover_modules(model_dir: str,
                     num_layers: int,
                     model_type: str = "",
                     chip: str = "bm1684x") -> List[Tuple[str, str, int]]:
    """Discover and group TPU-lowered MLIR files under model_dir.

    Returns list of (module_name, mlir_path, count) in logical order.
    Transformer blocks are grouped: block -> (block_0_tpu.mlir, N).
    For qwen3_5_text, block_0/block_3 are kept separate (3:1 ratio).
    """
    block_files: Dict[int, str] = {}
    cache_files: Dict[int, str] = {}
    other_files: Dict[str, str] = {}

    for entry in sorted(os.listdir(model_dir)):
        sub_dir = os.path.join(model_dir, entry)
        if not os.path.isdir(sub_dir):
            continue
        mlir_path = _find_tpu_mlir(sub_dir, entry, chip)
        if not mlir_path:
            continue
        m_block = re.match(r"^block_(\d+)$", entry)
        m_cache = re.match(r"^block_cache_(\d+)$", entry)
        if m_block:
            block_files[int(m_block.group(1))] = mlir_path
        elif m_cache:
            cache_files[int(m_cache.group(1))] = mlir_path
        else:
            other_files[entry] = mlir_path

    is_qwen3_5 = model_type in QWEN3_5_MODEL_TYPES
    modules = []
    for name in MODULE_ORDER:
        if name == "block":
            if not is_qwen3_5:
                modules.append(("block", block_files[0], num_layers))
            else:
                modules.append(("block(LinearAttention)", block_files[0], num_layers * 3 // 4))
                modules.append(("block(FullAttention)", block_files[3], num_layers // 4))
        elif name == "block_cache":
            if not is_qwen3_5:
                modules.append(("block_cache", cache_files[0], num_layers))
            else:
                modules.append(
                    ("block_cache(LinearAttention)", cache_files[0], num_layers * 3 // 4))
                modules.append(("block_cache(FullAttention)", cache_files[3], num_layers // 4))
        elif name in other_files:
            modules.append((name, other_files.pop(name), 1))

    for name in sorted(other_files):
        modules.append((name, other_files[name], 1))

    return modules


# ---------------------------------------------------------------------------
# Single-module analysis
# ---------------------------------------------------------------------------

COMPUTE_FP16 = "fp16"
COMPUTE_INT8 = "int8"
COMPUTE_FP8 = "fp8"
COMPUTE_VECTOR = "vector"


def _compute_type(base_opn: str) -> str:
    """Pick compute-power bucket for roofline estimate.

    Fp8MatMul computes in FP8; weight-only quant (w8/w4 f16/bf16) still
    computes in FP16, so all other key ops use FP16 TOPS.
    """
    if base_opn == "Fp8MatMul":
        return COMPUTE_FP8
    if base_opn in KEY_OPS:
        return COMPUTE_FP16
    return COMPUTE_VECTOR


def analyse_module(filepath: str):
    """Parse and analyse a single TPU-lowered MLIR file.

    The module must be in the ``TPU_LOWERED`` state (validated up front) and
    only ``tpu.*`` ops are analysed. The tpu.mlir carries the real per-tensor
    dtypes, so data volumes use each tensor's own storage size.

    Returns (rows_data, totals) where totals has keys: flops, read, write, io.
    """
    require_tpu_lowered(filepath)
    ops, _, ssa_op_map = parse_mlir_file(filepath)
    compute_ops = [op for op in ops if op.op_type not in SKIP_OPS]

    rows_data = []
    total_flops = total_read = total_write = 0

    for op in compute_ops:
        base_opn = op.op_type.split(".")[-1]
        flops = calc_flops(op)
        rb, wb = calc_data_volume(op, ssa_op_map=ssa_op_map)
        inp_shapes = ", ".join(t.shape_str() if t else "none" for t in op.input_types)
        out_shapes = ", ".join(t.shape_str() if t else "none" for t in op.output_types)
        rows_data.append(
            dict(
                opn=base_opn,
                loc=op.loc_name,
                inp_shapes=inp_shapes,
                out_shapes=out_shapes,
                flops=flops,
                rb=rb,
                wb=wb,
                total_io=rb + wb,
                is_key=base_opn in KEY_OPS,
                compute_type=_compute_type(base_opn),
            ))
        total_flops += flops
        total_read += rb
        total_write += wb

    totals = dict(flops=total_flops,
                  read=total_read,
                  write=total_write,
                  io=total_read + total_write)
    return rows_data, totals


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


def export_llm_excel(modules_data,
                     fp16_tops,
                     bw_gbps,
                     out_dir,
                     llm_path,
                     int8_tops=None,
                     fp8_tops=None,
                     vector_tops=None,
                     uarch_rate=0.8,
                     bw_util=0.7,
                     parallelism=0.5,
                     model_config=None,
                     seq_length=0,
                     max_pixels="",
                     cmdline=""):
    """Create combined LLM analysis Excel workbook.

    modules_data: list of (name, count, rows_data, totals)
    model_config: dict of LLM architecture info from AutoConfig
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter

    if int8_tops is None:
        int8_tops = fp16_tops * 2.0
    if fp8_tops is None:
        fp8_tops = fp16_tops * 2.0
    if vector_tops is None:
        vector_tops = fp16_tops / 8.0

    # ---------- Color palette ----------
    C_TITLE = "1F4E79"  # deep blue - main title
    C_SECTION = "2E75B6"  # medium blue - section banners
    C_COLHDR = "4472C4"  # column headers
    C_RESULT = "1F4E79"  # result highlight
    C_EDIT = "E2EFDA"  # editable (soft green)
    C_EDIT_TXT = "375623"
    C_KEY = "FFF2CC"  # key operator (soft yellow)
    C_SUM = "DDEBF7"  # summary row (soft blue)
    C_PHASE = "FCE4D6"  # phase total (soft peach)
    C_PHASE_TXT = "833C0B"
    C_ZEBRA = "F7F9FC"  # alternating row
    C_UTIL = "D9E7F5"  # utilization value background

    # ---------- Styles ----------
    wb = Workbook()
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill(start_color=C_TITLE, end_color=C_TITLE, fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color=C_SECTION, end_color=C_SECTION, fill_type="solid")
    col_hdr_font = Font(bold=True, color="FFFFFF", size=10)
    col_hdr_fill = PatternFill(start_color=C_COLHDR, end_color=C_COLHDR, fill_type="solid")
    key_fill = PatternFill(start_color=C_KEY, end_color=C_KEY, fill_type="solid")
    zebra_fill = PatternFill(start_color=C_ZEBRA, end_color=C_ZEBRA, fill_type="solid")
    sum_fill = PatternFill(start_color=C_SUM, end_color=C_SUM, fill_type="solid")
    edit_fill = PatternFill(start_color=C_EDIT, end_color=C_EDIT, fill_type="solid")
    edit_font = Font(bold=True, color=C_EDIT_TXT)
    phase_fill = PatternFill(start_color=C_PHASE, end_color=C_PHASE, fill_type="solid")
    phase_font = Font(bold=True, color=C_PHASE_TXT)
    result_fill = PatternFill(start_color=C_RESULT, end_color=C_RESULT, fill_type="solid")
    result_font = Font(bold=True, color="FFFFFF", size=12)
    util_fill = PatternFill(start_color=C_UTIL, end_color=C_UTIL, fill_type="solid")
    util_font = Font(bold=True, color=C_TITLE, size=10)
    bold = Font(bold=True)
    thin = Border(*(Side(style="thin", color="BFBFBF"), ) * 4)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _set_col_widths(ws, widths):
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    def _banner(ws, row, text, span, fill=section_fill, font=section_font, height=22):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = height

    # ============ Overview sheet ============
    ws0 = wb.active
    ws0.title = "Overview"
    TOTAL_COLS = 7

    # Row 1: Title banner
    _banner(ws0,
            1,
            "  LLM Performance Analysis",
            TOTAL_COLS,
            fill=title_fill,
            font=title_font,
            height=30)

    # --- Overview sheet layout (depends on module count, so precompute rows) ---
    # Order: Performance Summary -> Module Breakdown -> Hardware & Utilization
    # -> Special Op Ratios. Hardware & Utilization is pushed below the module
    # table, whose height depends on len(modules_data), so its absolute row
    # numbers (which feed cross-sheet formulas) are computed up front.
    NUM_MODULES = len(modules_data)
    PERF_BANNER = 3
    MOD_BANNER = PERF_BANNER + 3 + 1  # blank spacer row 6 -> banner at 7
    sum_hdr = MOD_BANNER + 1
    mod_start = sum_hdr + 1
    # 2 phase-summary rows follow the per-module rows
    phase_r = mod_start + NUM_MODULES + 1
    HW_BANNER = phase_r + 3  # blank spacer after phase rows
    HW_HDR = HW_BANNER + 1

    # --- Performance Summary (rows 3-5) ---
    _banner(ws0, PERF_BANNER, "Performance Summary", TOTAL_COLS)

    # We fill rows 4-5 with TTFT/Tokens/s; values reference phase totals below.
    # Placeholder cells styled now, formulas injected after we know phase_r.
    for r in (PERF_BANNER + 1, PERF_BANNER + 2):
        for c in range(1, TOTAL_COLS + 1):
            ws0.cell(row=r, column=c).border = thin
        ws0.row_dimensions[r].height = 22
    ws0.cell(row=PERF_BANNER + 1, column=1, value="TTFT (s)").font = result_font
    ws0.cell(row=PERF_BANNER + 1, column=1).fill = result_fill
    ws0.cell(row=PERF_BANNER + 1, column=1).alignment = center
    ws0.cell(row=PERF_BANNER + 2, column=1, value="Tokens/s").font = result_font
    ws0.cell(row=PERF_BANNER + 2, column=1).fill = result_fill
    ws0.cell(row=PERF_BANNER + 2, column=1).alignment = center
    for r in (PERF_BANNER + 1, PERF_BANNER + 2):
        vcell = ws0.cell(row=r, column=2)
        vcell.font = result_font
        vcell.fill = result_fill
        vcell.alignment = center
        # Utilization labels
        for col, label in ((3, "MFU Util"), (5, "BW Util")):
            lc = ws0.cell(row=r, column=col, value=label)
            lc.font = util_font
            lc.fill = util_fill
            lc.alignment = center
        for col in (4, 6):
            vc = ws0.cell(row=r, column=col)
            vc.font = util_font
            vc.fill = util_fill
            vc.alignment = center
            vc.number_format = "0.00%"

    # --- Module Breakdown ---
    _banner(ws0, MOD_BANNER, "Module Breakdown", TOTAL_COLS)
    sum_headers = [
        "Module", "Count", "GOPs", "I/O (MB)", "Est. Time (s)", "Total Time (s)", "Phase"
    ]
    for c, h in enumerate(sum_headers, 1):
        cell = ws0.cell(row=sum_hdr, column=c, value=h)
        cell.font = col_hdr_font
        cell.fill = col_hdr_fill
        cell.alignment = center
        cell.border = thin

    # --- Hardware & Utilization (now below the module table) ---
    _banner(ws0, HW_BANNER, "Hardware & Utilization", TOTAL_COLS)
    for c, h in enumerate(["Parameter", "Value", "Unit"], 1):
        cell = ws0.cell(row=HW_HDR, column=c, value=h)
        cell.font = col_hdr_font
        cell.fill = col_hdr_fill
        cell.alignment = center
        cell.border = thin

    params = [
        ("FP16 Compute Power", fp16_tops, "TOPS", "#,##0.##"),
        ("INT8 Compute Power", int8_tops, "TOPS", "#,##0.##"),
        ("FP8 Compute Power", fp8_tops, "TOPS", "#,##0.##"),
        ("Vector Compute Power", vector_tops, "TOPS", "#,##0.##"),
        ("Chip Bandwidth", bw_gbps, "GB/s", "#,##0.##"),
        ("uArch Rate", uarch_rate, "", "0%"),
        ("Bandwidth Utilization", bw_util, "", "0%"),
        ("Parallelism", parallelism, "", "0%"),
        ("Serialism", None, "", "0%"),  # filled below as =1-B<parallelism>
        ("CPU Call", 100, "us", "#,##0"),
        ("Preprocess Time", 0.1, "s", "#,##0.000"),
    ]
    PARAM_START = HW_HDR + 1
    for i, (label, val, unit, fmt) in enumerate(params):
        r = PARAM_START + i
        ws0.cell(row=r, column=1, value=label).border = thin
        if label == "Serialism":
            # Parallelism sits 1 row above Serialism; reference that cell.
            val = f"=1-B{r - 1}"
        c = ws0.cell(row=r, column=2, value=val)
        c.border = thin
        c.number_format = fmt
        c.alignment = center
        # All except computed Serialism are editable
        if label != "Serialism":
            c.fill = edit_fill
            c.font = edit_font
        else:
            c.font = bold
        uc = ws0.cell(row=r, column=3, value=unit)
        uc.border = thin
        uc.alignment = center

    # Formula references (must match absolute positions above)
    def _param_row(label):
        return PARAM_START + [p[0] for p in params].index(label)

    FP16_TOPS_REF = f"Overview!$B${_param_row('FP16 Compute Power')}"
    INT8_TOPS_REF = f"Overview!$B${_param_row('INT8 Compute Power')}"
    FP8_TOPS_REF = f"Overview!$B${_param_row('FP8 Compute Power')}"
    VECTOR_TOPS_REF = f"Overview!$B${_param_row('Vector Compute Power')}"
    BW_REF = f"Overview!$B${_param_row('Chip Bandwidth')}"
    CU_REF = f"Overview!$B${_param_row('uArch Rate')}"
    BU_REF = f"Overview!$B${_param_row('Bandwidth Utilization')}"
    # Parallelism feeds Serialism (= 1 - Parallelism); the Est.Time overlap
    # model uses Serialism as the non-overlapping fraction.
    SER_REF = f"Overview!$B${_param_row('Serialism')}"
    CPU_CALL_REF = f"Overview!$B${_param_row('CPU Call')}"
    PREPROCESS_TIME_REF = f"Overview!$B${_param_row('Preprocess Time')}"

    def _tops_ref_for_op(d):
        ctype = d.get("compute_type", COMPUTE_VECTOR)
        if ctype == COMPUTE_INT8:
            return INT8_TOPS_REF
        if ctype == COMPUTE_FP8:
            return FP8_TOPS_REF
        if ctype == COMPUTE_FP16:
            return FP16_TOPS_REF
        return VECTOR_TOPS_REF

    # --- Special Op Ratios ---
    RATIO_BANNER = PARAM_START + len(params) + 1
    _banner(ws0, RATIO_BANNER, "Special Op Ratios", TOTAL_COLS)
    for c, h in enumerate(["Operation", "Ratio"], 1):
        cell = ws0.cell(row=RATIO_BANNER + 1, column=c, value=h)
        cell.font = col_hdr_font
        cell.fill = col_hdr_fill
        cell.alignment = center
        cell.border = thin

    special_ratios = [
        ("FAttention (Prefill)", 3.0),
        ("FAttention (Decode)", 2.5),
        ("Gather", 5.0),
        ("Permute/Concat", 2.0),
        ("Mlp", 1.5),
        ("ChunkGatedDeltaRule", 3.0),
        ("RecurrentGatedDeltaRule", 2.0),
    ]
    RATIO_START = RATIO_BANNER + 2
    for i, (label, val) in enumerate(special_ratios):
        r = RATIO_START + i
        ws0.cell(row=r, column=1, value=label).border = thin
        c = ws0.cell(row=r, column=2, value=val)
        c.fill = edit_fill
        c.font = edit_font
        c.number_format = "0%"
        c.border = thin
        c.alignment = center
    FATTENTION_RATIO_REF = f"Overview!$B${RATIO_START}"
    FATTENTION_DECODE_RATIO_REF = f"Overview!$B${RATIO_START+1}"
    GATHER_RATIO_REF = f"Overview!$B${RATIO_START+2}"
    PERMUTE_CONCAT_RATIO_REF = f"Overview!$B${RATIO_START+3}"
    MLP_RATIO_REF = f"Overview!$B${RATIO_START+4}"
    CHUNKGATEDDELTARULE_RATIO_REF = f"Overview!$B${RATIO_START+5}"
    RECURRENTGATEDDELTARULE_RATIO_REF = f"Overview!$B${RATIO_START+6}"

    def _compute_formula(gops_cell, tops_ref=FP16_TOPS_REF):
        return f"=IF({tops_ref}=0,0,{gops_cell}/({tops_ref}*{CU_REF})*1000)"

    def _memory_formula(io_cell):
        return f"=IF({BW_REF}=0,0,{io_cell}/({BW_REF}*{BU_REF})*1000)"

    # Per-module sheet columns (no "No." column - Excel row numbers suffice)
    # 1:Op Type 2:Input Shapes 3:Output Shapes 4:GOPs 5:Read 6:Write 7:I/O
    # 8:Compute(us) 9:Memory(us) 10:Est.Time 11:Bottleneck 12:Name
    op_headers = [
        "Op Type",
        "Input Shapes",
        "Output Shapes",
        "GOPs",
        "Read (MB)",
        "Write (MB)",
        "I/O (MB)",
        "Compute (us)",
        "Memory (us)",
        "Est. Time (us)",
        "Bottleneck",
        "Name",
    ]
    mod_start = sum_hdr + 1
    prefill_rows = []
    decode_rows = []
    cur_ov_row = mod_start
    has_vit = False

    for mi, (mod_name, count, rows_data, totals) in enumerate(modules_data):
        # --- Create module sheet ---
        ws = wb.create_sheet(mod_name)
        for c, h in enumerate(op_headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = col_hdr_font
            cell.fill = col_hdr_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[1].height = 28
        is_decode_mod = mod_name in DECODE_MODULES and mod_name not in PREFILL_MODULES

        for idx, d in enumerate(rows_data, 1):
            r = idx + 1
            use_zebra = (idx % 2 == 0) and not d["is_key"]
            row_fill = key_fill if d["is_key"] else (zebra_fill if use_zebra else None)
            values = [
                d["opn"],
                d["inp_shapes"],
                d["out_shapes"],
                d["flops"] / 1e9,
                d["rb"] / 1024.0 / 1024.0,
                d["wb"] / 1024.0 / 1024.0,
                d["total_io"] / 1024.0 / 1024.0,
            ]
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = thin
                if row_fill is not None:
                    cell.fill = row_fill
                if c == 4:
                    cell.number_format = "#,##0.000000"
                elif c in (5, 6, 7):
                    cell.number_format = "#,##0.000"

            tops_ref = _tops_ref_for_op(d)
            # H: Compute(us) - from GOPs column D
            cell_h = ws.cell(row=r, column=8)
            cell_h.value = _compute_formula(f"D{r}", tops_ref)
            cell_h.number_format = "#,##0.000"
            cell_h.border = thin
            # I: Memory(us) - from I/O column G
            cell_i = ws.cell(row=r, column=9)
            cell_i.value = _memory_formula(f"G{r}")
            cell_i.number_format = "#,##0.000"
            cell_i.border = thin
            # J: Est.Time = MAX(H,I)+SER*MIN(H,I), with special OP ratio
            cell_j = ws.cell(row=r, column=10)
            base_time = f"MAX(H{r},I{r})+{SER_REF}*MIN(H{r},I{r})"
            if d["opn"] == "FAttention":
                fa_ref = FATTENTION_DECODE_RATIO_REF if is_decode_mod else FATTENTION_RATIO_REF
                cell_j.value = f"=({base_time})*{fa_ref}"
            elif d["opn"] == "Gather":
                cell_j.value = f"=({base_time})*{GATHER_RATIO_REF}"
            elif d["opn"] in ("Permute", "Concat"):
                cell_j.value = f"=({base_time})*{PERMUTE_CONCAT_RATIO_REF}"
            elif d["opn"] == "Mlp":
                cell_j.value = f"=({base_time})*{MLP_RATIO_REF}"
            elif d["opn"] == "ChunkGatedDeltaRule":
                cell_j.value = f"=({base_time})*{CHUNKGATEDDELTARULE_RATIO_REF}"
            elif d["opn"] == "RecurrentGatedDeltaRule":
                cell_j.value = f"=({base_time})*{RECURRENTGATEDDELTARULE_RATIO_REF}"
            else:
                cell_j.value = f"={base_time}"
            cell_j.number_format = "#,##0.000"
            cell_j.border = thin
            # K: Bottleneck
            cell_k = ws.cell(row=r, column=11, value=f'=IF(H{r}>=I{r},"Compute","Memory")')
            cell_k.border = thin
            cell_k.alignment = center
            # L: Name
            ws.cell(row=r, column=12, value=d["loc"]).border = thin

            if row_fill is not None:
                for c in range(8, 13):
                    ws.cell(row=r, column=c).fill = row_fill

        # Conditional formatting on Bottleneck column (K)
        nrows = len(rows_data)
        if nrows > 0:
            rng = f"K2:K{nrows + 1}"
            compute_rule = CellIsRule(operator="equal",
                                      formula=['"Compute"'],
                                      font=Font(bold=True, color="1F4E79"),
                                      fill=PatternFill(start_color="BDD7EE",
                                                       end_color="BDD7EE",
                                                       fill_type="solid"))
            memory_rule = CellIsRule(operator="equal",
                                     formula=['"Memory"'],
                                     font=Font(bold=True, color="9C0006"),
                                     fill=PatternFill(start_color="FFC7CE",
                                                      end_color="FFC7CE",
                                                      fill_type="solid"))
            ws.conditional_formatting.add(rng, compute_rule)
            ws.conditional_formatting.add(rng, memory_rule)

        # Summary row
        sr = len(rows_data) + 3
        for c in range(1, len(op_headers) + 1):
            cell = ws.cell(row=sr, column=c)
            cell.fill = sum_fill
            cell.font = bold
            cell.border = thin
        ws.cell(row=sr, column=1, value="TOTAL").alignment = center
        fd, ld = 2, len(rows_data) + 1
        ws.cell(row=sr, column=4, value=f"=SUM(D{fd}:D{ld})").number_format = "#,##0.000000"
        ws.cell(row=sr, column=5, value=f"=SUM(E{fd}:E{ld})").number_format = "#,##0.000"
        ws.cell(row=sr, column=6, value=f"=SUM(F{fd}:F{ld})").number_format = "#,##0.000"
        ws.cell(row=sr, column=7, value=f"=SUM(G{fd}:G{ld})").number_format = "#,##0.000"
        ws.cell(row=sr, column=10, value=f"=SUM(J{fd}:J{ld})").number_format = "#,##0.000"
        ws.freeze_panes = "A2"
        _set_col_widths(ws, [20, 24, 18, 11, 11, 11, 11, 12, 12, 12, 12, 24])

        # --- Fill Overview module row ---
        ov_r = cur_ov_row
        cur_ov_row += 1
        ws0.cell(row=ov_r, column=1, value=mod_name).border = thin
        # Count (editable for blocks)
        c_count = ws0.cell(row=ov_r, column=2, value=count)
        c_count.border = thin
        c_count.number_format = "0"
        c_count.alignment = center
        if count > 1:
            c_count.fill = edit_fill
            c_count.font = edit_font
        # C: GOPs
        c_gops = ws0.cell(row=ov_r, column=3)
        c_gops.value = f"='{mod_name}'!D{sr}"
        c_gops.number_format = "#,##0.000000"
        c_gops.border = thin
        # D: I/O (MB)
        c_io = ws0.cell(row=ov_r, column=4)
        c_io.value = f"='{mod_name}'!G{sr}"
        c_io.number_format = "#,##0.000"
        c_io.border = thin
        # E: Est. Time (s)
        c_est = ws0.cell(row=ov_r, column=5)
        c_est.value = f"='{mod_name}'!J{sr}/1000000"
        c_est.number_format = "#,##0.000000"
        c_est.border = thin
        # F: Total Time (s)
        c_total = ws0.cell(row=ov_r, column=6)
        c_total.value = f"=E{ov_r}*B{ov_r}+B{ov_r}*{CPU_CALL_REF}/1000000"
        c_total.number_format = "#,##0.000000"
        c_total.border = thin
        # G: Phase
        if mod_name in PREFILL_MODULES and mod_name in DECODE_MODULES:
            phase = "Both"
        elif mod_name in PREFILL_MODULES:
            phase = "Prefill"
        elif mod_name in DECODE_MODULES:
            phase = "Decode"
        else:
            phase = ""
        c_phase = ws0.cell(row=ov_r, column=7, value=phase)
        c_phase.border = thin
        c_phase.alignment = center

        if mod_name in PREFILL_MODULES:
            prefill_rows.append(ov_r)
        if mod_name in DECODE_MODULES:
            decode_rows.append(ov_r)
        if mod_name == "vit":
            has_vit = True

    # ============ Phase summary rows ============
    phase_r = cur_ov_row + 1
    for pr, (label, rows) in enumerate([
        ("Prefill Total", prefill_rows),
        ("Decode Total", decode_rows),
    ]):
        r = phase_r + pr
        for c in range(1, len(sum_headers) + 1):
            cell = ws0.cell(row=r, column=c)
            cell.fill = phase_fill
            cell.font = phase_font
            cell.border = thin
            cell.alignment = center
        ws0.cell(row=r, column=1, value=label).alignment = left
        if rows:
            total_time_formula = "+".join(f"F{x}" for x in rows)
            if label == "Prefill Total" and has_vit:
                total_time_formula += f"+{PREPROCESS_TIME_REF}"
            ws0.cell(row=r, column=3,
                     value="=" + "+".join(f"C{x}*B{x}"
                                          for x in rows)).number_format = "#,##0.000000"
            ws0.cell(row=r, column=4,
                     value="=" + "+".join(f"D{x}*B{x}" for x in rows)).number_format = "#,##0.000"
            ws0.cell(row=r, column=6, value="=" + total_time_formula).number_format = "#,##0.000000"

    # ============ Back-fill TTFT / Tokens/s at top (rows 4, 5) ============
    # TTFT row (4)
    ws0.cell(row=4, column=2, value=f"=F{phase_r}").number_format = "#,##0.000000"
    ws0.cell(row=4,
             column=4,
             value=f"=IF(F{phase_r}=0,0,C{phase_r}/{FP16_TOPS_REF}/1000/F{phase_r})")
    ws0.cell(row=4, column=6, value=f"=IF(F{phase_r}=0,0,D{phase_r}/{BW_REF}/1000/F{phase_r})")
    # Tokens/s row (5)
    ws0.cell(row=5, column=2,
             value=f"=IF(F{phase_r+1}=0,0,1/F{phase_r+1})").number_format = "#,##0.00"
    ws0.cell(row=5,
             column=4,
             value=f"=IF(F{phase_r+1}=0,0,C{phase_r+1}/{FP16_TOPS_REF}/1000/F{phase_r+1})")
    ws0.cell(row=5,
             column=6,
             value=f"=IF(F{phase_r+1}=0,0,D{phase_r+1}/{BW_REF}/1000/F{phase_r+1})")
    # Re-apply number format for util cells (set before formula insertion was overwritten)
    for r in (4, 5):
        for c in (4, 6):
            ws0.cell(row=r, column=c).number_format = "0.00%"
            ws0.cell(row=r, column=c).font = util_font
            ws0.cell(row=r, column=c).fill = util_fill
            ws0.cell(row=r, column=c).alignment = center
        ws0.cell(row=r, column=2).font = result_font
        ws0.cell(row=r, column=2).fill = result_fill
        ws0.cell(row=r, column=2).alignment = center

    # ============ Model Architecture ============
    # Placed after Special Op Ratios (which now sit below the module table).
    arch_banner = RATIO_START + len(special_ratios) + 1
    if model_config:
        _banner(ws0, arch_banner, "Model Architecture", TOTAL_COLS)
        arch_fields = [
            ("Model", llm_path),
            ("Architecture", model_config.get("architectures", "")),
            ("Num Hidden Layers", model_config.get("num_hidden_layers", "")),
            ("Hidden Size", model_config.get("hidden_size", "")),
            ("Num Attention Heads", model_config.get("num_attention_heads", "")),
            ("Num Key Value Heads", model_config.get("num_key_value_heads", "")),
            ("Intermediate Size", model_config.get("intermediate_size", "")),
            ("Vocab Size", model_config.get("vocab_size", "")),
            ("Head Dim", model_config.get("head_dim", "")),
            ("Seq Length", seq_length if seq_length else ""),
            ("Max Pixels", max_pixels if max_pixels else ""),
            ("Command", cmdline if cmdline else ""),
        ]
        arch_fields = [(k, v) for k, v in arch_fields if v != ""]
        for r_off, (k, v) in enumerate(arch_fields, 1):
            kc = ws0.cell(row=arch_banner + r_off, column=1, value=k)
            kc.border = thin
            kc.font = bold
            # Span value over remaining columns for readability
            ws0.merge_cells(start_row=arch_banner + r_off,
                            start_column=2,
                            end_row=arch_banner + r_off,
                            end_column=TOTAL_COLS)
            c_val = ws0.cell(row=arch_banner + r_off, column=2, value=v)
            c_val.border = thin
            c_val.alignment = left
            if isinstance(v, (int, float)):
                c_val.number_format = "#,##0"
        info_banner = arch_banner + len(arch_fields) + 2
    else:
        info_banner = arch_banner

    # ============ Notes ============
    _banner(ws0, info_banner, "Notes", TOTAL_COLS)
    notes = [
        "Only TPU-lowered modules (module.state = \"TPU_LOWERED\") are analysed; tpu.* ops only.",
        "Green cells (parameters, ratios, block counts) are editable; all estimates auto-update.",
        "Est.Time = max(Compute, Memory) + Serialism * min(Compute, Memory).",
        "block / block_cache use the first block as representative, multiplied by Count.",
        "Data volumes read the real dtypes from the TPU-lowered *_tpu.mlir (bf16, ui8 for packed w4/w8, ...).",
        "Key operators: Fp8MatMul uses FP8 TOPS; others (incl. w4/w8 A16MatMul) use FP16 TOPS; non-key ops use vector TOPS.",
    ]
    for i, txt in enumerate(notes, 1):
        ws0.merge_cells(start_row=info_banner + i,
                        start_column=1,
                        end_row=info_banner + i,
                        end_column=TOTAL_COLS)
        cell = ws0.cell(row=info_banner + i, column=1, value=f"• {txt}")
        cell.alignment = left
        cell.font = Font(italic=True, color="595959")

    # Freeze top summary rows for navigation
    ws0.freeze_panes = "A6"
    _set_col_widths(ws0, [26, 16, 14, 14, 16, 16, 12])
    file = os.path.join(out_dir, f"{os.path.basename(out_dir)}.xlsx")
    wb.save(file)
    print(f"\nExcel saved: {file}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    # yapf: disable
    parser = argparse.ArgumentParser(
        description="Analyze LLM MLIR modules: per-operator FLOPs, data volume, "
                    "estimated runtime (Roofline model)")
    parser.add_argument('-m', '--model_path', type=str, required=True,
                        help='original weight, like ./Qwen2-7B-Instruct')
    parser.add_argument('-s', '--seq_length', type=int, required=True,
                        help="sequence length")
    parser.add_argument("-t", "--tops", type=float, required=True,
                        help="FP16 compute power in TOPS")
    parser.add_argument("--int8_tops", type=float, default=None,
                        help="INT8 compute power in TOPS (default: 2 * tops)")
    parser.add_argument("--fp8_tops", type=float, default=None,
                        help="FP8 compute power in TOPS (default: 2 * tops)")
    parser.add_argument("-b", "--bandwidth", type=float, required=True,
                        help="Chip memory bandwidth in GB/s")
    parser.add_argument("-q", "--quantize", default="auto",
                        choices=["auto", "f16", "w8f16", "w4f16", "bf16", "w8bf16", "w4bf16"],
                        help="Quantization mode (default: auto, inferred from model config)")
    parser.add_argument("-c", "--chip", default="bm1684x",
                        choices=["bm1684x", "bm1688", "cv186x", "bm1690", "bm1684x2"],
                        help="Chip type (default: bm1684x)")
    parser.add_argument("-v", "--vector_tops", type=float, default=None,
                        help="Vector compute power in TOPS (default: tops/8)")
    parser.add_argument("-r", "--uarch_rate", type=float, default=0.8,
                        help="uArch Rate ratio (default: 0.8)")
    parser.add_argument("-u", "--bw_util", type=float, default=0.8,
                        help="Bandwidth utilization ratio (default: 0.8)")
    parser.add_argument("-p", "--parallelism", type=float, default=0.5,
                        help="Parallelism ratio for Est.Time (default: 0.5)")
    parser.add_argument('--max_input_length', type=int, default=0,
                        help='max input length for prefill, default 0 means the same as seq_length')
    parser.add_argument('--max_pixels', type=str, default="",
                        help='max input pixels for vision models, default "" means no vision input')
    parser.add_argument("-o", "--out_dir", required=True,
                        help="Output directory path (default: <out_dir>/<out_dir>_analysis.xlsx)")
    args = parser.parse_args()
    # yapf: enable
    from llm.transformers_compat import load_auto_config
    config = load_auto_config(args.model_path, trust_remote_code=True)
    if hasattr(config, "text_config"):
        llm_config = config.text_config
    else:
        llm_config = config
    # ------------------------------------------------------------------
    # TODO: Add llm_converter.py call here to generate MLIRs if needed
    # e.g. convert_model(args.model_path, args.model_dir, ...)
    # ------------------------------------------------------------------
    if args.max_pixels:
        max_pixels = args.max_pixels
    elif config.model_type in ["qwen2_5_vl", "qwen2_vl"]:
        max_pixels = "672,896"
    else:
        max_pixels = "768,768"
    cmds = [
        "llm_convert.py", f"-m {args.model_path}", f"-s {args.seq_length}", f"-q {args.quantize}",
        f"-c {args.chip}", f"--out_dir {args.out_dir}", "--only_mlir", "--debug",
        f"--max_pixels {max_pixels}"
    ]
    if args.max_input_length > 0:
        cmds.append(f"--max_input_length {args.max_input_length}")
    print("\nRunning LLM conversion to generate MLIR files...")
    cmd = " ".join(cmds)
    print(f"Command: {cmd}")
    mlir_dir = os.path.join(args.out_dir, "tmp_mlir_analyse")
    try:
        subprocess.check_call(cmd, shell=True)
    except subprocess.CalledProcessError as e:
        # Codegen (bmodel) may fail for some quant modes (e.g. Fp8MatMul on a
        # chip whose backend lacks codegen), but the *_tpu.mlir files needed
        # for analysis are produced earlier in the pipeline. Proceed if they
        # exist; only bail out when there is nothing to analyse.
        if glob.glob(os.path.join(mlir_dir, "*", f"*_{args.chip}_*_tpu.mlir")):
            print(
                f"Warning: LLM conversion exited with {e.returncode} (likely bmodel "
                f"codegen failure), but *_tpu.mlir files are present - proceeding "
                f"with analysis.",
                file=sys.stderr)
        else:
            print(f"Error during LLM conversion: {e}", file=sys.stderr)
            sys.exit(1)

    model_type = getattr(llm_config, "model_type", "")
    # Extract model config info
    model_config = {}
    for key in [
            "num_hidden_layers", "hidden_size", "num_attention_heads", "num_key_value_heads",
            "intermediate_size", "vocab_size", "head_dim"
    ]:
        val = getattr(llm_config, key, None)
        if val is not None:
            model_config[key] = val
    archs = getattr(config, "architectures", None)
    if archs:
        model_config["architectures"] = archs[0] if len(archs) == 1 else ", ".join(archs)
    # Use num_hidden_layers to correct block counts if available
    num_layers = model_config.get("num_hidden_layers")
    # Step 1: Discover modules
    print(f"Scanning: {mlir_dir}")
    modules = discover_modules(mlir_dir, llm_config.num_hidden_layers, model_type, args.chip)
    if not modules:
        print("No MLIR files found.", file=sys.stderr)
        sys.exit(1)
    # Correct block counts from config if available. discover_modules already
    # applies the right counts (qwen3_5 uses 3:1 linear:full via integer
    # division, matching the [linear,linear,linear,full] layer-type pattern);
    # only the non-qwen3_5 path needs the num_layers override here.
    if num_layers and model_type not in QWEN3_5_MODEL_TYPES:
        modules = [(n, p, num_layers if n in ("block", "block_cache") and c > 1 else c)
                   for n, p, c in modules]
    print(f"Found {len(modules)} module(s): "
          f"{', '.join(f'{n}(x{c})' if c > 1 else n for n, _, c in modules)}")

    # Step 2: Analyse each module
    modules_data = []
    for name, path, count in modules:
        print(f"  Analysing: {name} ({os.path.basename(path)})")
        rows_data, totals = analyse_module(path)
        modules_data.append((name, count, rows_data, totals))

    # Step 3: Export Excel
    cmdline = "python " + " ".join(sys.argv)
    fp16_tops = args.tops
    int8_tops = args.int8_tops if args.int8_tops is not None else fp16_tops * 2.0
    fp8_tops = args.fp8_tops if args.fp8_tops is not None else fp16_tops * 2.0
    export_llm_excel(modules_data, fp16_tops, args.bandwidth, args.out_dir, args.model_path,
                     int8_tops, fp8_tops, args.vector_tops, args.uarch_rate, args.bw_util,
                     args.parallelism, model_config, args.seq_length, max_pixels, cmdline)


if __name__ == "__main__":
    main()
